"""
Asset classifier pipeline: coordinates or street address -> NAIP aerial chip
-> Claude vision classification.

Flow:
  1. Read sites from assets.csv (columns: id; plus lat+lon OR address; optional
     label, input_confidence: high | medium | low for source trust). Addresses
     are geocoded to lat/lon via the free US Census Geocoder (CONUS) with
     OpenStreetMap Nominatim as fallback.
  2. Query Microsoft Planetary Computer STAC API for the newest NAIP scene at each point
  3. Windowed-read a chip around the point from the Cloud-Optimized GeoTIFF (no full download)
  4. Optional: if NEARMAP_API_KEY is set, also pull a high-res Nearmap vertical
     and 45-degree oblique panoramas (N/E/S/W) via the Tile API
  5. Send all views to Claude: classify the site (tower vs rooftop), locate the
     asset with a bounding box, and assess visible cellular equipment
  6. If still unidentified (rural vert-only), widen the Nearmap AOI to match NAIP
  7. If still unidentified, run a two-stage zoom: scout candidate regions, crop
     and magnify them, then re-classify on the zoomed views
  8. Convert the detection box on the georeferenced NAIP chip to asset lat/lon
  9. Write results.csv, chips for spot-checking, and an executive summary markdown

Setup:
  pip install -r requirements.txt
  Get an API key at https://console.anthropic.com/ then:
  export ANTHROPIC_API_KEY=sk-ant-...

Notes:
  - NAIP covers the continental US only (~0.6-1m resolution, public domain).
    For other regions, swap the STAC collection (e.g. state orthoimagery, OpenAerialMap).
  - Chip size: 250m at 0.6m GSD = ~417px square. Good balance of context vs detail.
"""

import base64
import io
import json
import math
import os
import time
import argparse
import shutil
from pathlib import Path

import numpy as np
import pandas as pd
import rasterio
import requests
from rasterio.windows import from_bounds
from pyproj import Transformer
from pystac_client import Client
import planetary_computer
from PIL import Image
import anthropic
from anthropic import Anthropic
from google import genai
from google.genai import types as genai_types
from dotenv import load_dotenv

try:
    from tqdm import tqdm
except ImportError:
    def tqdm(iterable, **kwargs):
        return iterable

load_dotenv()  # picks up ANTHROPIC_API_KEY / GEMINI_API_KEY from .env if present


def _env_flag(name: str, default: str = "0") -> bool:
    return os.environ.get(name, default).strip().lower() in ("1", "true", "yes")


NEARMAP_TIERED = _env_flag("NEARMAP_TIERED")
BIFURCATED_AI = _env_flag("BIFURCATED_AI")
NAIP_ONLY = _env_flag("NAIP_ONLY")
TIER_CONF_HIGH = float(os.environ.get("TIER_CONF_HIGH", "0.75"))
TIER_CONF_MEDIUM = float(os.environ.get("TIER_CONF_MEDIUM", "0.6"))
GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")
CLAUDE_ESCALATION_MODEL = os.environ.get(
    "CLAUDE_ESCALATION_MODEL", "claude-sonnet-4-6")
OBLIQUE_VIEWS = ["North", "East", "South", "West"]

# ----------------------------- configuration --------------------------------

STAC_URL = "https://planetarycomputer.microsoft.com/api/stac/v1"
COLLECTION = "naip"
CHIP_SIZE_M = 250          # side length of the extracted chip, in meters
# Primary model first; hops to the next on persistent rate limits or 404.
# claude-sonnet-4-20250514 was retired 2026-06-15; use current IDs from:
# https://docs.anthropic.com/en/docs/about-claude/models/overview
_default_models = "claude-sonnet-4-6,claude-haiku-4-5-20251001"
MODELS = [
    m.strip() for m in os.environ.get("CLAUDE_MODELS", _default_models).split(",")
    if m.strip()
]
_model_idx = 0
API_DELAY_S = float(os.environ.get("CLAUDE_DELAY_S", "12"))
INPUT_CSV = "assets.csv"    # columns: id; lat+lon OR address; optional: label, input_confidence
INPUT_CONFIDENCE_LEVELS = ("high", "medium", "low")
OUTPUT_CSV = "results.csv"
CHIP_DIR = Path("chips")
RUNS_DIR = Path("runs")
RUN_DIR = Path(".")

# Geocoding: free US Census API first (good for CONUS rooftop addresses), then
# OpenStreetMap Nominatim. Set GEOCODER=nominatim to skip Census.
GEOCODER = os.environ.get("GEOCODER", "auto").strip().lower()  # auto | census | nominatim
GEOCODER_USER_AGENT = os.environ.get(
    "GEOCODER_USER_AGENT", "site-classifier/1.0 (cell-site imagery pipeline)")
CENSUS_GEOCODE_URL = (
    "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress")
CENSUS_BENCHMARK = "Public_AR_Current"
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
GEOCODE_DELAY_S = 1.1      # Nominatim usage policy: max 1 request/second

# Optional Nearmap integration (Tile API).
# When NEARMAP_API_KEY is set, each asset also gets a high-res top-down view
# plus 45-degree oblique panoramas from the four compass directions - obliques
# show the vertical sides of structures, which is what makes rooftop antennas
# and towers actually visible. Without the key the pipeline runs NAIP-only.
# The Tile API bills against the subscription's monthly GB allowance; the
# Transactional Content API was tried first but needs a separate credits
# add-on (the coverage/v2/tx call returns 403 on this subscription).
NEARMAP_API_KEY = os.environ.get("NEARMAP_API_KEY")
NEARMAP_TILE_URL = "https://api.nearmap.com/tiles/v3/{content}/{z}/{x}/{y}.jpg"
NEARMAP_COVERAGE_POINT_URL = "https://api.nearmap.com/coverage/v2/point/{lon},{lat}"
NEARMAP_VIEWS = ["Vert", "North", "East", "South", "West"]
NEARMAP_CHIP_M = 100       # side length of the Nearmap AOI, in meters
# Wide-AOI fallback: rural sites often have vert-only Nearmap coverage and the
# recorded coordinates can put the real asset outside the narrow AOI. When the
# first pass can't identify a site and no obliques were available, the Nearmap
# fetch is retried at this size (matching the NAIP chip) and re-classified.
NEARMAP_FALLBACK_CHIP_M = 250
NEARMAP_VERT_ZOOM = 21     # ~6 cm/px Web Mercator ground resolution
NEARMAP_OBLIQUE_ZOOM = 20  # panorama max zoom, ~11 cm/px
NEARMAP_MAX_PX = 2048      # downscale stitched views before saving/sending

# Two-stage zoom: after primary + wide-AOI passes still return other/unclear,
# scout suspicious regions on the best top-down image, magnify them, and
# re-classify. Critical for rural sites where towers are tiny in wide chips.
ZOOM_GRID = 3              # 3x3 grid fallback when scout finds nothing
ZOOM_MAX_CANDIDATES = 6    # max zoom crops sent to stage-2 classifier
ZOOM_OUTPUT_PX = 1024      # magnified crop size in pixels
ZOOM_MIN_FRAC = 0.10       # minimum crop side as fraction of source image
ZOOM_PAD_FRAC = 0.15       # padding around each candidate box
EXECUTIVE_SUMMARY_MD = "EXECUTIVE_SUMMARY.md"

CLASSIFICATION_PROMPT = """\
You are analyzing aerial imagery of one location where a cellular-infrastructure \
asset is expected. One or more views are provided, each preceded by a text label:
- "NAIP top-down": wide straight-down chip (~250 m across, ~1 m resolution). \
The recorded coordinates can be off by tens of meters, so the asset may appear \
ANYWHERE in this chip, not just at the center.
- "Nearmap top-down": recent high-resolution (~7 cm) straight-down view of the \
same location, usually covering a smaller area than the NAIP chip.
- "Nearmap oblique (North/East/South/West)": 45-degree angled views of the same \
location. These reveal the vertical sides of structures - towers, masts, and \
rooftop antennas that are nearly invisible from straight above stand out \
clearly here. Weight them heavily in every task.

Definitions:
- TOWER SITE: a ground-based, purpose-built vertical structure carrying \
antennas - monopole, lattice/self-support tower, or guyed mast. Top-down cues: \
tiny footprint, long thin linear shadow, lattice cross-pattern, guy wires, \
small cleared/fenced compound with equipment cabinets. Oblique cues: a tall \
thin structure rising far above its surroundings.
- ROOFTOP SITE: a building whose roof hosts the cellular equipment. Cues: \
panel antennas / sector frames at roof corners or edges (often 3 sectors), \
triangular/rectangular antenna mounts, microwave backhaul dishes, equipment \
cabinets with cable trays, short masts on the parapet.
- STEALTH / BUILDING-TOWER SITE: a structure that looks like a building but \
has a tall narrow tower section - church steeple, clock tower, faux-building \
monopole, or a tower segment rising from one corner of a larger footprint. \
From above: a compact building with an unusually tall shadow from one corner \
or a square tower block on the roofline; antennas may sit on the tower cap.

Perform three tasks:

TASK 1 - site_type. Search the ENTIRE extent of EVERY view - edges and corners \
included, never just the center - and classify the site:
- "tower": a tower (as defined above) is visible anywhere in the imagery. A \
tower outranks nearby buildings - never call a site "rooftop" merely because a \
large building is more prominent than a thin mast.
- "rooftop": no tower present, and a building roof hosts (or most plausibly \
hosts) the equipment.
- "other": neither applies (water tank, silo, bare field, etc.) - describe it.
- "unclear": image quality or ambiguity prevents a confident call.
Set site_confidence to at most 0.6 unless two or more independent cues or \
views corroborate the call.

TASK 2 - locate the asset. Identify the exact asset (the tower structure and \
its compound, or the host building's roof) and report:
- asset_box_2d: [ymin, xmin, ymax, xmax], integers in 0-1000 normalized image \
coordinates, drawn TIGHTLY around the asset ON THE FIRST IMAGE provided. If \
the asset is not visible in the first image, box it on the view where it is \
clearest instead.
- asset_view: the exact label of the view the box was drawn on (e.g. "NAIP \
top-down"). Set both fields to null only if no asset can be located at all.

TASK 3 - cell_equipment: is cellular equipment visible on the located asset \
(antennas, sector frames, dishes, cabinets, cable trays)?
- true: visible evidence; false: none visible; null: cannot assess (resolution \
or viewing angle insufficient).
On rooftops, equipment is often missed when it sits in **building shadow**, \
along shaded parapets, or reads similarly to HVAC. In oblique views, inspect \
sunlit AND shaded roof edges, corners, and mechanical zones before calling \
false.

Field meanings: site_evidence and cell_equipment_evidence are one short \
sentence each, citing the specific views and cues used.
"""

INPUT_CONFIDENCE_PROMPTS = {
    "high": (
        "\n\nSOURCE TRUST: HIGH. The coordinate comes from a trusted source that "
        "expects an active cellular asset at this location. Prioritize finding "
        "the host structure and any cellular equipment. Do not call "
        "cell_equipment false from a quick scan - inspect roof edges, parapets, "
        "and shaded areas on oblique views. Prefer null over false when "
        "imagery is ambiguous."
    ),
    "medium": (
        "\n\nSOURCE TRUST: MEDIUM. The coordinate likely points to a cellular "
        "site but may be approximate. Weight oblique views when assessing "
        "rooftop equipment in shadow."
    ),
    "low": (
        "\n\nSOURCE TRUST: LOW. The coordinate is exploratory; apply normal "
        "evidence standards."
    ),
}

EQUIPMENT_RECHECK_PROMPT = """\
You are re-checking ONLY for visible cellular equipment at a trusted site. \
The first pass called cell_equipment false, but the data source expects gear \
here.

Re-examine every view - especially Nearmap obliques and shaded roof areas:
- Thin rectangular panel antennas on parapets or short masts
- Sector frames at roof corners (often three sectors)
- Microwave dishes, RRUs, cable trays, equipment cabinets
- Gear hidden in building shadow or mistaken for HVAC

Return the same JSON schema. If any plausible cellular equipment is visible, \
set cell_equipment true and explain which view and shaded/sunlit area shows it.
"""

# Enforced via Claude tool input_schema so every reply parses into this shape.
RESPONSE_SCHEMA = {
    "type": "object",
    "properties": {
        "site_type": {
            "type": "string",
            "enum": ["tower", "rooftop", "other", "unclear"],
        },
        "site_confidence": {"type": "number"},
        "site_evidence": {"type": "string"},
        "asset_box_2d": {
            "type": "array",
            "items": {"type": "integer"},
        },
        "asset_view": {"type": "string"},
        "cell_equipment": {"type": "boolean"},
        "cell_equipment_confidence": {"type": "number"},
        "cell_equipment_evidence": {"type": "string"},
    },
    "required": ["site_type", "site_confidence", "site_evidence"],
}

# Gemini SDK uses a distinct schema dialect; kept in sync with RESPONSE_SCHEMA.
GEMINI_RESPONSE_SCHEMA = {
    "type": "OBJECT",
    "properties": {
        "site_type": {
            "type": "STRING",
            "enum": ["tower", "rooftop", "other", "unclear"],
        },
        "site_confidence": {"type": "NUMBER"},
        "site_evidence": {"type": "STRING"},
        "asset_box_2d": {
            "type": "ARRAY",
            "items": {"type": "INTEGER"},
            "nullable": True,
        },
        "asset_view": {"type": "STRING", "nullable": True},
        "cell_equipment": {"type": "BOOLEAN", "nullable": True},
        "cell_equipment_confidence": {"type": "NUMBER"},
        "cell_equipment_evidence": {"type": "STRING"},
    },
    "required": ["site_type", "site_confidence", "site_evidence"],
}

SCAN_PROMPT = """\
You are reviewing a single top-down aerial image where a cellular tower or \
rooftop site is expected, but a first-pass classifier could not identify it. \
The asset may be anywhere in the frame and is often subtle: a small lattice \
mast, monopole shadow, fenced compound, rooftop antenna cluster, or a stealth \
tower disguised as a building with a tall tower section (steeple, clock tower, \
faux-building cell site). Check the area just below image center especially \
when coordinates are approximate.

Search the ENTIRE image - especially edges and corners - and return up to four \
candidate regions that could plausibly be a tower site or rooftop cellular host. \
Prioritize: tiny footprints with long shadows, lattice cross-patterns, fenced \
pads with equipment cabinets, or building roofs with sector-frame mounts.

Return ONLY JSON with a "candidates" array. Each entry needs:
- box_2d: [ymin, xmin, ymax, xmax] in 0-1000 normalized coordinates
- reason: one short phrase citing the visual cue

If nothing looks plausible, return an empty candidates array.
"""

SCAN_SCHEMA = {
    "type": "object",
    "properties": {
        "candidates": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "box_2d": {
                        "type": "array",
                        "items": {"type": "integer"},
                    },
                    "reason": {"type": "string"},
                },
                "required": ["box_2d", "reason"],
            },
        },
    },
    "required": ["candidates"],
}

GEMINI_SCAN_SCHEMA = {
    "type": "OBJECT",
    "properties": {
        "candidates": {
            "type": "ARRAY",
            "items": {
                "type": "OBJECT",
                "properties": {
                    "box_2d": {
                        "type": "ARRAY",
                        "items": {"type": "INTEGER"},
                    },
                    "reason": {"type": "STRING"},
                },
                "required": ["box_2d", "reason"],
            },
        },
    },
    "required": ["candidates"],
}

ZOOM_CLASSIFICATION_PROMPT = """\
You are performing a SECOND-PASS review on magnified zoom crops from a site \
that was not identified in wide imagery. A cellular tower or rooftop site is \
still expected at this location.

One or more "Zoom crop" views are provided - each is a magnified section of a \
top-down image. Also included may be the original wide "NAIP top-down" or \
"Nearmap top-down" view for context.

Use the zoom crops as primary evidence. A tower site often appears as a \
lattice mast, monopole, guyed structure, or small fenced compound with \
equipment. A stealth site may be a building with a tall tower block or steeple \
on one corner and a long shadow from that tower section. A rooftop site shows \
panel antennas, sector frames, or dishes on a building roof.

Perform the same three tasks as the primary classifier:
1. site_type: tower | rooftop | other | unclear
2. asset_box_2d + asset_view on the view where the asset is clearest
3. cell_equipment: true | false | null

Set site_confidence to at most 0.7 unless zoom crops show unambiguous equipment.
"""


def normalize_input_confidence(value) -> str:
    """Return high | medium | low. Missing or invalid values default to medium."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "medium"
    level = str(value).strip().lower()
    return level if level in INPUT_CONFIDENCE_LEVELS else "medium"


def normalize_confidence(value, default: float | None = None) -> float | None:
    """Clamp model confidence to 0-1. Values > 1 are treated as wrong-scale output."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return default
    try:
        v = float(value)
    except (TypeError, ValueError):
        return default
    if v > 1.0:
        if v <= 10.0:
            v /= 10.0
        elif v <= 100.0:
            v /= 100.0
        else:
            v = 1.0
    return max(0.0, min(1.0, v))


def normalize_model_result(res: dict) -> dict:
    """Fix site_confidence and cell_equipment_confidence after a model JSON reply."""
    if "site_confidence" in res:
        norm = normalize_confidence(res.get("site_confidence"))
        if norm is not None:
            res["site_confidence"] = norm
    if "cell_equipment_confidence" in res:
        norm = normalize_confidence(res.get("cell_equipment_confidence"))
        if norm is not None:
            res["cell_equipment_confidence"] = norm
    return res


def build_classification_prompt(row) -> str:
    """Assemble the full classification prompt from base + label + source trust."""
    prompt = CLASSIFICATION_PROMPT
    label_hint = str(row.get("label", "")).strip().lower()
    if label_hint == "stealth":
        prompt += (
            "\n\nNOTE: This site is tagged STEALTH. Expect a tower "
            "disguised as or integrated into a building - steeple, "
            "clock tower, faux facade, or tower segment on a corner. "
            "Do not dismiss a tall narrow shadow or tower block as "
            "merely architectural unless clearly non-telecom."
        )
    prompt += INPUT_CONFIDENCE_PROMPTS[normalize_input_confidence(
        row.get("input_confidence"))]
    return prompt


def maybe_recheck_equipment(provider: str, clients: dict, res: dict, views: list,
                            input_confidence: str) -> dict:
    """Second pass when a trusted source expects gear but the model said false."""
    if input_confidence not in ("high", "medium"):
        return res
    if res.get("cell_equipment") is not False:
        return res
    if len(views) < 2:
        return res
    print("  equipment recheck (trusted source, obliques/shadow pass)")
    recheck = classify_site(
        provider, clients, views, prompt=EQUIPMENT_RECHECK_PROMPT)
    if recheck.get("cell_equipment") is True:
        res["cell_equipment"] = True
        res["cell_equipment_confidence"] = recheck.get(
            "cell_equipment_confidence", res.get("cell_equipment_confidence"))
        res["cell_equipment_evidence"] = recheck.get(
            "cell_equipment_evidence", res.get("cell_equipment_evidence"))
        if recheck.get("site_type") in ("tower", "rooftop"):
            res["site_type"] = recheck["site_type"]
            res["site_confidence"] = recheck.get(
                "site_confidence", res.get("site_confidence"))
            res["site_evidence"] = recheck.get(
                "site_evidence", res.get("site_evidence"))
        normalize_model_result(res)
    return res

# ----------------------------- geocoding ------------------------------------

_last_geocode_at = 0.0


def _has_coordinates(row) -> bool:
    lat, lon = row.get("lat"), row.get("lon")
    if lat is None or lon is None:
        return False
    if isinstance(lat, float) and pd.isna(lat):
        return False
    if isinstance(lon, float) and pd.isna(lon):
        return False
    if str(lat).strip() == "" or str(lon).strip() == "":
        return False
    return True


def _clean_address(row) -> str | None:
    address = row.get("address")
    if address is None or (isinstance(address, float) and pd.isna(address)):
        return None
    text = str(address).strip()
    return text or None


def _throttle_nominatim():
    global _last_geocode_at
    elapsed = time.time() - _last_geocode_at
    if elapsed < GEOCODE_DELAY_S:
        time.sleep(GEOCODE_DELAY_S - elapsed)
    _last_geocode_at = time.time()


def geocode_census(address: str) -> dict | None:
    """US Census Bureau oneline geocoder - free, no API key, CONUS-focused."""
    resp = requests.get(
        CENSUS_GEOCODE_URL,
        params={
            "address": address,
            "benchmark": CENSUS_BENCHMARK,
            "format": "json",
        },
        timeout=30,
    )
    resp.raise_for_status()
    matches = resp.json().get("result", {}).get("addressMatches", [])
    if not matches:
        return None
    match = matches[0]
    coords = match["coordinates"]
    return {
        "lat": float(coords["y"]),
        "lon": float(coords["x"]),
        "geocode_source": "census",
        "geocode_matched_address": match.get("matchedAddress"),
        "geocode_quality": "census_match",
    }


def geocode_nominatim(address: str) -> dict | None:
    """OpenStreetMap Nominatim - free, worldwide, 1 req/sec usage policy."""
    _throttle_nominatim()
    resp = requests.get(
        NOMINATIM_URL,
        params={"q": address, "format": "json", "limit": 1},
        headers={"User-Agent": GEOCODER_USER_AGENT},
        timeout=30,
    )
    resp.raise_for_status()
    results = resp.json()
    if not results:
        return None
    hit = results[0]
    return {
        "lat": float(hit["lat"]),
        "lon": float(hit["lon"]),
        "geocode_source": "nominatim",
        "geocode_matched_address": hit.get("display_name"),
        "geocode_quality": hit.get("type") or hit.get("class"),
    }


def geocode_address(address: str) -> dict:
    """Resolve a street address to lat/lon. Raises ValueError if no match."""
    errors = []
    if GEOCODER in ("auto", "census"):
        try:
            result = geocode_census(address)
            if result:
                return result
            errors.append("census: no match")
        except Exception as e:
            errors.append(f"census: {e}")

    if GEOCODER in ("auto", "nominatim"):
        try:
            result = geocode_nominatim(address)
            if result:
                return result
            errors.append("nominatim: no match")
        except Exception as e:
            errors.append(f"nominatim: {e}")

    raise ValueError(
        f"could not geocode address ({'; '.join(errors)}): {address}")


def resolve_row_coordinates(row) -> tuple[float, float, dict]:
    """Return (lat, lon, geocode_metadata). metadata is empty when coords given."""
    if _has_coordinates(row):
        return float(row["lat"]), float(row["lon"]), {}

    address = _clean_address(row)
    if not address:
        raise ValueError(
            "each row needs lat+lon or a non-empty address column")

    geo = geocode_address(address)
    meta = {k: v for k, v in geo.items() if k not in ("lat", "lon")}
    meta["input_address"] = address
    return geo["lat"], geo["lon"], meta


def validate_input_csv(df: pd.DataFrame):
    """Ensure each row has id and either coordinates or an address."""
    if "id" not in df.columns:
        raise SystemExit(f"{INPUT_CSV} is missing required column: id")
    has_coords = "lat" in df.columns and "lon" in df.columns
    has_address = "address" in df.columns
    if not has_coords and not has_address:
        raise SystemExit(
            f"{INPUT_CSV} needs lat+lon columns, an address column, or both")

    missing = []
    for _, row in df.iterrows():
        if _has_coordinates(row):
            continue
        if _clean_address(row):
            continue
        missing.append(row["id"])
    if missing:
        raise SystemExit(
            f"{INPUT_CSV}: these rows have no lat/lon and no address: "
            f"{missing[:5]}{'...' if len(missing) > 5 else ''}")

# ----------------------------- imagery stage --------------------------------

_catalog = None

def get_catalog():
    global _catalog
    if _catalog is None:
        _catalog = Client.open(STAC_URL, modifier=planetary_computer.sign_inplace)
    return _catalog


def fetch_chip(lat: float, lon: float, chip_m: float = CHIP_SIZE_M):
    """Return (PIL.Image, acquisition_date, geo) for the newest NAIP scene at a
    point, or (None, None, None) if no imagery covers the location. `geo` holds
    the chip's CRS and projected bounds so a detection box drawn on the image
    can be converted back to real-world coordinates."""
    search = get_catalog().search(
        collections=[COLLECTION],
        intersects={"type": "Point", "coordinates": [lon, lat]},
    )
    items = sorted(search.items(), key=lambda i: i.datetime, reverse=True)
    if not items:
        return None, None, None

    item = items[0]
    href = item.assets["image"].href

    with rasterio.open(href) as src:
        # NAIP rasters are in UTM; project the WGS84 point into the raster CRS
        transformer = Transformer.from_crs("EPSG:4326", src.crs, always_xy=True)
        x, y = transformer.transform(lon, lat)
        half = chip_m / 2.0
        window = from_bounds(x - half, y - half, x + half, y + half, src.transform)
        # Read RGB bands only; boundless handles points near scene edges
        data = src.read([1, 2, 3], window=window, boundless=True, fill_value=0)
        geo = {"crs": str(src.crs),
               "x_min": x - half, "x_max": x + half,
               "y_min": y - half, "y_max": y + half}

    img = Image.fromarray(np.transpose(data, (1, 2, 0)).astype(np.uint8))
    return img, item.datetime.date().isoformat(), geo


def box_to_latlon(geo: dict, box) -> tuple[float, float, float] | None:
    """Convert a [ymin, xmin, ymax, xmax] box in 0-1000 normalized image
    coordinates on the NAIP chip into (lat, lon, offset_m), where offset_m is
    the distance from the box center to the chip center (the input coordinate).
    Returns None if the box is malformed."""
    try:
        ymin, xmin, ymax, xmax = (float(v) for v in box[:4])
    except (TypeError, ValueError):
        return None
    if not (0 <= ymin <= ymax <= 1000 and 0 <= xmin <= xmax <= 1000):
        return None
    # Normalized box center -> projected coordinates (y axis is flipped:
    # image row 0 is the chip's northern edge / max projected y)
    cx_n = (xmin + xmax) / 2000.0
    cy_n = (ymin + ymax) / 2000.0
    x = geo["x_min"] + cx_n * (geo["x_max"] - geo["x_min"])
    y = geo["y_max"] - cy_n * (geo["y_max"] - geo["y_min"])
    to_wgs84 = Transformer.from_crs(geo["crs"], "EPSG:4326", always_xy=True)
    lon, lat = to_wgs84.transform(x, y)
    center_x = (geo["x_min"] + geo["x_max"]) / 2.0
    center_y = (geo["y_min"] + geo["y_max"]) / 2.0
    offset_m = math.hypot(x - center_x, y - center_y)
    return lat, lon, offset_m


_nearmap_session = requests.Session()


def _nearmap_get(url: str) -> requests.Response:
    """GET with header auth (keeps the API key out of logged URLs) and a
    short retry on rate-limit/transient errors."""
    for attempt in range(3):
        resp = _nearmap_session.get(
            url, headers={"Authorization": f"Apikey {NEARMAP_API_KEY}"},
            timeout=60)
        if resp.status_code in (429, 502, 503) and attempt < 2:
            time.sleep(2 * (attempt + 1))
            continue
        return resp
    return resp


def _tile_range(lat: float, lon: float, half_m: float, zoom: int):
    """Slippy-tile x/y index range covering a half_m-radius box at a zoom."""
    dlat = half_m / 111_320.0
    dlon = half_m / (111_320.0 * math.cos(math.radians(lat)))
    n = 2 ** zoom

    def tile_xy(la, lo):
        x = (lo + 180.0) / 360.0 * n
        y = (1.0 - math.asinh(math.tan(math.radians(la))) / math.pi) / 2.0 * n
        return x, y

    x_west, y_north = tile_xy(lat + dlat, lon - dlon)
    x_east, y_south = tile_xy(lat - dlat, lon + dlon)
    return int(x_west), int(x_east), int(y_north), int(y_south)


def fetch_nearmap_views(lat: float, lon: float, chip_m: float = NEARMAP_CHIP_M,
                        views: list[str] | None = None):
    """Fetch Nearmap content for a point via the Tile API: high-res vertical
    plus 45-degree oblique panoramas (N/E/S/W), stitched from XYZ tiles.

    Returns ({view_name: PIL.Image}, capture_date). Empty dict when the key is
    not set or the location has no Nearmap coverage.

    Optional `views` limits which orientations to fetch (e.g. ["Vert"] or
    OBLIQUE_VIEWS). Defaults to all NEARMAP_VIEWS when omitted.
    """
    if not NEARMAP_API_KEY:
        return {}, None

    fetch_views = views if views is not None else NEARMAP_VIEWS
    result = {}
    for view in fetch_views:
        zoom = NEARMAP_VERT_ZOOM if view == "Vert" else NEARMAP_OBLIQUE_ZOOM
        x0, x1, y0, y1 = _tile_range(lat, lon, chip_m / 2.0, zoom)
        cols, rows = x1 - x0 + 1, y1 - y0 + 1

        # Canvas dimensions follow the tile orientation: East/West mosaics
        # have the slippy x axis running vertically
        if view in ("East", "West"):
            canvas = Image.new("RGB", (rows * 256, cols * 256))
        else:
            canvas = Image.new("RGB", (cols * 256, rows * 256))

        got_any = False
        for ty in range(y0, y1 + 1):
            for tx in range(x0, x1 + 1):
                resp = _nearmap_get(NEARMAP_TILE_URL.format(
                    content=view, z=zoom, x=tx, y=ty))
                if resp.status_code == 404:   # no coverage for this tile/view
                    continue
                resp.raise_for_status()
                tile = Image.open(io.BytesIO(resp.content)).convert("RGB")
                got_any = True
                if view in ("Vert", "North"):     # north-up
                    pos = ((tx - x0) * 256, (ty - y0) * 256)
                elif view == "South":             # south-up: both axes flip
                    pos = ((x1 - tx) * 256, (y1 - ty) * 256)
                elif view == "East":              # east-up: up = +x, right = +y
                    pos = ((ty - y0) * 256, (x1 - tx) * 256)
                else:                             # west-up: up = -x, right = -y
                    pos = ((y1 - ty) * 256, (tx - x0) * 256)
                canvas.paste(tile, pos)

        if not got_any:
            continue
        if view != "Vert":
            # Compensate the 45-degree foreshortening (256 -> 192 height)
            canvas = canvas.resize(
                (canvas.width, max(1, int(canvas.height * 0.75))))
        canvas.thumbnail((NEARMAP_MAX_PX, NEARMAP_MAX_PX))
        result[view] = canvas

    if not result:
        return {}, None

    # Capture date metadata via the standard (non-transactional) coverage API;
    # purely informational, so failures are ignored
    capture_date = None
    try:
        resp = _nearmap_get(
            NEARMAP_COVERAGE_POINT_URL.format(lon=lon, lat=lat) + "?limit=1")
        if resp.ok:
            surveys = resp.json().get("surveys") or []
            if surveys:
                capture_date = surveys[0].get("captureDate")
    except Exception:
        pass
    return result, capture_date

# --------------------------- classification stage ---------------------------

def _image_block(img: Image.Image) -> dict:
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=90)
    data = base64.standard_b64encode(buf.getvalue()).decode("ascii")
    return {
        "type": "image",
        "source": {
            "type": "base64",
            "media_type": "image/jpeg",
            "data": data,
        },
    }


def _parse_json_fallback(text: str, default: dict) -> dict:
    text = (text or "").strip()
    start, end = text.find("{"), text.rfind("}")
    if start != -1 and end > start:
        text = text[start:end + 1]
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return {**default, "site_evidence": f"unparseable model reply: {text[:200]}"}


def _extract_tool_result(resp, tool_name: str, default: dict) -> dict:
    for block in resp.content:
        if block.type == "tool_use" and block.name == tool_name:
            if isinstance(block.input, dict):
                return dict(block.input)
    text_parts = [block.text for block in resp.content if block.type == "text"]
    return _parse_json_fallback("\n".join(text_parts), default)


def _call_claude_json(client: Anthropic, content: list, schema: dict,
                      tool_name: str, retries: int = 3,
                      model: str | None = None) -> dict:
    """Shared Claude vision call with tool-based JSON and retry logic."""
    global _model_idx
    attempt = 0
    default = {"site_type": "unclear", "site_confidence": 0.0}
    while True:
        use_model = model or MODELS[_model_idx]
        try:
            resp = client.messages.create(
                model=use_model,
                max_tokens=1000,
                tools=[{
                    "name": tool_name,
                    "description": "Return structured analysis as JSON.",
                    "input_schema": schema,
                }],
                tool_choice={"type": "tool", "name": tool_name},
                messages=[{"role": "user", "content": content}],
            )
            res = _extract_tool_result(resp, tool_name, default)
            normalize_model_result(res)
            res["model"] = use_model
            return res
        except anthropic.RateLimitError:
            if model:
                raise
            if attempt < retries:
                attempt += 1
                wait = 15 * attempt
                print(f"  rate limit on {use_model}, retrying in {wait}s "
                      f"({attempt}/{retries})...")
                time.sleep(wait)
                continue
            if _model_idx + 1 < len(MODELS):
                print(f"  {use_model} rate limited -> hopping to {MODELS[_model_idx + 1]}")
                _model_idx += 1
                attempt = 0
                continue
            raise
        except anthropic.APIStatusError as e:
            if model:
                raise
            if e.status_code == 404 and _model_idx + 1 < len(MODELS):
                print(f"  {use_model} not found (404) -> hopping to {MODELS[_model_idx + 1]}")
                _model_idx += 1
                attempt = 0
                continue
            if e.status_code == 404:
                raise SystemExit(
                    f"\nClaude model '{use_model}' returned 404 (not found). "
                    f"Tried: {', '.join(MODELS)}\n"
                    "Set CLAUDE_MODELS to valid IDs, e.g. "
                    "claude-sonnet-4-6,claude-haiku-4-5-20251001\n"
                    "See https://docs.anthropic.com/en/docs/about-claude/models/overview"
                ) from e
            if e.status_code in (429, 529, 503, 500) and attempt < retries:
                attempt += 1
                wait = 15 * attempt
                print(f"  transient {e.status_code}, retrying in {wait}s "
                      f"({attempt}/{retries})...")
                time.sleep(wait)
                continue
            if e.status_code in (429, 529) and _model_idx + 1 < len(MODELS):
                print(f"  {use_model} overloaded -> hopping to {MODELS[_model_idx + 1]}")
                _model_idx += 1
                attempt = 0
                continue
            raise


def _gemini_image_part(img: Image.Image) -> genai_types.Part:
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=90)
    return genai_types.Part.from_bytes(data=buf.getvalue(), mime_type="image/jpeg")


def _call_gemini_json(client: genai.Client, contents: list, schema: dict,
                      retries: int = 3) -> dict:
    """Gemini vision call with structured JSON output (single model)."""
    attempt = 0
    while True:
        try:
            resp = client.models.generate_content(
                model=GEMINI_MODEL,
                contents=contents,
                config=genai_types.GenerateContentConfig(
                    response_mime_type="application/json",
                    response_schema=schema,
                    max_output_tokens=1000,
                    **({"thinking_config": genai_types.ThinkingConfig(thinking_budget=0)}
                       if not GEMINI_MODEL.startswith("gemini-2.0") else {}),
                ),
            )
            break
        except genai.errors.APIError as e:
            if e.code in (429, 503) and attempt < retries:
                attempt += 1
                wait = 15 * attempt
                print(f"  transient Gemini {e.code}, retrying in {wait}s "
                      f"({attempt}/{retries})...")
                time.sleep(wait)
                continue
            raise
    text = (resp.text or "").strip()
    start, end = text.find("{"), text.rfind("}")
    if start != -1 and end > start:
        text = text[start:end + 1]
    try:
        res = json.loads(text)
    except json.JSONDecodeError:
        res = {"site_type": "unclear", "site_confidence": 0.0,
               "site_evidence": f"unparseable model reply: {text[:200]}"}
    normalize_model_result(res)
    res["model"] = GEMINI_MODEL
    return res


def _views_to_claude_content(views: list[tuple[str, Image.Image]], prompt: str) -> list:
    content = []
    for label, img in views:
        content.append({"type": "text", "text": f"View: {label}"})
        content.append(_image_block(img))
    content.append({"type": "text", "text": prompt})
    return content


def _views_to_gemini_contents(views: list[tuple[str, Image.Image]], prompt: str) -> list:
    contents = []
    for label, img in views:
        contents.append(f"View: {label}")
        contents.append(_gemini_image_part(img))
    contents.append(prompt)
    return contents


def classify_site(provider: str, clients: dict,
                  views: list[tuple[str, Image.Image]],
                  prompt: str = CLASSIFICATION_PROMPT, retries: int = 3,
                  scan: bool = False, claude_model: str | None = None) -> dict:
    """Classify one asset via Gemini or Claude using the same prompt."""
    if scan:
        claude_schema, gemini_schema = SCAN_SCHEMA, GEMINI_SCAN_SCHEMA
        tool_name = "scan_candidates"
        classify_prompt = SCAN_PROMPT
    else:
        claude_schema, gemini_schema = RESPONSE_SCHEMA, GEMINI_RESPONSE_SCHEMA
        tool_name = "classify_site"
        classify_prompt = prompt
    if provider == "gemini":
        contents = _views_to_gemini_contents(views, classify_prompt)
        return _call_gemini_json(clients["gemini"], contents, gemini_schema, retries)
    content = _views_to_claude_content(views, classify_prompt)
    return _call_claude_json(
        clients["claude"], content, claude_schema, tool_name, retries,
        model=claude_model)


def site_confidence_band(res: dict) -> str:
    """Map numeric site_confidence to high | medium | low for tier gating."""
    conf = normalize_confidence(res.get("site_confidence"))
    if conf is None:
        return "low"
    if conf >= TIER_CONF_HIGH:
        return "high"
    if conf >= TIER_CONF_MEDIUM:
        return "medium"
    return "low"


def tier_confident_stop(res: dict) -> bool:
    """True when tiered fetch can stop without pulling the next Nearmap tier."""
    if res.get("site_type") not in ("tower", "rooftop"):
        return False
    if site_confidence_band(res) == "low":
        return False
    if res.get("cell_equipment") is None:
        return False
    return True


def escalation_reason(res: dict) -> str | None:
    """Why a Gemini result should escalate to Claude; None if no escalation."""
    if res.get("site_type") == "other":
        return "other_type"
    if res.get("site_type") == "unclear":
        return "unclear_type"
    if site_confidence_band(res) == "low":
        return "low_confidence"
    return None


def classify_with_tiers(lat: float, lon: float, img: Image.Image | None,
                        provider: str, clients: dict, prompt: str,
                        input_confidence: str,
                        build_views) -> tuple[dict, dict, str | None, str, list]:
    """Tier 0 (NAIP) -> Tier 1 (Vert) -> Tier 2 (obliques). Returns
    (result, nearmap_views, nearmap_date, nearmap_tier, views)."""
    nearmap_views: dict = {}
    nearmap_date = None

    views = build_views({})
    res = classify_site(provider, clients, views, prompt=prompt)
    res = maybe_recheck_equipment(provider, clients, res, views, input_confidence)
    if tier_confident_stop(res):
        return res, nearmap_views, nearmap_date, "naip_only", views

    if not NEARMAP_API_KEY:
        return res, nearmap_views, nearmap_date, "naip_only", views

    vert_views, vert_date = fetch_nearmap_views(lat, lon, views=["Vert"])
    nearmap_views.update(vert_views)
    nearmap_date = vert_date or nearmap_date
    if not nearmap_views:
        return res, nearmap_views, nearmap_date, "naip_only", views

    views = build_views(nearmap_views)
    res = classify_site(provider, clients, views, prompt=prompt)
    res = maybe_recheck_equipment(provider, clients, res, views, input_confidence)
    if tier_confident_stop(res):
        return res, nearmap_views, nearmap_date, "vert_only", views

    missing = [v for v in OBLIQUE_VIEWS if v not in nearmap_views]
    if missing:
        oblique_views, ob_date = fetch_nearmap_views(lat, lon, views=missing)
        nearmap_views.update(oblique_views)
        nearmap_date = ob_date or nearmap_date

    views = build_views(nearmap_views)
    res = classify_site(provider, clients, views, prompt=prompt)
    res = maybe_recheck_equipment(provider, clients, res, views, input_confidence)
    return res, nearmap_views, nearmap_date, "full", views


def classify_chip(client: Anthropic, views: list[tuple[str, Image.Image]],
                  prompt: str = CLASSIFICATION_PROMPT, retries: int = 3) -> dict:
    """Legacy wrapper — prefer classify_site()."""
    return classify_site(
        "claude", {"claude": client}, views, prompt=prompt, retries=retries)


def _valid_box(box) -> list[int] | None:
    try:
        ymin, xmin, ymax, xmax = (int(v) for v in box[:4])
    except (TypeError, ValueError):
        return None
    if not (0 <= ymin < ymax <= 1000 and 0 <= xmin < xmax <= 1000):
        return None
    if (ymax - ymin) < ZOOM_MIN_FRAC * 1000 or (xmax - xmin) < ZOOM_MIN_FRAC * 1000:
        return None
    return [ymin, xmin, ymax, xmax]


def _grid_boxes(grid: int = ZOOM_GRID) -> list[list[int]]:
    """Return normalized boxes for an NxN grid covering the full image."""
    step = 1000 // grid
    boxes = []
    for row in range(grid):
        for col in range(grid):
            ymin = row * step
            xmin = col * step
            ymax = 1000 if row == grid - 1 else (row + 1) * step
            xmax = 1000 if col == grid - 1 else (col + 1) * step
            boxes.append([ymin, xmin, ymax, xmax])
    return boxes


def _crop_zoom(img: Image.Image, box: list[int]) -> Image.Image:
    """Magnify a normalized box from a source image to ZOOM_OUTPUT_PX."""
    w, h = img.size
    ymin, xmin, ymax, xmax = box
    pad_y = int((ymax - ymin) * ZOOM_PAD_FRAC)
    pad_x = int((xmax - xmin) * ZOOM_PAD_FRAC)
    ymin = max(0, ymin - pad_y)
    xmin = max(0, xmin - pad_x)
    ymax = min(1000, ymax + pad_y)
    xmax = min(1000, xmax + pad_x)
    left = int(xmin / 1000.0 * w)
    upper = int(ymin / 1000.0 * h)
    right = max(left + 1, int(xmax / 1000.0 * w))
    lower = max(upper + 1, int(ymax / 1000.0 * h))
    crop = img.crop((left, upper, right, lower))
    crop = crop.resize((ZOOM_OUTPUT_PX, ZOOM_OUTPUT_PX), Image.Resampling.LANCZOS)
    return crop


def scout_candidates(provider: str, clients: dict, label: str,
                     img: Image.Image) -> list[dict]:
    """Ask the vision model to propose candidate regions on one top-down image."""
    views = [(label, img)]
    res = classify_site(provider, clients, views, scan=True)
    return res.get("candidates") or []


def _anchor_candidates() -> list[dict]:
    """Default crops around the recorded coordinate (chip center) and the
    band just below center, where assets often sit when coords are approximate."""
    return [
        {"box_2d": [350, 350, 650, 650],
         "reason": "coordinate anchor (center)"},
        {"box_2d": [480, 380, 680, 620],
         "reason": "coordinate anchor (just below center)"},
    ]


def build_zoom_views(asset_id: str, source_label: str, source_img: Image.Image,
                     candidates: list[dict]) -> list[tuple[str, Image.Image]]:
    """Turn scout candidates into magnified zoom crops; save each to chips/."""
    zoom_views = []
    seen = set()
    for i, cand in enumerate(candidates[:ZOOM_MAX_CANDIDATES], start=1):
        box = _valid_box(cand.get("box_2d"))
        if box is None:
            continue
        key = tuple(box)
        if key in seen:
            continue
        seen.add(key)
        crop = _crop_zoom(source_img, box)
        reason = (cand.get("reason") or "candidate").replace("\n", " ")[:80]
        path = CHIP_DIR / f"{asset_id}_zoom_{i}.jpg"
        crop.save(path, quality=92)
        zoom_views.append((f"Zoom crop {i} ({reason})", crop))
    return zoom_views


def run_zoom_stage(provider: str, clients: dict, asset_id: str,
                   context_views: list[tuple[str, Image.Image]],
                   source_label: str, source_img: Image.Image,
                   max_crops: int = ZOOM_MAX_CANDIDATES) -> tuple[dict, int]:
    """Scout + magnify + re-classify. Returns (result dict, zoom crop count)."""
    scouted = scout_candidates(provider, clients, source_label, source_img)
    candidates = _anchor_candidates() + scouted
    if not scouted:
        print(f"  [{asset_id}] scout found no extra candidates")
    if not candidates:
        print(f"  [{asset_id}] no candidates -> {ZOOM_GRID}x{ZOOM_GRID} grid")
        candidates = [{"box_2d": b, "reason": "grid sweep"} for b in _grid_boxes()]

    zoom_views = build_zoom_views(asset_id, source_label, source_img,
                                  candidates[:max_crops])
    if not zoom_views:
        return {"site_type": "unclear", "site_confidence": 0.0,
                "site_evidence": "Zoom stage could not build valid crops."}, 0

    context = context_views[:1] if context_views else []
    res = classify_site(
        provider, clients, context + zoom_views, prompt=ZOOM_CLASSIFICATION_PROMPT)
    res["classification_stage"] = "zoom"
    return res, len(zoom_views)


def _row_error(record: dict) -> str | None:
    """Return a non-empty error string, or None if the row succeeded."""
    err = record.get("error")
    if err is None or (isinstance(err, float) and pd.isna(err)):
        return None
    err = str(err).strip()
    return err or None


def _format_asset_label(record: dict) -> str:
    label = str(record.get("label", "")).strip()
    aid = record.get("id", "")
    return f"{aid} ({label})" if label else str(aid)


def _format_located(record: dict) -> str:
    off = record.get("asset_offset_m")
    if off is not None and not (isinstance(off, float) and pd.isna(off)):
        return f"{off:.0f} m off"
    if record.get("asset_view"):
        return f"on {record['asset_view']}"
    return "—"


def _format_cell_equip(record: dict) -> str:
    ce = record.get("cell_equipment")
    ev = str(record.get("cell_equipment_evidence") or "").strip()
    if ce is True:
        return f"true — {ev}" if ev else "true"
    if ce is False:
        return f"false — {ev}" if ev else "false"
    if ce is None:
        return "unknown"
    return str(ce)


def _format_confidence(record: dict) -> str | float:
    conf = normalize_confidence(record.get("site_confidence"))
    if conf is None:
        return "—"
    return round(conf, 2)


def pick_review_image_path(asset_id: str, record: dict) -> Path | None:
    """Pick the best saved chip for stakeholder review (oblique > NAIP > zoom)."""
    chip = CHIP_DIR
    nm = (record.get("nearmap_views") or "").lower()
    asset_view = (record.get("asset_view") or "").lower()

    oblique_dirs = {
        "north": "north", "east": "east", "south": "south", "west": "west",
    }
    if record.get("cell_equipment") is True:
        for name, suffix in oblique_dirs.items():
            if name in nm or name in asset_view:
                path = chip / f"{asset_id}_nearmap_{suffix}.jpg"
                if path.exists():
                    return path

    for name, suffix in oblique_dirs.items():
        if name in asset_view:
            path = chip / f"{asset_id}_nearmap_{suffix}.jpg"
            if path.exists():
                return path

    if "nearmap top-down" in asset_view or "vert" in nm:
        path = chip / f"{asset_id}_nearmap_vert.jpg"
        if path.exists():
            return path

    chip_path = record.get("chip_path")
    if chip_path:
        path = Path(chip_path)
        if path.exists():
            return path

    for name in (f"{asset_id}_nearmap_vert.jpg", f"{asset_id}_NAIP.jpg"):
        path = chip / name
        if path.exists():
            return path

    zooms = sorted(chip.glob(f"{asset_id}_zoom_*.jpg"))
    return zooms[0] if zooms else None


def build_stakeholder_row(record: dict) -> dict:
    err = _row_error(record)
    if err:
        return {
            "Asset": _format_asset_label(record),
            "Site type": "error",
            "Conf": "—",
            "Located": "—",
            "Cell equip": err[:120],
            "Views": record.get("view_count", 0),
            "Review image": record.get("review_image"),
        }
    if record.get("site_type") == "no_imagery":
        return {
            "Asset": _format_asset_label(record),
            "Site type": "no imagery",
            "Conf": "—",
            "Located": "—",
            "Cell equip": "—",
            "Views": 0,
            "Review image": None,
        }
    return {
        "Asset": _format_asset_label(record),
        "Site type": record.get("site_type"),
        "Conf": _format_confidence(record),
        "Located": _format_located(record),
        "Cell equip": _format_cell_equip(record),
        "Views": record.get("view_count", 0),
        "Review image": record.get("review_image"),
    }


def write_stakeholder_report(results: list[dict], report_csv: str, report_xlsx: str):
    """Write a clean CSV + Excel workbook with embedded review images."""
    rows = [build_stakeholder_row(r) for r in results]
    report_df = pd.DataFrame(rows)
    report_df.to_csv(report_csv, index=False)

    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"Stakeholder CSV written to {report_csv} "
              "(install openpyxl for Excel export)")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["Asset", "Site type", "Conf", "Located", "Cell equip", "Views", "Photo"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[row_idx].height = 95
        ws.cell(row=row_idx, column=1, value=row["Asset"])
        ws.cell(row=row_idx, column=2, value=row["Site type"])
        ws.cell(row=row_idx, column=3, value=row["Conf"])
        ws.cell(row=row_idx, column=4, value=row["Located"])
        ws.cell(row=row_idx, column=5, value=row["Cell equip"])
        ws.cell(row=row_idx, column=6, value=row["Views"])

        img_path = row.get("Review image")
        if img_path and Path(img_path).exists():
            thumb = CHIP_DIR / f"_thumb_{Path(img_path).name}"
            with Image.open(img_path) as im:
                im = im.convert("RGB")
                im.thumbnail((160, 120))
                thumb_w, thumb_h = im.size
                im.save(thumb, quality=85)
            xl_img = XLImage(str(thumb))
            xl_img.width, xl_img.height = thumb_w, thumb_h
            col = get_column_letter(7)
            ws.add_image(xl_img, f"{col}{row_idx}")

    widths = {"A": 22, "B": 12, "C": 8, "D": 14, "E": 42, "F": 8, "G": 24}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    try:
        wb.save(report_xlsx)
    except PermissionError:
        alt = str(Path(report_xlsx).with_stem(Path(report_xlsx).stem + "_updated"))
        wb.save(alt)
        print(f"Stakeholder report written to {report_csv} and {alt}")
        print(f"  (Could not overwrite {report_xlsx} — file is open. Close it and "
              f"re-run --regenerate-report to refresh.)")
        return

    print(f"Stakeholder report written to {report_csv} and {report_xlsx}")


def write_executive_summary(results: list[dict], assets_df: pd.DataFrame):
    """Write a stakeholder-friendly markdown summary of the latest run."""
    total = len(assets_df)
    errors = [r for r in results if _row_error(r)]
    classified = [r for r in results if r.get("site_type") not in
                  (None, "no_imagery") and not _row_error(r)]
    towers = sum(1 for r in classified if r.get("site_type") == "tower")
    rooftops = sum(1 for r in classified if r.get("site_type") == "rooftop")
    cell_hits = sum(1 for r in classified if r.get("cell_equipment") is True)
    located = [r for r in classified if r.get("asset_lat") is not None
               or r.get("asset_view")]

    lines = [
        "# Asset Classifier — Executive Summary",
        "",
        f"*Generated: {time.strftime('%B %d, %Y at %I:%M %p')}*",
        "",
        "## At a glance",
        "",
        "We built an automated pipeline that takes a list of coordinates and "
        "determines whether each location is a **tower site** or **rooftop "
        "cellular site**, whether **cellular equipment is visible**, and where "
        "the asset sits relative to the recorded point.",
        "",
        f"- **Assets evaluated:** {total}",
        f"- **Successfully classified:** {len(classified)}",
        f"- **Tower sites identified:** {towers}",
        f"- **Rooftop sites identified:** {rooftops}",
        f"- **Cellular equipment detected:** {cell_hits}",
        f"- **Assets with a located position:** {len(located)}",
        f"- **Errors:** {len(errors)}",
        "",
        "## Pilot run highlights",
        "",
        "This proof-of-concept run combined **free public NAIP imagery**, "
        "**Nearmap high-resolution + 45° oblique views**, and **Anthropic Claude "
        "vision AI** to classify six sample assets (2 urban NJ, 4 rural W/C).",
        "",
        "**What worked well:**",
        "",
        "- **Rooftop detection (urban):** Nearmap obliques revealed rooftop "
        "antenna sectors on a classical building (asset_001, 95% confidence)",
        "- **Tower detection (rural):** Monopoles identified from shadow "
        "signatures even when 60 m from the recorded coordinate (asset_003)",
        "- **Disguised towers:** A monopine (tower disguised as a pine tree) "
        "was correctly identified only from oblique imagery (asset_004)",
        "- **Off-center assets:** The pipeline searches the full image, not "
        "just the center — critical when coordinates are imprecise",
        "",
        "**Improvement in progress:**",
        "",
        "- asset_005 (rural Oregon) was missed in wide imagery but a human "
        "reviewer confirmed a lattice tower in the top-right of the NAIP chip. "
        "A **two-stage zoom** pass (now implemented) magnifies suspicious "
        "regions before re-classifying — designed specifically for this case.",
        "",
        "## How it works",
        "",
        "```",
        "Coordinates (CSV)",
        "       |",
        "       v",
        "  +----+----+",
        "  |  NAIP   |  Wide public aerial (~250 m, ~1 m resolution)",
        "  +----+----+",
        "       |",
        "       v",
        "  +----+----+",
        "  | Nearmap |  High-res top-down + 45-degree obliques (urban/suburban)",
        "  +----+----+",
        "       |",
        "       v",
        "  +----+----+",
        "  | Claude  |  AI vision: classify site, locate asset, detect equipment",
        "  +----+----+",
        "       |",
        "       v (if rural / still unidentified)",
        "  +----+----+",
        "  |  Zoom   |  Magnify suspicious regions and re-classify",
        "  +----+----+",
        "       |",
        "       v",
        "  results.csv + review chips + this summary",
        "```",
        "",
        "### Imagery sources",
        "",
        "| Source | What it provides | Why it matters |",
        "|---|---|---|",
        "| **NAIP** (free, public) | Wide top-down context around each point | "
        "Catches off-center towers; cheap baseline for the full US |",
        "| **Nearmap** (subscription) | ~7 cm top-down + 45° oblique views | "
        "Makes rooftop antennas and disguised towers (e.g. monopines) visible |",
        "| **Claude** (Anthropic) | Structured classification from multi-image input | "
        "Turns imagery into site type, equipment call, and location |",
        "",
        "### Confidence safeguards",
        "",
        "1. **Whole-image search** — never assumes the asset is at the exact center",
        "2. **Multi-view fusion** — NAIP context + Nearmap detail + oblique angles",
        "3. **Rural fallback** — widens Nearmap area when only vertical imagery exists",
        "4. **Two-stage zoom** — magnifies subtle structures the wide view missed",
        "5. **Human review chips** — every image sent to the model is saved for audit",
        "",
        "## Results by asset",
        "",
        "| Asset | Region | Site type | Confidence | Cell equip. | Located | Method | Key finding |",
        "|---|---|---|---:|---|---|---|---|",
    ]

    label_map = {k: str(v).strip()
                 for k, v in zip(assets_df["id"], assets_df.get("label", pd.Series(dtype=str)))}
    for r in results:
        aid = r.get("id", "")
        region = label_map.get(aid, str(r.get("label", "")).strip())
        err = _row_error(r)
        if err:
            short = err if len(err) <= 60 else err[:60] + "…"
            row = f"| {aid} | {region} | **ERROR** | — | — | — | — | {short} |"
        elif r.get("site_type") == "no_imagery":
            row = f"| {aid} | {region} | No imagery | — | — | — | — | Outside coverage |"
        else:
            conf = r.get("site_confidence")
            conf_s = f"{conf:.0%}" if isinstance(conf, (int, float)) else "—"
            cell = r.get("cell_equipment")
            cell_s = {True: "Yes", False: "No", None: "Unknown"}.get(cell, "—")
            if r.get("asset_lat") is not None:
                off = r.get("asset_offset_m")
                if off is not None and not (isinstance(off, float) and pd.isna(off)):
                    loc = f"{off:.0f} m off"
                else:
                    loc = "yes"
            elif r.get("asset_view"):
                loc = f"on {r['asset_view']}"
            else:
                loc = "—"
            stage = r.get("classification_stage") or "primary"
            if r.get("nearmap_aoi_m") == NEARMAP_FALLBACK_CHIP_M:
                stage = "wide AOI"
            evidence = (r.get("site_evidence") or "")[:90]
            if len(r.get("site_evidence") or "") > 90:
                evidence += "…"
            row = (f"| {aid} | {region} | {r.get('site_type', '—')} | {conf_s} | "
                   f"{cell_s} | {loc} | {stage} | {evidence} |")
        lines.append(row)

    lines.extend([
        "",
        "## Operational notes",
        "",
        f"- **Nearmap data usage (pilot run):** ~15 MB for {total} assets "
        "(under 1% of the 2.49 GB/month subscription allowance)",
        "- **Review folder:** saved images in `chips/` — NAIP chips named "
        "`*_NAIP.jpg`, Nearmap views `*_nearmap_*.jpg`, zoom crops `*_zoom_*.jpg`",
        "- **Machine-readable output:** `results.csv` for downstream systems",
        "",
        "## Known limitations",
        "",
        "- Rural sites may have Nearmap vertical imagery only (no 45° obliques)",
        "- Very small lattice towers can still be missed until the zoom stage runs",
        "- Recorded coordinates can be tens of meters off the true asset",
        "- AI calls should be spot-checked on low-confidence results (< 60%)",
        "",
        "## Recommended next steps",
        "",
        "1. Spot-check chips for any low-confidence or unexpected classifications",
        "2. Scale to the full asset list once stakeholders approve the approach",
        "3. Feed confirmed results back into the coordinate/enrichment workflow",
        "",
    ])

    Path(EXECUTIVE_SUMMARY_MD).write_text("\n".join(lines), encoding="utf-8")


def setup_run_directory(prefix: str, run_dir: str | None) -> Path:
    """Create or reopen a timestamped run folder for all outputs."""
    if run_dir:
        run_root = Path(run_dir)
        if not run_root.is_dir():
            raise SystemExit(f"Run directory not found: {run_root}")
        print(f"Resuming run folder: {run_root}", flush=True)
    else:
        stamp = time.strftime("%Y-%m-%d_%H%M%S")
        run_root = RUNS_DIR / f"{stamp}_{prefix}"
        run_root.mkdir(parents=True, exist_ok=True)
        print(f"Created run folder: {run_root}", flush=True)
    (run_root / "chips").mkdir(exist_ok=True)
    return run_root


def _print_run_banner(total: int, pending: int, skipped: int, input_csv: str,
                      run_dir: Path, output_csv: str, report_csv: str | None):
    """Startup summary so the operator can monitor the run in the terminal."""
    est_min = max(1, round(pending * 0.35))  # ~12s pacing + API work per site
    print("\n" + "=" * 60, flush=True)
    print("  SITE CLASSIFIER RUN", flush=True)
    print("=" * 60, flush=True)
    print(f"  Run folder: {run_dir}", flush=True)
    print(f"  Input:      {input_csv}", flush=True)
    print(f"  Detail CSV: {output_csv}", flush=True)
    if report_csv:
        print(f"  Report:     {report_csv} (+ Excel with photos)", flush=True)
    print(f"  Chips:      {run_dir / 'chips'}", flush=True)
    print(f"  Total:      {total} assets | {skipped} already done | {pending} to run",
          flush=True)
    print(f"  Est. time:  ~{est_min} min for {pending} remaining", flush=True)
    print("=" * 60 + "\n", flush=True)


def _print_asset_start(idx: int, total: int, asset_id: str, row):
    """Mark the start of each asset in the terminal log."""
    addr = _clean_address(row)
    coord = (f"{row['lat']}, {row['lon']}" if _has_coordinates(row)
             else (addr[:55] + "…" if addr and len(addr) > 55 else addr))
    print(f"\n>>> [{idx}/{total}] {asset_id} — {coord}", flush=True)


def _print_asset_result(record: dict):
    """One-line success/failure summary after each asset."""
    err = _row_error(record)
    if err:
        print(f"    RESULT: ERROR — {err[:100]}", flush=True)
        return
    site = record.get("site_type", "—")
    conf = record.get("site_confidence")
    conf_s = f"{conf:.2f}" if isinstance(conf, (int, float)) else "—"
    cell = record.get("cell_equipment")
    cell_s = {True: "yes", False: "no", None: "?"}.get(cell, str(cell))
    located = _format_located(record)
    views = record.get("view_count", 0)
    photo = "photo saved" if record.get("review_image") else "no photo"
    print(f"    RESULT: {site} | conf {conf_s} | cell {cell_s} | "
          f"{located} | {views} views | {photo}", flush=True)


def _print_run_complete(results: list[dict], run_dir: Path, output_csv: str,
                        report_csv: str | None, report_xlsx: str | None,
                        summary_md: str):
    """Final terminal summary when the batch finishes."""
    errors = sum(1 for r in results if _row_error(r))
    ok = len(results) - errors
    towers = sum(1 for r in results if r.get("site_type") == "tower")
    rooftops = sum(1 for r in results if r.get("site_type") == "rooftop")
    cell_hits = sum(1 for r in results if r.get("cell_equipment") is True)
    print("\n" + "=" * 60, flush=True)
    print("  RUN COMPLETE", flush=True)
    print("=" * 60, flush=True)
    print(f"  Folder:   {run_dir}", flush=True)
    print(f"  Classified: {ok}/{len(results)}  |  Errors: {errors}", flush=True)
    print(f"  Towers: {towers}  |  Rooftops: {rooftops}  |  Cell equip: {cell_hits}",
          flush=True)
    print(f"  Detail:   {output_csv}", flush=True)
    if report_csv:
        print(f"  Report:   {report_csv}", flush=True)
        print(f"  Excel:    {report_xlsx}  (includes embedded photos)", flush=True)
    print(f"  Summary:  {summary_md}", flush=True)
    print("=" * 60 + "\n", flush=True)


def regenerate_reports_from_detail(run_root: Path, output_csv: str,
                                   report_csv: str | None, report_xlsx: str | None,
                                   summary_md: str, input_csv: str):
    """Fix confidence values and rebuild report files without API calls."""
    if not os.path.exists(output_csv):
        raise SystemExit(f"No detail file found: {output_csv}")

    results = pd.read_csv(output_csv).to_dict("records")
    fixed = 0
    for record in results:
        before = (record.get("site_confidence"), record.get("cell_equipment_confidence"))
        normalize_model_result(record)
        after = (record.get("site_confidence"), record.get("cell_equipment_confidence"))
        if before != after:
            fixed += 1

    pd.DataFrame(results).to_csv(output_csv, index=False)

    input_path = run_root / Path(input_csv).name
    assets_df = (pd.read_csv(input_path) if input_path.exists()
                 else pd.read_csv(input_csv))

    write_executive_summary(results, assets_df)
    if report_csv and report_xlsx:
        write_stakeholder_report(results, report_csv, report_xlsx)

    print(f"\nRegenerated reports in {run_root}", flush=True)
    print(f"  Fixed confidence on {fixed} row(s)", flush=True)
    print(f"  Detail:  {output_csv}", flush=True)
    if report_csv:
        print(f"  Report:  {report_csv}", flush=True)
        print(f"  Excel:   {report_xlsx}", flush=True)
    print(f"  Summary: {summary_md}", flush=True)

def classify_with_routing(provider: str, clients: dict, views: list,
                          prompt: str, input_confidence: str
                          ) -> tuple[dict, str, str | None, str | None]:
    """Run primary classification; optionally escalate Gemini -> Claude."""
    primary_model = provider
    res = classify_site(provider, clients, views, prompt=prompt)
    res = maybe_recheck_equipment(provider, clients, res, views, input_confidence)

    escalation_model = None
    escalation_reason_str = None
    if BIFURCATED_AI and provider == "gemini":
        reason = escalation_reason(res)
        if reason:
            escalation_reason_str = reason
            escalation_model = "claude"
            print(f"  escalating to Claude ({reason.replace('_', ' ')})")
            res = classify_site(
                "claude", clients, views, prompt=prompt,
                claude_model=CLAUDE_ESCALATION_MODEL)
            res = maybe_recheck_equipment(
                "claude", clients, res, views, input_confidence)
    return res, primary_model, escalation_model, escalation_reason_str


def _effective_provider(primary_model: str, escalation_model: str | None) -> str:
    return escalation_model or primary_model


def main():
    global INPUT_CSV, OUTPUT_CSV, EXECUTIVE_SUMMARY_MD, CHIP_DIR, RUN_DIR

    parser = argparse.ArgumentParser(description="Classify cell sites from aerial imagery")
    parser.add_argument("--input", "-i", default=INPUT_CSV,
                        help=f"Input CSV (default: {INPUT_CSV})")
    parser.add_argument("--run-dir", default=None,
                        help="Resume an existing timestamped run folder under runs/")
    parser.add_argument("--output", "-o", default=None,
                        help="Detail output CSV filename inside the run folder")
    parser.add_argument("--report-csv", default=None,
                        help="Stakeholder summary CSV filename inside the run folder")
    parser.add_argument("--report-xlsx", default=None,
                        help="Stakeholder Excel with photos inside the run folder")
    parser.add_argument("--regenerate-report", action="store_true",
                        help="Fix confidence values and rebuild reports from detail CSV")
    args = parser.parse_args()

    if args.regenerate_report and not args.run_dir:
        raise SystemExit("--regenerate-report requires --run-dir pointing at an "
                         "existing run folder (e.g. runs\\2026-06-12_184824_WI)")

    INPUT_CSV = args.input
    stem = Path(INPUT_CSV).stem
    prefix = stem.replace("_assets", "") if stem.endswith("_assets") else stem

    run_root = setup_run_directory(prefix, args.run_dir)
    RUN_DIR = run_root
    CHIP_DIR = run_root / "chips"

    detail_name = (Path(args.output).name if args.output
                   else f"{prefix}_results_detail.csv"
                   if stem.endswith("_assets") else "results_detail.csv")
    OUTPUT_CSV = str(run_root / detail_name)

    report_csv = None
    report_xlsx = None
    if stem.endswith("_assets") or args.report_csv or args.report_xlsx:
        report_csv = str(run_root / (Path(args.report_csv).name if args.report_csv
                                       else f"{prefix}_results.csv"))
        report_xlsx = str(run_root / (Path(args.report_xlsx).name if args.report_xlsx
                                        else f"{prefix}_results.xlsx"))

    EXECUTIVE_SUMMARY_MD = str(run_root / (
        f"{prefix}_EXECUTIVE_SUMMARY.md"
        if stem.endswith("_assets") else "EXECUTIVE_SUMMARY.md"))

    if not args.run_dir and not args.regenerate_report:
        shutil.copy2(INPUT_CSV, run_root / Path(INPUT_CSV).name)

    if args.regenerate_report:
        regenerate_reports_from_detail(
            run_root, OUTPUT_CSV, report_csv, report_xlsx,
            EXECUTIVE_SUMMARY_MD, INPUT_CSV)
        return

    if not os.environ.get("ANTHROPIC_API_KEY"):
        raise SystemExit(
            "ANTHROPIC_API_KEY is not set. Get a key at https://console.anthropic.com/\n"
            '  PowerShell: $env:ANTHROPIC_API_KEY="sk-ant-..."\n'
            "  bash/zsh:   export ANTHROPIC_API_KEY=sk-ant-..."
        )
    if BIFURCATED_AI and not os.environ.get("GEMINI_API_KEY"):
        raise SystemExit(
            "BIFURCATED_AI=1 requires GEMINI_API_KEY. Get a key at "
            "https://aistudio.google.com/apikey"
        )

    clients = {
        "claude": Anthropic(),
        "gemini": genai.Client(api_key=os.environ.get("GEMINI_API_KEY"))
        if BIFURCATED_AI else None,
    }
    primary_provider = "gemini" if BIFURCATED_AI else "claude"

    if NAIP_ONLY:
        print("[BREAKPOINT] NAIP_ONLY=1 — Nearmap fetch disabled. "
              "Running on NAIP imagery only.", flush=True)
    if NEARMAP_TIERED and not NAIP_ONLY:
        print("NEARMAP_TIERED=1 — tiered Nearmap fetch enabled "
              "(NAIP -> Vert -> obliques)", flush=True)
    if BIFURCATED_AI:
        print(f"BIFURCATED_AI=1 — Gemini first ({GEMINI_MODEL}), "
              f"Claude escalation ({CLAUDE_ESCALATION_MODEL})", flush=True)

    print(f"Input:  {INPUT_CSV}")
    print(f"Output: {OUTPUT_CSV} (detail/resume)")
    if report_csv:
        print(f"Report: {report_csv} + {report_xlsx}")

    df = pd.read_csv(INPUT_CSV)
    validate_input_csv(df)

    # Resume support: keep successfully classified rows from a previous run and
    # skip them, so quota is only spent on assets that still need work
    results = []
    done_ids = set()
    if os.path.exists(OUTPUT_CSV):
        prev = pd.read_csv(OUTPUT_CSV)
        if "site_type" in prev.columns:
            mask = prev["site_type"].notna()
            if "error" in prev.columns:
                mask &= prev["error"].isna()
            kept = prev[mask]
            results = kept.to_dict("records")
            done_ids = set(kept["id"])
            if done_ids:
                print(f"Resuming: {len(done_ids)} assets already done, "
                      f"{len(df) - len(done_ids)} remaining")

    pending_rows = [row for _, row in df.iterrows() if row["id"] not in done_ids]
    _print_run_banner(len(df), len(pending_rows), len(done_ids),
                      INPUT_CSV, run_root, OUTPUT_CSV, report_csv)

    progress = tqdm(pending_rows, desc="Progress", unit="site", dynamic_ncols=True)
    for row in progress:
        progress.set_postfix_str(str(row["id"]), refresh=False)
        _print_asset_start(len(done_ids) + progress.n + 1, len(df), row["id"], row)
        # Carry all input columns (id, lat, lon, address, label, ...) into results
        record = row.to_dict()
        try:
            lat, lon, geocode_meta = resolve_row_coordinates(row)
            record["lat"] = lat
            record["lon"] = lon
            record.update(geocode_meta)
            if geocode_meta:
                print(f"    geocoded ({geocode_meta.get('geocode_source')}): "
                      f"{lat:.6f}, {lon:.6f}", flush=True)

            img, img_date, naip_geo = fetch_chip(lat, lon)

            nearmap_views, nearmap_date = {}, None
            if not NAIP_ONLY:
                try:
                    if not NEARMAP_TIERED:
                        nearmap_views, nearmap_date = fetch_nearmap_views(lat, lon)
                except Exception as e:
                    print(f"  [{row['id']}] nearmap fetch failed: {e}")

            if img is None and not nearmap_views:
                record["site_type"] = "no_imagery"
                results.append(record)
                _print_asset_result(record)
                pd.DataFrame(results).to_csv(OUTPUT_CSV, index=False)
                time.sleep(API_DELAY_S)
                continue

            def build_views(nm_views):
                """Save Nearmap chips (overwriting any narrow-AOI versions)
                and assemble the labeled view list for the model."""
                v = []
                if img is not None:
                    v.append(("NAIP top-down", img))
                for name, vimg in nm_views.items():
                    vpath = CHIP_DIR / f"{row['id']}_nearmap_{name.lower()}.jpg"
                    vimg.save(vpath, quality=90)
                    label = ("Nearmap top-down" if name == "Vert"
                             else f"Nearmap oblique ({name})")
                    v.append((label, vimg))
                return v

            chip_path = None
            if img is not None:
                chip_path = CHIP_DIR / f"{row['id']}_NAIP.jpg"
                img.save(chip_path, quality=90)

            label_hint = str(row.get("label", "")).strip().lower()
            input_confidence = normalize_input_confidence(row.get("input_confidence"))
            prompt = build_classification_prompt(row)

            primary_model = primary_provider
            escalation_model = None
            escalation_reason_str = None

            if NAIP_ONLY:
                views = build_views({})
                nearmap_tier = "naip_only"
                res, primary_model, escalation_model, escalation_reason_str = (
                    classify_with_routing(
                        primary_provider, clients, views, prompt, input_confidence))
            elif NEARMAP_TIERED:
                res, nearmap_views, nearmap_date, nearmap_tier, views = (
                    classify_with_tiers(
                        lat, lon, img, primary_provider, clients, prompt,
                        input_confidence, build_views))
                if BIFURCATED_AI:
                    reason = escalation_reason(res)
                    if reason:
                        escalation_reason_str = reason
                        escalation_model = "claude"
                        print(f"  escalating to Claude ({reason.replace('_', ' ')})")
                        res = classify_site(
                            "claude", clients, views, prompt=prompt,
                            claude_model=CLAUDE_ESCALATION_MODEL)
                        res = maybe_recheck_equipment(
                            "claude", clients, res, views, input_confidence)
            else:
                nearmap_tier = "full" if nearmap_views else "naip_only"
                views = build_views(nearmap_views)
                res, primary_model, escalation_model, escalation_reason_str = (
                    classify_with_routing(
                        primary_provider, clients, views, prompt, input_confidence))

            nearmap_aoi_m = NEARMAP_CHIP_M if nearmap_views else None
            classification_stage = "primary"
            zoom_count = 0
            stage_provider = _effective_provider(primary_model, escalation_model)

            # Wide-AOI fallback: rural sites often have vert-only Nearmap
            # coverage, and the asset may sit outside the narrow AOI entirely
            has_obliques = any(n != "Vert" for n in nearmap_views)
            if (not NAIP_ONLY and NEARMAP_API_KEY
                    and res.get("site_type") in ("other", "unclear")
                    and not has_obliques):
                print(f"  [{row['id']}] unidentified with narrow AOI -> "
                      f"retrying Nearmap at {NEARMAP_FALLBACK_CHIP_M}m")
                try:
                    wide_views, wide_date = fetch_nearmap_views(
                        lat, lon, NEARMAP_FALLBACK_CHIP_M)
                except Exception as e:
                    wide_views, wide_date = {}, None
                    print(f"  [{row['id']}] wide nearmap fetch failed: {e}")
                if wide_views:
                    nearmap_views = wide_views
                    nearmap_date = wide_date or nearmap_date
                    nearmap_aoi_m = NEARMAP_FALLBACK_CHIP_M
                    views = build_views(wide_views)
                    res, _, esc_m, esc_r = classify_with_routing(
                        stage_provider, clients, views, prompt, input_confidence)
                    if esc_m:
                        escalation_model = esc_m
                        escalation_reason_str = esc_r or escalation_reason_str
                    classification_stage = "wide_aoi"
                    nearmap_tier = "wide_aoi"
                    stage_provider = _effective_provider(primary_model, escalation_model)

            # Two-stage zoom: scout suspicious regions, magnify, re-classify.
            # Stealth-tagged sites always get a zoom pass - wide imagery often
            # misreads building-integrated towers as generic rooftops.
            force_zoom = label_hint == "stealth"
            if force_zoom or res.get("site_type") in ("other", "unclear"):
                source_label, source_img = None, None
                if nearmap_views.get("Vert"):
                    source_label, source_img = "Nearmap top-down", nearmap_views["Vert"]
                elif img is not None:
                    source_label, source_img = "NAIP top-down", img
                if source_img is not None:
                    print(f"  [{row['id']}] running two-stage zoom on {source_label}")
                    zoom_res, zoom_count = run_zoom_stage(
                        stage_provider, clients, row["id"], views,
                        source_label, source_img,
                        max_crops=3 if force_zoom else ZOOM_MAX_CANDIDATES)
                    prior_conf = res.get("site_confidence") or 0
                    zoom_conf = zoom_res.get("site_confidence") or 0
                    zoom_wins = (
                        zoom_res.get("site_type") in ("tower", "rooftop")
                        or zoom_conf > prior_conf
                        or (force_zoom and zoom_res.get("cell_equipment") is True)
                        or force_zoom
                    )
                    if zoom_wins:
                        res = zoom_res
                        classification_stage = "zoom"
                        nearmap_tier = "zoom"

            # Convert the detection box to real-world coordinates - only valid
            # when the box was drawn on the georeferenced NAIP chip
            asset_lat = asset_lon = asset_offset_m = None
            box, box_view = res.get("asset_box_2d"), res.get("asset_view")
            if box and box_view == "NAIP top-down" and naip_geo:
                located = box_to_latlon(naip_geo, box)
                if located:
                    asset_lat, asset_lon, asset_offset_m = located

            record.update({
                "image_date": img_date,
                "nearmap_date": nearmap_date,
                "nearmap_views": ",".join(nearmap_views) or None,
                "nearmap_aoi_m": nearmap_aoi_m,
                "chip_path": str(chip_path) if chip_path else None,
                "view_count": len(views),
                "site_type": res.get("site_type"),
                "site_confidence": res.get("site_confidence"),
                "site_evidence": res.get("site_evidence"),
                "asset_lat": asset_lat,
                "asset_lon": asset_lon,
                "asset_offset_m": (round(asset_offset_m, 1)
                                   if asset_offset_m is not None else None),
                "asset_box_2d": json.dumps(box) if box else None,
                "asset_view": box_view,
                "cell_equipment": res.get("cell_equipment"),
                "cell_equipment_confidence": res.get("cell_equipment_confidence"),
                "cell_equipment_evidence": res.get("cell_equipment_evidence"),
                "classification_stage": classification_stage,
                "zoom_crops": zoom_count or None,
                "input_confidence": input_confidence,
                "source_trust_mismatch": (
                    input_confidence == "high"
                    and res.get("cell_equipment") is False),
                "model": res.get("model"),
                "nearmap_tier": nearmap_tier,
                "primary_model": primary_model,
                "escalation_model": escalation_model,
                "escalation_reason": escalation_reason_str,
            })
            review_path = pick_review_image_path(row["id"], record)
            if review_path:
                record["review_image"] = str(review_path)
            loc = (f"({asset_lat:.6f},{asset_lon:.6f}, {asset_offset_m:.0f}m off)"
                   if asset_lat is not None else f"(box on: {box_view})")
            print(f"    {record['site_type']} ({record['site_confidence']}) {loc} "
                  f"| cell: {record['cell_equipment']} | stage: {classification_stage}",
                  flush=True)
            _print_asset_result(record)

        except Exception as e:
            record["error"] = str(e)
            print(f"    ERROR: {e}", flush=True)
            _print_asset_result(record)

        results.append(record)
        # Rewrite after every row so a mid-run crash never loses completed work
        pd.DataFrame(results).to_csv(OUTPUT_CSV, index=False)
        time.sleep(API_DELAY_S)

    write_executive_summary(results, df)
    if report_csv and report_xlsx:
        write_stakeholder_report(results, report_csv, report_xlsx)
    _print_run_complete(results, run_root, OUTPUT_CSV, report_csv, report_xlsx,
                      EXECUTIVE_SUMMARY_MD)


if __name__ == "__main__":
    main()
