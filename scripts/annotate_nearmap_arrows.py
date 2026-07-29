"""Annotate Nearmap chips with a blue POI arrow and write an Excel workbook.

The requested lat/lon is the center of each Nearmap AOI chip, so the arrow tip
is placed at image center. Works offline on an existing run folder (no Nearmap
API calls).

Example:
  python scripts/annotate_nearmap_arrows.py --run-dir runs/2026-07-24_132449_Alma_Realty_Locations
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from PIL import Image, ImageDraw

_REPO_ROOT = Path(__file__).resolve().parents[1]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

PHOTO_COLS = (
    "Photo Vert 50m",
    "Photo Oblique 50m",
    "Photo Oblique 250m",
)
BLUE = (0, 102, 255)
BLUE_OUTLINE = (0, 40, 140)
HALO = (255, 255, 255)


def draw_poi_arrow(im: Image.Image,
                   tip: tuple[int, int] | None = None) -> Image.Image:
    """Draw a blue arrow pointing down at the POI (default: image center)."""
    out = im.convert("RGB").copy()
    draw = ImageDraw.Draw(out)
    w, h = out.size
    cx, cy = tip if tip is not None else (w // 2, h // 2)

    size = max(28, min(w, h) // 10)
    thickness = max(4, size // 5)
    head_w = size * 0.55
    head_h = size * 0.7
    shaft_top = cy - int(size * 2.4)
    shaft_bot = cy - int(head_h * 0.85)

    def _arrow(fill, inflate: int = 0):
        t = thickness + inflate * 2
        hw = head_w + inflate
        hh = head_h + inflate
        # shaft
        draw.rectangle(
            [cx - t // 2, shaft_top - inflate, cx + t // 2, shaft_bot + inflate],
            fill=fill,
        )
        # arrow head (triangle) tip at (cx, cy)
        tip_y = cy + inflate
        draw.polygon(
            [
                (cx, tip_y),
                (cx - hw, tip_y - hh),
                (cx + hw, tip_y - hh),
            ],
            fill=fill,
        )

    _arrow(HALO, inflate=3)
    _arrow(BLUE_OUTLINE, inflate=1)
    _arrow(BLUE, inflate=0)
    # small target dot under the tip
    r = max(3, thickness // 2)
    draw.ellipse([cx - r - 1, cy - r - 1, cx + r + 1, cy + r + 1], fill=HALO)
    draw.ellipse([cx - r, cy - r, cx + r, cy + r], fill=BLUE)
    return out


def annotate_path(src: Path, dest: Path | None = None) -> Path | None:
    if not src.is_file():
        return None
    dest = dest or src.with_name(src.stem + "_marked" + src.suffix)
    with Image.open(src) as im:
        marked = draw_poi_arrow(im)
        marked.save(dest, quality=92)
    return dest


def write_xlsx(df: pd.DataFrame, xlsx_path: Path, photo_cols: tuple[str, ...],
               thumb_max: int = 220) -> Path:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Alma Nearmap"

    # Text columns for paths; embed images instead
    headers = [c for c in df.columns if c not in photo_cols] + list(photo_cols)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    text_cols = [c for c in headers if c not in photo_cols]
    photo_start = len(text_cols) + 1

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        ws.row_dimensions[row_idx].height = 120
        for col_idx, col_name in enumerate(text_cols, start=1):
            val = row.get(col_name, "")
            if pd.isna(val):
                val = ""
            ws.cell(row=row_idx, column=col_idx, value=val)

        for i, col_name in enumerate(photo_cols):
            col_idx = photo_start + i
            path_str = str(row.get(col_name) or "").strip()
            if not path_str:
                continue
            img_path = Path(path_str)
            if not img_path.is_file():
                img_path = _REPO_ROOT / path_str
            if not img_path.is_file():
                ws.cell(row=row_idx, column=col_idx, value=path_str)
                continue

            thumb = img_path.with_name(f"_thumb_{img_path.name}")
            with Image.open(img_path) as im:
                im = im.convert("RGB")
                im.thumbnail((thumb_max, thumb_max))
                tw, th = im.size
                im.save(thumb, quality=85)
            xl_img = XLImage(str(thumb))
            xl_img.width, xl_img.height = tw, th
            ws.add_image(xl_img, f"{get_column_letter(col_idx)}{row_idx}")

    for col_idx, col_name in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        if col_name in photo_cols:
            ws.column_dimensions[letter].width = 32
        elif col_name in ("Address", "Current Tenants"):
            ws.column_dimensions[letter].width = 28
        else:
            ws.column_dimensions[letter].width = max(10, min(18, len(col_name) + 2))

    try:
        wb.save(xlsx_path)
    except PermissionError:
        alt = xlsx_path.with_stem(xlsx_path.stem + "_updated")
        wb.save(alt)
        return alt
    return xlsx_path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--run-dir",
        required=True,
        help="Existing run folder with *_with_nearmap.csv and chips/",
    )
    parser.add_argument(
        "--csv",
        default=None,
        help="Override path to the with_nearmap CSV",
    )
    args = parser.parse_args()

    run_dir = Path(args.run_dir)
    if not run_dir.is_dir():
        raise SystemExit(f"Run directory not found: {run_dir}")

    csv_path = Path(args.csv) if args.csv else None
    if csv_path is None:
        matches = list(run_dir.glob("*_with_nearmap.csv"))
        if not matches:
            raise SystemExit(f"No *_with_nearmap.csv in {run_dir}")
        csv_path = matches[0]

    df = pd.read_csv(csv_path)
    missing = [c for c in PHOTO_COLS if c not in df.columns]
    if missing:
        raise SystemExit(f"CSV missing columns {missing}: {csv_path}")

    annotated = 0
    for idx, row in df.iterrows():
        for col in PHOTO_COLS:
            path_str = str(row.get(col) or "").strip()
            if not path_str:
                continue
            src = Path(path_str)
            if not src.is_file():
                src = _REPO_ROOT / path_str
            if not src.is_file():
                print(f"  missing: {path_str}", flush=True)
                continue
            # Prefer annotating the unmarked original if a prior marked path
            if src.stem.endswith("_marked"):
                candidate = src.with_name(src.stem[: -len("_marked")] + src.suffix)
                if candidate.is_file():
                    src = candidate
            dest = src.with_name(src.stem + "_marked" + src.suffix)
            result = annotate_path(src, dest)
            if result is None:
                continue
            # Store path relative to repo when under repo
            try:
                rel = result.resolve().relative_to(_REPO_ROOT.resolve())
                df.at[idx, col] = rel.as_posix()
            except ValueError:
                df.at[idx, col] = str(result.as_posix())
            annotated += 1
            print(f"  marked {result.name}", flush=True)

    out_csv = csv_path.with_name(csv_path.stem + "_marked.csv")
    df.to_csv(out_csv, index=False)

    xlsx_path = run_dir / (csv_path.stem.replace("_with_nearmap", "") +
                           "_with_nearmap_marked.xlsx")
    saved = write_xlsx(df, xlsx_path, PHOTO_COLS)
    print(
        f"\nDone. Annotated {annotated} images.\n"
        f"  csv:  {out_csv}\n"
        f"  xlsx: {saved}",
        flush=True,
    )


if __name__ == "__main__":
    main()
