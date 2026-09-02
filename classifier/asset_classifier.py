"""
Asset classifier pipeline: coordinates or street address -> NAIP aerial chip
-> Claude vision classification.

Flow:
  1. Read sites from assets.csv (columns: id; plus lat+lon OR address; optional
     label, input_confidence: high | medium | low for source trust). Addresses
     are geocoded to lat/lon via the free US Census Geocoder (CONUS) with
     OpenStreetMap Nominatim as fallback.
  2. Query Microsoft Planetary Computer STAC API for the newest NAIP scene at each point
  3. Windowed-read a chip around the point from the Cloud-Optimized GeoTIFF (no full download)
  4. Optional: if NEARMAP_API_KEY is set, also pull a high-res Nearmap vertical
     and 45-degree oblique panoramas (N/E/S/W) via the Tile API
  5. Send all views to Claude: classify the site (tower vs rooftop), locate the
     asset with a bounding box, and assess visible cellular equipment
  6. If still unidentified (rural vert-only), widen the Nearmap AOI to match NAIP
  7. If still unidentified, run a two-stage zoom: scout candidate regions, crop
     and magnify them, then re-classify on the zoomed views
  8. Convert the detection box on the georeferenced NAIP chip to asset lat/lon
  9. Write results.csv, chips for spot-checking, and an executive summary markdown

Setup:
  pip install -r requirements.txt
  Get an API key at https://console.anthropic.com/ then:
  export ANTHROPIC_API_KEY=sk-ant-...

Notes:
  - NAIP covers the continental US only (~0.6-1m resolution, public domain).
    For other regions, swap the STAC collection (e.g. state orthoimagery, OpenAerialMap).
  - Chip size: 250m at 0.6m GSD = ~417px square. Good balance of context vs detail.
"""

import base64
import io
import json
import math
import os
import random
import sys
import time
import argparse
import shutil
from datetime import date
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parents[1]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

import numpy as np
import pandas as pd
import rasterio
import requests
from rasterio.windows import from_bounds
from pyproj import Transformer
from pystac_client import Client
import planetary_computer
from PIL import Image
import anthropic
from anthropic import Anthropic
from google import genai
from google.genai import types as genai_types
from dotenv import load_dotenv

try:
    from tqdm import tqdm
except ImportError:
    def tqdm(iterable, **kwargs):
        return iterable

load_dotenv()  # picks up ANTHROPIC_API_KEY / GEMINI_API_KEY from .env if present

# When True (orchestrator default), suppress banners/tqdm/API retry chatter but still
# emit short per-step "— done" progress lines (NAIP, Nearmap, classify, zoom, etc.).
QUIET = False


def _out(msg: str = "", *, important: bool = False) -> None:
    """Print unless QUIET; important=True always prints (stage/result progress)."""
    if QUIET and not important:
        return
    print(msg, flush=True)


def _env_flag(name: str, default: str = "0") -> bool:
    return os.environ.get(name, default).strip().lower() in ("1", "true", "yes")


NEARMAP_TIERED = _env_flag("NEARMAP_TIERED", default="1")
BIFURCATED_AI = _env_flag("BIFURCATED_AI", default="1")
GEMINI_ONLY = _env_flag("GEMINI_ONLY")
TOWER_ONLY = _env_flag("TOWER_ONLY")
NAIP_ONLY = _env_flag("NAIP_ONLY")
ZOOM_STAGE = _env_flag("ZOOM_STAGE", default="1")
# Widen NAIP AOI and re-classify when primary pass is other/unclear
# (towers just outside the 250 m frame). One extra Gemini call — cheaper than zoom.
WIDE_AOI_STAGE = _env_flag("WIDE_AOI_STAGE", default="1")
TIER_CONF_HIGH = float(os.environ.get("TIER_CONF_HIGH", "0.75"))
TIER_CONF_MEDIUM = float(os.environ.get("TIER_CONF_MEDIUM", "0.6"))
# Rooftop cell_equipment_confidence bar for early-stop / Claude skip (matches
# enrichment MIN_ROOFTOP_CELL_CONFIDENCE). Below this → keep fetching / escalate.
ROOFTOP_CELL_CONF_MIN = float(os.environ.get("ROOFTOP_CELL_CONF_MIN", "0.75"))
# Stale NAIP cannot early-stop Nearmap for rooftops (equipment may post-date the chip),
# unless confidence is at/above this override (definitive NAIP-only identification).
NAIP_MAX_AGE_YEARS = float(os.environ.get("NAIP_MAX_AGE_YEARS", "2"))
NAIP_AGE_HIGH_CONF_OVERRIDE = float(os.environ.get("NAIP_AGE_HIGH_CONF_OVERRIDE", "0.85"))
GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-3-flash-preview")
# Cheap NAIP screen; promote to GEMINI_MODEL for tower/rooftop/unclear/weak other.
GEMINI_SCREEN_MODEL = os.environ.get(
    "GEMINI_SCREEN_MODEL", "gemini-3.5-flash-lite"
)
# Gemini 3.x thinking_level: NAIP screen LOW; Nearmap/confirm MEDIUM.
_THINKING_LEVELS = frozenset({"MINIMAL", "LOW", "MEDIUM", "HIGH"})


def _parse_thinking_level(raw: str | None, default: str) -> str:
    text = (raw or "").strip().upper()
    return text if text in _THINKING_LEVELS else default


GEMINI_SCREEN_THINKING_LEVEL = _parse_thinking_level(
    os.environ.get("GEMINI_SCREEN_THINKING_LEVEL"), "LOW"
)
GEMINI_THINKING_LEVEL = _parse_thinking_level(
    os.environ.get("GEMINI_THINKING_LEVEL"), "MEDIUM"
)
CLAUDE_ESCALATION_MODEL = os.environ.get(
    "CLAUDE_ESCALATION_MODEL", "claude-sonnet-4-6")
# Dual-model cell crops use Haiku; Sonnet stays on rooftop HVAC localize.
CLAUDE_CROP_MODEL = os.environ.get(
    "CLAUDE_CROP_MODEL", "claude-haiku-4-5-20251001"
)
# Thinking tokens for Gemini vision calls. Empty/auto: 1024 for non-lite
# Gemini 2.5/3.x, 0 for *-lite* (keeps Flash-Lite cheap/fast).
GEMINI_THINKING_BUDGET_ENV = os.environ.get("GEMINI_THINKING_BUDGET", "").strip()
# Soft-keep / solo-trust cell confidence floors.
GEMINI_SOFT_KEEP_CELL_CONF = float(os.environ.get("GEMINI_SOFT_KEEP_CELL_CONF", "0.85"))
# Skip Claude (escalation + tower dual-model) when Gemini site_confidence is
# at/above this. Rooftop HVAC FPs still go through Claude dual-confirm.
GEMINI_SOLO_CELL_CONF = float(os.environ.get("GEMINI_SOLO_CELL_CONF", "0.90"))
# Do not burn Claude on weak Gemini — below this site_confidence, skip
# full-scene escalation and dual-model Claude. Weak calls stay holdout.
# 0.60 matches the NAIP scout floor so imagery-only towers at 0.65 still get Claude.
CLAUDE_ESCALATE_MIN_SITE_CONF = float(
    os.environ.get("CLAUDE_ESCALATE_MIN_SITE_CONF", "0.60")
)
# Min IoU between Gemini and Claude boxes on localize dual-confirm.
GEMINI_CLAUDE_BOX_IOU = float(os.environ.get("GEMINI_CLAUDE_BOX_IOU", "0.20"))


def _csv_oblique_views(raw: str, default: tuple[str, ...] = ("North", "East")) -> list[str]:
    allowed = {"North", "East", "South", "West"}
    parts = [p.strip().title() for p in str(raw or "").split(",") if p.strip()]
    views = [p for p in parts if p in allowed]
    return views or list(default)


OBLIQUE_VIEWS = _csv_oblique_views(
    os.environ.get("NEARMAP_OBLIQUE_VIEWS", "North,East")
)

# ----------------------------- configuration --------------------------------

STAC_URL = "https://planetarycomputer.microsoft.com/api/stac/v1"
COLLECTION = "naip"
# Primary NAIP chip side length (meters). Set to 500 to skip a prior 250 m pass
# and go straight to wide-frame classify (then ZOOM_STAGE if still other/unclear).
CHIP_SIZE_M = float(os.environ.get("CHIP_SIZE_M", "250"))
# Wide-AOI (zoom-out) retry size when primary NAIP pass finds no tower.
# Only runs when NAIP_WIDE_CHIP_M > CHIP_SIZE_M.
NAIP_WIDE_CHIP_M = float(os.environ.get("NAIP_WIDE_CHIP_M", "500"))
# Primary model first; hops to the next on persistent rate limits or 404.
# claude-sonnet-4-20250514 was retired 2026-06-15; use current IDs from:
# https://docs.anthropic.com/en/docs/about-claude/models/overview
_default_models = "claude-sonnet-4-6,claude-haiku-4-5-20251001"
MODELS = [
    m.strip() for m in os.environ.get("CLAUDE_MODELS", _default_models).split(",")
    if m.strip()
]
_model_idx = 0
API_DELAY_S = float(os.environ.get("CLAUDE_DELAY_S", "12"))
GEMINI_DELAY_S = float(os.environ.get("GEMINI_DELAY_S", "30"))
GEMINI_RETRIES = int(os.environ.get("GEMINI_RETRIES", "6"))
GEMINI_RETRY_BASE_S = float(os.environ.get("GEMINI_RETRY_BASE_S", "20"))
INPUT_CSV = "data/assets.csv"    # columns: id; lat+lon OR address; optional: label, input_confidence
INPUT_CONFIDENCE_LEVELS = ("high", "medium", "low")
OUTPUT_CSV = "results.csv"
CHIP_DIR = Path("chips")
RUNS_DIR = Path("runs")
RUN_DIR = Path(".")

# Geocoding: shared with orchestrator ingest via ingest/geocoder.py
# (US Census first, Nominatim fallback). Configure with GEOCODER=auto in .env.
# Optional Nearmap integration (Tile API).
# When NEARMAP_API_KEY is set, each asset also gets a high-res top-down view
# plus 45-degree oblique panoramas from the four compass directions - obliques
# show the vertical sides of structures, which is what makes rooftop antennas
# and towers actually visible. Without the key the pipeline runs NAIP-only.
# The Tile API bills against the subscription's monthly GB allowance; the
# Transactional Content API was tried first but needs a separate credits
# add-on (the coverage/v2/tx call returns 403 on this subscription).
NEARMAP_API_KEY = os.environ.get("NEARMAP_API_KEY")
NEARMAP_TILE_URL = "https://api.nearmap.com/tiles/v3/{content}/{z}/{x}/{y}.jpg"
NEARMAP_COVERAGE_POINT_URL = "https://api.nearmap.com/coverage/v2/point/{lon},{lat}"
NEARMAP_VIEWS = ["Vert", *OBLIQUE_VIEWS]
NEARMAP_CHIP_M = 100       # side length of the Nearmap AOI, in meters
# Wide-AOI fallback: rural sites often have vert-only Nearmap coverage and the
# recorded coordinates can put the real asset outside the narrow AOI. When the
# first pass can't identify a site and no obliques were available, the Nearmap
# fetch is retried at this size (matching the NAIP chip) and re-classified.
NEARMAP_FALLBACK_CHIP_M = 250
NEARMAP_VERT_ZOOM = int(os.environ.get("NEARMAP_VERT_ZOOM", "20"))
NEARMAP_OBLIQUE_ZOOM = int(os.environ.get("NEARMAP_OBLIQUE_ZOOM", "20"))
NEARMAP_MAX_PX = int(os.environ.get("NEARMAP_MAX_PX", "1024"))
# Downscale copies sent to vision models (saved chips stay at NEARMAP_MAX_PX).
MODEL_IMAGE_MAX_PX = int(os.environ.get("MODEL_IMAGE_MAX_PX", "768"))
MODEL_MAX_OBLIQUES = int(os.environ.get("MODEL_MAX_OBLIQUES", "2"))
# Lite NAIP screen uses a smaller image than Flash confirm.
SCREEN_IMAGE_MAX_PX = int(os.environ.get("SCREEN_IMAGE_MAX_PX", "768"))

# Two-stage zoom: after primary + wide-AOI passes still return other/unclear,
# scout suspicious regions on the best top-down image, magnify them, and
# re-classify. Critical for rural sites where towers are tiny in wide chips.
ZOOM_GRID = 3              # 3x3 grid fallback when scout finds nothing
ZOOM_MAX_CANDIDATES = int(os.environ.get("ZOOM_MAX_CANDIDATES", "3"))
ZOOM_OUTPUT_PX = 1024      # magnified crop size in pixels
ZOOM_MIN_FRAC = 0.10       # minimum crop side as fraction of source image
ZOOM_PAD_FRAC = 0.15       # padding around each candidate box
# Dual-model / cell-recheck crops need more context than zoom scout. 15% pad
# clips panel arrays and Claude then votes false on foliage / tank rim / HVAC.
CELL_CONFIRM_PAD_FRAC = 0.40
# Rooftop gear boxes are intentionally tight (sector panels). Do not reuse
# ZOOM_MIN_FRAC here — that rejects valid antenna boxes as "invalid".
ASSET_BOX_MIN_FRAC = 0.03  # ~30/1000 normalized
ASSET_BOX_MAX_SIDE = 500   # reject whole-roof / whole-scene boxes
EXECUTIVE_SUMMARY_MD = "EXECUTIVE_SUMMARY.md"

TOWER_SUBTYPE_SCHEMA_VALUES = [
    "monopole",
    "guyed",
    "self_support",
    "stealth",
    "steeple",
    "water_tower",
    "silo",
    "flagpole",
    "smokestack",
    "other_tower",
    "unclear",
]

CLASSIFICATION_PROMPT = """\
You are analyzing aerial imagery of one location where a cellular-infrastructure \
asset is expected. One or more views are provided, each preceded by a text label:
- "NAIP top-down": wide straight-down chip (~250 m across, ~1 m resolution). \
The recorded coordinates can be off by tens of meters, so the asset may appear \
ANYWHERE in this chip, not just at the center.
- "Nearmap top-down": recent high-resolution (~7 cm) straight-down view of the \
same location, usually covering a smaller area than the NAIP chip.
- "Nearmap oblique (North/East/South/West)": 45-degree angled views of the same \
location. These reveal the vertical sides of structures - towers, masts, and \
rooftop antennas that are nearly invisible from straight above stand out \
clearly here. Weight them heavily in every task.

Definitions:
- TOWER SITE: a ground-based, purpose-built vertical structure carrying \
antennas - monopole, lattice/self-support tower, guyed mast, or a disguised \
mast (monopalm / monopine / canister shroud). Top-down cues: tiny footprint, \
long thin linear shadow, lattice cross-pattern, guy wires, small \
cleared/fenced compound with equipment cabinets, or a palm/pine that is \
far taller and straighter than its neighbors. Oblique cues: a tall thin \
structure rising far above its surroundings — including a faux palm/pine.
- ROOFTOP SITE: a building whose roof hosts cellular telecom equipment. True \
cues: panel antennas / sector frames at roof corners or edges (often 3 sectors), \
triangular or steel antenna mounts, microwave backhaul dishes, RRUs, telecom \
cabinets with cable trays, short masts or poles carrying panel antennas on the \
parapet. Ordinary building mechanicals alone are NOT a rooftop cell site.
- STEALTH TREE / DISGUISED MAST: a purpose-built telecom mast camouflaged as \
a palm (monopalm) or pine (monopine), or a slim canister/flagpole shroud. \
Cues: unnaturally straight/thick trunk, crown taller and more regular than \
nearby real trees, antenna cylinder or panel bulge in the fronds, equipment \
cabinets or a fenced pad at the base. Often sits at a lot edge, behind a \
billboard, or in a corner of the chip — search edges. Natural palms/pines \
have tapered trunks, irregular fronds, and no cabinets; those are not towers.
- STEALTH / BUILDING-TOWER SITE: a structure that looks like a building but \
has a tall narrow tower section - church steeple, clock tower, faux-building \
monopole, or a tower segment rising from one corner of a larger footprint. \
From above: a compact building with an unusually tall shadow from one corner \
or a square tower block on the roofline; antennas may sit on the tower cap.

Perform three tasks:

TASK 1 - site_type. Search the ENTIRE extent of EVERY view - edges and corners \
included, never just the center - and classify the site:
- The recorded pin is often tens of meters off. If the center is a parking lot, \
empty pavement, driveway, or landscaping next to a larger commercial / mall \
building, search adjacent rooftops and tower compounds in the chip for cellular \
gear — do not stop at "other" just because the pin itself is pavement.
- "tower": a PURPOSE-BUILT cellular/telecom tower is visible. Must show cell \
platforms, sector racks, microwave dishes, a fenced telecom compound, or a \
disguised mast (monopalm/monopine/canister) — NOT a wood utility pole, street \
light, traffic signal, power transmission lattice, or a natural tree. A \
too-tall straight palm/pine with a regular crown and cabinets or an antenna \
bulge is stealth, not "other". Power/transmission lattices and bare utility \
poles are "other".
- "rooftop": no tower present, and a building roof hosts (or most plausibly \
hosts) the equipment.
- "other": neither applies (water tank, silo, bare field, power lattice, \
utility pole without cell racks, etc.) - describe it.
- "unclear": image quality or ambiguity prevents a confident call.
When site_type is "tower", also set tower_subtype to the best match:
- "monopole": single thin pole, minimal footprint, no lattice faces
- "guyed": guy wires or anchor pads visible from above or oblique
- "self_support": lattice or solid self-supporting tower legs
- "stealth": ONLY a purpose-built disguised telecom mast — e.g. monopalm \
(faux palm), monopine (faux pine), slim flagpole/canister shroud, or a \
freestanding faux-steeple mast with visible antenna bays/slots. A palm- or \
pine-disguised mast is stealth, never flagpole. Do NOT use stealth for \
ordinary church steeples, cupolas, elevator penthouses, decorative \
parking-lot towers, building corners, real trees, or architecture that \
merely "could" hide antennas. If unsure, use other_tower or unclear — never \
guess stealth.
- "water_tower": elevated tank on legs
- "silo": agricultural/industrial silo hosting antennas
- "flagpole": slim mast with a flag or ball finial and no foliage disguise \
and no RF canister/shroud. A cylindrical telecom shroud is stealth; faux \
fronds/needles are stealth.
- "smokestack": industrial stack
- "other_tower": tower present but subtype unclear
- "unclear": tower confirmed but subtype not discernible
When site_type is not "tower", set tower_subtype to null.
Never set cell_equipment=true on a stealth call from "likely conceals" or \
"typical for stealth" alone. For monopalm/monopine, true when you see \
equipment cabinets or a fenced pad at the base, a cylindrical antenna bay \
or panel bulge in the crown, or panels/dishes through the fronds — do not \
require exposed sector racks like a bare monopole. Real trees with no \
cabinets or bulge stay cell_equipment=false.
Set site_confidence to at most 0.6 unless two or more independent cues or \
views corroborate the call.

TASK 2 - locate the asset / cellular gear. Report:
- asset_box_2d: [ymin, xmin, ymax, xmax], integers in 0-1000 normalized image \
coordinates for ONE view only.
- asset_view: the exact label of that same view (e.g. "Nearmap oblique (South)").
Boxing rules (critical):
- Draw the box TIGHTLY around the cellular hardware you are claiming \
(sector panels, dishes, RRUs, parapet masts, facade mounts) — not the whole \
building, whole roof, parking deck, or HVAC plant.
- Prefer the Nearmap oblique where those antennas are clearest. Use Nearmap \
top-down only when gear is unmistakable from above. Prefer NAIP only when no \
Nearmap view shows the gear.
- The box must match the view named in asset_view; never reuse one box across \
views. Leave both fields null only if no asset/gear can be located.
- Keep the box compact: typically 3–25% of the image on each side \
unless the tower compound truly fills more. Never return inverted \
coordinates (ymin must be < ymax, xmin < xmax).

TASK 3 - cell_equipment: is cellular telecom equipment visible on the located asset?
- true: clear visible evidence of cellular gear; false: none visible; null: cannot \
assess (resolution or viewing angle insufficient). Prefer null over true when unsure.
TRUE only for cellular/telecom hardware, for example:
- Panel / sector antennas (flat vertical rectangles on frames, often ~3 sectors) \
on the roof, parapet, or building facade / wall (side-mounted)
- Microwave / backhaul dishes, RRUs, cable trays feeding antenna mounts
- Short masts or poles on the parapet carrying antennas
- Rooftop or wall-mounted radio panels / sector frames (not HVAC condensers)
- Telecom equipment cabinets clearly paired with antenna mounts (not alone)
- Monopalm/monopine/canister mast with cabinets or a fenced pad at the base, \
or an antenna cylinder/panel bulge in the faux crown
Default when unsure: cell_equipment=null (not true). Do not mark cell_equipment \
true from Nearmap Vert / top-down alone unless sector panels or dishes are \
unmistakable; prefer an oblique confirmation. Never set cell_equipment=true \
and never draw asset_box_2d around HVAC condensers, chillers, vents, or \
generic mechanical clusters.
FALSE / NOT cellular when they are the ONLY roof objects (HVAC and cell often \
coexist — do not call false just because condensers/vents are also present):
- HVAC condensers, chillers, air handlers, cooling towers alone
- Roof vents, pipes, stacks, drains, skylights, access hatches alone
- Solar panels, satellite TV dishes for building use, water tanks alone
- Random dark rectangles in neat rows across a roof (typical multi-zone HVAC) \
with no parapet sector frames, dishes, or antenna masts elsewhere on the roof
On rooftops, gear is often missed when it sits in **building shadow** or along \
shaded parapets, or when it sits next to HVAC. In oblique views, inspect sunlit \
AND shaded roof edges, corners, and mechanical zones before calling false. \
Top-down-only HVAC clusters must never be called cellular by themselves — but \
HVAC next to panel antennas is still cell_equipment=true.

Also set cell_gear_kind to exactly one of:
sector_panel | facade_mount | microwave | rru | parapet_mast | none | unclear
Use none when cell_equipment is false; unclear only when you cannot tell.

Field meanings: site_evidence and cell_equipment_evidence are one short \
sentence each, citing the specific views and cues used. For cell_equipment true, \
evidence must name the antenna/dish/mount cue, or for stealth the cabinets/\
shroud/crown bulge — not "equipment on roof". \
site_evidence describes the host structure only; it must NOT claim cellular \
gear unless cell_equipment is also true. If cell_equipment is false/null, \
site_evidence must not say "cellular site" / "sector panels" / "antenna mounts".
"""

TOWER_ONLY_CLASSIFICATION_PROMPT = """\
You are analyzing NAIP top-down aerial imagery of one location where a \
ground-based cellular TOWER may exist. Detection mode is TOWER ONLY.

The recorded coordinates can be off by tens of meters, so a tower may appear \
ANYWHERE in the chip, not just at the center.

Definitions:
- TOWER: a ground-based, purpose-built vertical structure carrying antennas - \
monopole, lattice/self-support tower, guyed mast, stealth monopalm/monopine/\
canister, stealth steeple/clock tower, water tower on legs, silo, flagpole, \
or smokestack with telecom gear. Top-down cues: tiny footprint, long thin \
shadow, lattice cross-pattern, guy wires, small fenced compound with \
equipment cabinets, or a palm/pine far taller and straighter than neighbors.
- NOT A TOWER: rooftop cellular hosts, bare fields, unrelated buildings, parking \
lots, or structures with no vertical tower mast. Classify those as "other".

Perform three tasks:

TASK 1 - site_type. Search the ENTIRE chip - edges and corners included:
- "tower": a tower (as defined above) is visible anywhere in the imagery
- "other": no tower visible (including rooftop sites and non-telecom structures)
- "unclear": image quality or ambiguity prevents a confident call
When site_type is "tower", also set tower_subtype to the best match:
monopole | guyed | self_support | stealth | steeple | water_tower | silo | \
flagpole | smokestack | other_tower | unclear
When site_type is not "tower", set tower_subtype to null.
Set site_confidence to at most 0.6 unless two or more independent cues \
corroborate a tower.

TASK 2 - locate the tower. If site_type is tower, report asset_box_2d \
[ymin, xmin, ymax, xmax] in 0-1000 normalized coordinates on the clearest view, \
plus asset_view. Otherwise set both to null.

TASK 3 - cell_equipment: is cellular equipment visible on the located tower?
true | false | null

Field meanings: site_evidence and cell_equipment_evidence are one short \
sentence each, citing the specific cues used.
"""

TOWER_ONLY_SCAN_PROMPT = """\
You are reviewing a NAIP top-down image where a ground-based cellular tower \
may exist, but a first pass could not identify it. Search the ENTIRE image for \
tower cues: tiny footprints with long shadows, lattice cross-patterns, guy \
wires, fenced compounds, monopoles, or stealth tower blocks on a building corner.

Return ONLY JSON with a "candidates" array (up to four entries). Each entry needs:
- box_2d: [ymin, xmin, ymax, xmax] in 0-1000 normalized coordinates
- reason: one short phrase citing the visual cue

Ignore rooftop-only hosts. If nothing looks like a tower, return an empty array.
"""

TOWER_ONLY_ZOOM_CLASSIFICATION_PROMPT = """\
You are performing a SECOND-PASS tower review on magnified NAIP zoom crops from \
a site that was not identified in wide imagery. A ground-based cellular tower \
may still be present.

Use the zoom crops as primary evidence. Look for lattice masts, monopoles, guyed \
structures, fenced compounds, stealth tower sections, or long thin tower shadows.

Detection mode is TOWER ONLY:
1. site_type: tower | other | unclear (rooftop hosts are "other")
2. When site_type is tower, tower_subtype: monopole | guyed | self_support | \
stealth | steeple | water_tower | silo | flagpole | smokestack | other_tower | unclear
3. asset_box_2d + asset_view on the clearest view when tower is found
4. cell_equipment: true | false | null

Set site_confidence to at most 0.7 unless zoom crops show an unambiguous tower.
"""

INPUT_CONFIDENCE_PROMPTS = {
    "high": (
        "\n\nSOURCE TRUST: HIGH. This is an outreach-verified / carrier-claimed "
        "site — an active cellular asset is likely present somewhere in the "
        "chip, often tens of meters off the pin and sometimes in a corner or "
        "lot edge. Search EVERY view including edges. Expect rooftop gear, a "
        "ground mast, or stealth (monopalm / monopine / canister) mixed with "
        "real trees. Do not call other just because the pin is pavement or "
        "the center building looks empty. Prefer null over false when "
        "imagery is ambiguous — but never mark HVAC or ordinary roof "
        "mechanicals as cell_equipment true."
    ),
    "medium": (
        "\n\nSOURCE TRUST: MEDIUM. The coordinate likely points to a cellular "
        "site but may be approximate. Weight oblique views when assessing "
        "rooftop equipment in shadow. Do not confuse HVAC with cellular gear."
    ),
    "low": (
        "\n\nSOURCE TRUST: LOW. The coordinate is exploratory; apply normal "
        "evidence standards. Prefer null/false over true for cell_equipment "
        "when cues are ambiguous."
    ),
}

EQUIPMENT_RECHECK_PROMPT = """\
You are re-checking ONLY for visible cellular equipment at a trusted \
outreach-verified site. The first pass called cell_equipment false, but an \
asset is expected here.

Re-examine every view — especially Nearmap obliques, chip edges/corners, \
vacant-lot margins, and tree clusters:
- Thin rectangular panel antennas on parapets, short masts, or building walls
- Sector frames at roof corners (often three sectors) or facade mounts
- Microwave dishes, RRUs, cable trays, telecom cabinets paired with mounts
- A monopalm/monopine (too-tall straight palm/pine with a regular crown, \
cabinets or antenna bulge) or a canister/flagpole shroud — not a natural tree

Do NOT flip to true for HVAC condensers, vents, pipes, skylights, drains, \
rows of identical mechanical units, or ordinary street palms. Prefer null \
over true when ambiguous.

Return the same JSON schema. If any clear cellular equipment or disguised \
mast is visible, set cell_equipment true and explain which view shows it.
"""

EQUIPMENT_FALSE_POSITIVE_RECHECK_PROMPT = """\
You previously set cell_equipment=true on a rooftop. Re-check carefully.

IMPORTANT: HVAC and cellular gear VERY OFTEN share the same roof. Seeing \
condensers, vents, pipes, or mechanical screens does NOT mean there is no \
cellular equipment. Look past the HVAC plant for telecom hardware.

Keep cell_equipment=true when ANY of these are visible on the roof OR building \
sides/facades (edges, corners, parapets, walls, and behind/beside HVAC) — use \
oblique views for side-mounted gear:
- Panel / sector antennas (flat vertical rectangles on frames, often ~3 sectors)
- Wall- or facade-mounted radio panels / sector frames
- Microwave / backhaul dishes, RRUs, cable trays to antenna mounts
- Short masts or poles carrying antennas
- Telecom cabinets clearly feeding antenna mounts

Set cell_equipment=false ONLY when you can rule out telecom hardware and the \
visible objects are explained entirely by ordinary mechanicals (HVAC-only, \
vents, drains, skylights, solar for building use). Say so explicitly in \
cell_equipment_evidence (e.g. "no panel antennas or dishes; HVAC only").

Set cell_equipment=null if resolution/angle cannot separate HVAC from antennas.

Also set cell_gear_kind to sector_panel | facade_mount | microwave | rru | \
parapet_mast | none | unclear.

Return the same JSON schema. Evidence must state whether antennas/dishes are \
present or absent — do not flip to false merely because HVAC is visible.
"""

ROOFTOP_CELL_CROP_RECHECK_PROMPT = """\
You are inspecting a MAGNIFIED crop PLUS the source view of a rooftop/facade \
region previously flagged as having cellular equipment.

Decide whether cellular telecom hardware is visible in the crop OR the source \
view:
- true: sector panels, facade mounts, dishes, RRUs, or parapet masts
- false: ONLY when the boxed region is clearly HVAC/vents/pipes/solar/skylights \
with no telecom gear in the crop or the source view
- null: crop is too tight, blurry, or foliage-obscured — do NOT vote false

If the crop clips the array, use the source view. Prefer null over false when \
unsure. Set cell_gear_kind accordingly.
Return the same JSON schema.
"""

ROOFTOP_CELL_DUAL_CONFIRM_PROMPT = """\
Independent confirmation pass: does this site show CLEAR cellular equipment on \
a rooftop or facade (sector panels, facade mounts, dishes, RRUs, parapet masts)?

Answer cell_equipment true only with unmistakable telecom hardware. HVAC-only \
roofs are false. Set cell_gear_kind and cite the view in cell_equipment_evidence.
Return the same JSON schema.
"""

ROOFTOP_CELL_CROP_CONFIRM_PROMPT = """\
You are checking a magnified crop PLUS the source view of a rooftop/facade \
region another model boxed as cellular gear.

TRUE when the crop OR the source view shows sector panels, facade mounts, \
microwave dishes, RRUs, or parapet masts. FALSE only when that boxed region is \
clearly HVAC/vents/pipes/solar/skylights with no telecom gear in either view.

If the crop is too tight, blurry, or foliage-obscured, use the source view. \
If still unsure, set cell_equipment null (not false) and cell_gear_kind=unclear. \
Leave asset_box_2d null — the prior box is fixed.
Return the same JSON schema.
"""

ROOFTOP_CELL_LOCALIZE_CONFIRM_PROMPT = """\
Independent localization pass: does this rooftop/facade show CLEAR cellular \
equipment anywhere in the views (not only a prior crop)?

If cell_equipment=true you MUST draw your own tight asset_box_2d on the Nearmap \
oblique where the gear is clearest, and set asset_view to that exact view label. \
Never box HVAC alone. Do not vote false merely because a prior crop was too \
tight — search the full obliques. Prefer null over false when unsure.
Return the same JSON schema.
"""

TOWER_CELL_CROP_CONFIRM_PROMPT = """\
You are checking a magnified crop PLUS the source view of a suspected telecom \
TOWER (not a rooftop HVAC cluster).

Vote cell_equipment true if the crop OR the source view shows a purpose-built \
telecom tower with cellular/microwave gear: sector panels, RRUs, microwave \
dishes, or an antenna platform/array. Lattice/self-support, monopole, guyed, \
stealth (monopine/palm/canister), water tower, silo, flagpole, and smokestack \
hosts WITH antennas are towers — never treat those as HVAC.

If the crop is foliage, a tank rim, the mast base, or otherwise inconclusive \
but the source view shows the tower and its antenna array, vote true from the \
source view. A monopalm/monopine crop that shows the faux crown, cabinets, or \
antenna bulge is still a tower — do not vote false as "just a tree". \
Satellite-only earth-station dishes without cellular arrays → false. \
Leave asset_box_2d null — the prior box is fixed.
Return the same JSON schema.
"""

TOWER_CELL_LOCALIZE_CONFIRM_PROMPT = """\
Independent confirmation of a candidate GROUND-BASED telecom tower (not a \
rooftop HVAC check). Search EVERY view, including edges and corners — \
monopalm/monopine masts are often at a lot edge or in a corner cluster of \
real trees.

A tower may be a monopole, lattice/self-support, guyed mast, stealth \
monopine/palm/canister, water tower, silo, flagpole, or smokestack with \
telecom gear.

1. site_type: "tower" if a purpose-built tower/mast is visible; "rooftop" only \
if there is no tower and a building hosts gear; "other" if neither. A \
too-tall straight palm/pine with a regular synthetic crown is a stealth \
tower, not a natural tree.
2. cell_equipment: true when cellular or microwave gear is visible on that \
tower (sector panels, RRUs, dishes, antenna platforms/arrays, side-arm \
mounts, a canister/antenna bulge in faux fronds, or equipment cabinets / \
fenced pad at a disguised mast). Satellite-only earth-station dishes without \
cellular arrays → false. HVAC, natural trees, bare tanks, and vehicles are \
not cell gear — but do not dismiss a monopalm/monopine as a tree.
3. If cell_equipment=true, draw a tight asset_box_2d on the clearest view \
(prefer the antenna array; the full mast is OK) and set asset_view to that \
view's exact label.

Do NOT classify a visible lattice/monopole/guyed/water/stealth tower as \
rooftop or other. Do NOT vote false merely because individual panels are \
small, hidden in fronds, or the top is slightly soft — platforms, arrays, \
shrouds, and cabinets count. Prefer false only when no tower is present, or \
the structure has no telecom gear.
Return the same JSON schema.
"""


CLAIMED_SITE_TOWER_CONFIRM_NOTE = """

CLAIMED SITE: an outreach-verified record expects an asset here. Search \
edges and corners for a monopalm/monopine/canister mixed with real trees. \
Do not vote cell_equipment false with "no tower exists" unless you have \
checked lot edges and tree clusters. Prefer null over false when a tall \
straight palm/pine or canister mast is visible but panels are hidden in \
the crown.
"""


def cell_confirm_prompt(site_type: str, *, used_crop: bool) -> str:
    """Prompt for dual-model cell confirm: tower vs rooftop, crop vs localize."""
    site = str(site_type or "").strip().lower()
    if site == "tower":
        return (
            TOWER_CELL_CROP_CONFIRM_PROMPT
            if used_crop
            else TOWER_CELL_LOCALIZE_CONFIRM_PROMPT
        )
    return (
        ROOFTOP_CELL_CROP_CONFIRM_PROMPT
        if used_crop
        else ROOFTOP_CELL_LOCALIZE_CONFIRM_PROMPT
    )

ROOFTOP_BOX_REPAIR_PROMPT = """\
Prior classification suggested a rooftop cellular site but did not return a \
usable asset_box_2d (missing, inverted, too small, or too large).

Your ONLY job: locate cellular gear and draw ONE valid box.

Rules:
- Prefer a Nearmap oblique (North/East/South/West) where sector panels, dishes, \
RRUs, or parapet masts are clearest. Do not box HVAC alone.
- asset_box_2d = [ymin, xmin, ymax, xmax] integers in 0-1000 with ymin < ymax \
and xmin < xmax. Draw TIGHTLY around the cellular hardware (typically 3–25% of \
the image on each side — never the whole roof or whole building).
- asset_view must be the exact label of that same view.
- If cellular gear is clearly visible: site_type=rooftop, cell_equipment=true, \
cell_gear_kind set, and both box fields filled.
- If no cellular gear is visible: cell_equipment=false, cell_gear_kind=none, \
leave asset_box_2d and asset_view null.
Return the same JSON schema.
"""

# Enforced via Claude tool input_schema so every reply parses into this shape.
RESPONSE_SCHEMA = {
    "type": "object",
    "properties": {
        "site_type": {
            "type": "string",
            "enum": ["tower", "rooftop", "other", "unclear"],
        },
        "tower_subtype": {
            "type": "string",
            "enum": list(TOWER_SUBTYPE_SCHEMA_VALUES),
        },
        "site_confidence": {"type": "number"},
        "site_evidence": {"type": "string"},
        "asset_box_2d": {
            "type": "array",
            "items": {"type": "integer"},
        },
        "asset_view": {"type": "string"},
        "cell_equipment": {"type": "boolean"},
        "cell_equipment_confidence": {"type": "number"},
        "cell_equipment_evidence": {"type": "string"},
        "cell_gear_kind": {
            "type": "string",
            "enum": [
                "sector_panel",
                "facade_mount",
                "microwave",
                "rru",
                "parapet_mast",
                "none",
                "unclear",
            ],
        },
    },
    "required": ["site_type", "site_confidence", "site_evidence"],
}


# Gemini SDK uses a distinct schema dialect; kept in sync with RESPONSE_SCHEMA.
GEMINI_RESPONSE_SCHEMA = {
    "type": "OBJECT",
    "properties": {
        "site_type": {
            "type": "STRING",
            "enum": ["tower", "rooftop", "other", "unclear"],
        },
        "tower_subtype": {
            "type": "STRING",
            "enum": TOWER_SUBTYPE_SCHEMA_VALUES,
            "nullable": True,
        },
        "site_confidence": {"type": "NUMBER"},
        "site_evidence": {"type": "STRING"},
        "asset_box_2d": {
            "type": "ARRAY",
            "items": {"type": "INTEGER"},
            "nullable": True,
        },
        "asset_view": {"type": "STRING", "nullable": True},
        "cell_equipment": {"type": "BOOLEAN", "nullable": True},
        "cell_equipment_confidence": {"type": "NUMBER"},
        "cell_equipment_evidence": {"type": "STRING"},
        "cell_gear_kind": {
            "type": "STRING",
            "enum": [
                "sector_panel",
                "facade_mount",
                "microwave",
                "rru",
                "parapet_mast",
                "none",
                "unclear",
            ],
            "nullable": True,
        },
    },
    "required": ["site_type", "site_confidence", "site_evidence"],
}


def _site_type_values() -> list[str]:
    if TOWER_ONLY:
        return ["tower", "other", "unclear"]
    return ["tower", "rooftop", "other", "unclear"]


def _positive_site_types() -> tuple[str, ...]:
    if TOWER_ONLY:
        return ("tower",)
    return ("tower", "rooftop")


def _classification_response_schemas() -> tuple[dict, dict]:
    site_types = _site_type_values()
    claude_schema = {
        **RESPONSE_SCHEMA,
        "properties": {
            **RESPONSE_SCHEMA["properties"],
            "site_type": {"type": "string", "enum": site_types},
        },
    }
    gemini_schema = {
        **GEMINI_RESPONSE_SCHEMA,
        "properties": {
            **GEMINI_RESPONSE_SCHEMA["properties"],
            "site_type": {"type": "STRING", "enum": site_types},
        },
    }
    return claude_schema, gemini_schema


def resolve_ai_mode() -> tuple[str, bool]:
    """Return (primary_provider, allow_claude_escalation)."""
    if GEMINI_ONLY:
        return "gemini", False
    if BIFURCATED_AI:
        return "gemini", True
    return "claude", False


SCAN_PROMPT = """\
You are reviewing a single top-down aerial image where a cellular tower or \
rooftop site is expected, but a first-pass classifier could not identify it. \
The asset may be anywhere in the frame and is often subtle: a small lattice \
mast, monopole shadow, fenced compound, rooftop antenna cluster, a stealth \
tower disguised as a building with a tall tower section (steeple, clock tower, \
faux-building cell site), or a monopalm/monopine (too-tall straight palm/pine \
in a lot corner or tree cluster). Check the area just below image center \
especially when coordinates are approximate.

Search the ENTIRE image - especially edges and corners - and return up to four \
candidate regions that could plausibly be a tower site or rooftop cellular host. \
Prioritize: tiny footprints with long shadows, lattice cross-patterns, fenced \
pads with equipment cabinets, a palm/pine that is taller and straighter than \
its neighbors, or building roofs with sector-frame mounts.

Return ONLY JSON with a "candidates" array. Each entry needs:
- box_2d: [ymin, xmin, ymax, xmax] in 0-1000 normalized coordinates
- reason: one short phrase citing the visual cue

If nothing looks plausible, return an empty candidates array.
"""

SCAN_SCHEMA = {
    "type": "object",
    "properties": {
        "candidates": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "box_2d": {
                        "type": "array",
                        "items": {"type": "integer"},
                    },
                    "reason": {"type": "string"},
                },
                "required": ["box_2d", "reason"],
            },
        },
    },
    "required": ["candidates"],
}

GEMINI_SCAN_SCHEMA = {
    "type": "OBJECT",
    "properties": {
        "candidates": {
            "type": "ARRAY",
            "items": {
                "type": "OBJECT",
                "properties": {
                    "box_2d": {
                        "type": "ARRAY",
                        "items": {"type": "INTEGER"},
                    },
                    "reason": {"type": "STRING"},
                },
                "required": ["box_2d", "reason"],
            },
        },
    },
    "required": ["candidates"],
}

ZOOM_CLASSIFICATION_PROMPT = """\
You are performing a SECOND-PASS review on magnified zoom crops from a site \
that was not identified in wide imagery. A cellular tower or rooftop site is \
still expected at this location.

One or more "Zoom crop" views are provided - each is a magnified section of a \
top-down image. Also included may be the original wide "NAIP top-down" or \
"Nearmap top-down" view for context.

Use the zoom crops as primary evidence. A tower site often appears as a \
lattice mast, monopole, guyed structure, small fenced compound with \
equipment, or a monopalm/monopine (straight trunk, regular crown, cabinets \
at the base). A stealth site may also be a building with a tall tower block \
or steeple on one corner and a long shadow from that tower section. A \
rooftop site shows panel antennas, sector frames, or dishes on a building roof.

Perform the same three tasks as the primary classifier:
1. site_type: tower | rooftop | other | unclear
2. When site_type is tower, tower_subtype: monopole | guyed | self_support | \
stealth | steeple | water_tower | silo | flagpole | smokestack | other_tower | unclear
3. asset_box_2d + asset_view on the view where the asset is clearest
4. cell_equipment: true | false | null

Set site_confidence to at most 0.7 unless zoom crops show unambiguous equipment.
"""


def _active_scan_prompt() -> str:
    return TOWER_ONLY_SCAN_PROMPT if TOWER_ONLY else SCAN_PROMPT


def _active_zoom_prompt() -> str:
    return TOWER_ONLY_ZOOM_CLASSIFICATION_PROMPT if TOWER_ONLY else ZOOM_CLASSIFICATION_PROMPT


def source_expects_asset(res: dict | None = None, *, input_confidence: str | None = None) -> bool:
    """True when outreach-verified / high source trust says an asset is likely."""
    conf = input_confidence
    if conf is None and res is not None:
        conf = res.get("input_confidence")
    return normalize_input_confidence(conf) == "high"


def normalize_input_confidence(value) -> str:
    """Return high | medium | low. Missing or invalid values default to medium."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "medium"
    level = str(value).strip().lower()
    return level if level in INPUT_CONFIDENCE_LEVELS else "medium"


def normalize_confidence(value, default: float | None = None) -> float | None:
    """Clamp model confidence to 0-1. Values > 1 are treated as wrong-scale output."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return default
    try:
        v = float(value)
    except (TypeError, ValueError):
        return default
    if v > 1.0:
        if v <= 10.0:
            v /= 10.0
        elif v <= 100.0:
            v /= 100.0
        else:
            v = 1.0
    return max(0.0, min(1.0, v))


def normalize_model_result(res: dict) -> dict:
    """Fix site_confidence and cell_equipment_confidence after a model JSON reply."""
    if "site_confidence" in res:
        norm = normalize_confidence(res.get("site_confidence"))
        if norm is not None:
            res["site_confidence"] = norm
    if "cell_equipment_confidence" in res:
        norm = normalize_confidence(res.get("cell_equipment_confidence"))
        if norm is not None:
            res["cell_equipment_confidence"] = norm
    if res.get("site_type") != "tower":
        res["tower_subtype"] = None
    elif res.get("tower_subtype") in ("", "null"):
        res["tower_subtype"] = None
    kind = str(res.get("cell_gear_kind") or "").strip().lower().replace(" ", "_")
    allowed = {
        "sector_panel",
        "facade_mount",
        "microwave",
        "rru",
        "parapet_mast",
        "none",
        "unclear",
    }
    if kind in allowed:
        res["cell_gear_kind"] = kind
    elif res.get("cell_equipment") is True:
        res["cell_gear_kind"] = "unclear"
    elif res.get("cell_equipment") is False:
        res["cell_gear_kind"] = "none"
    return res


def build_classification_prompt(row) -> str:
    """Assemble the full classification prompt from base + label + source trust."""
    prompt = TOWER_ONLY_CLASSIFICATION_PROMPT if TOWER_ONLY else CLASSIFICATION_PROMPT
    label_hint = str(row.get("label", "")).strip().lower()
    if label_hint == "stealth":
        prompt += (
            "\n\nNOTE: This site is tagged STEALTH. Expect a disguised mast "
            "(monopalm, monopine, canister shroud) or a tower integrated into "
            "a building (steeple, clock tower, faux facade, corner tower "
            "block). Do not dismiss a too-tall palm/pine, a tall narrow "
            "shadow, or a tower block as merely a tree or architecture "
            "unless clearly non-telecom."
        )
    prompt += INPUT_CONFIDENCE_PROMPTS[normalize_input_confidence(
        row.get("input_confidence"))]
    return prompt


def maybe_recheck_equipment(provider: str, clients: dict, res: dict, views: list,
                            input_confidence: str) -> dict:
    """Second pass when a trusted source expects gear but the model said false."""
    if not source_expects_asset(res, input_confidence=input_confidence):
        return res
    if res.get("cell_equipment") is True:
        return res
    site = str(res.get("site_type") or "").strip().lower()
    # Claimed-site empties and false cell calls get a second look.
    if res.get("cell_equipment") is not False and site not in {"other", "unclear"}:
        return res
    if len(views) < 2:
        return res
    recheck = classify_site(
        provider, clients, views, prompt=EQUIPMENT_RECHECK_PROMPT)
    if recheck.get("cell_equipment") is True:
        res["cell_equipment"] = True
        res["cell_equipment_confidence"] = recheck.get(
            "cell_equipment_confidence", res.get("cell_equipment_confidence"))
        res["cell_equipment_evidence"] = recheck.get(
            "cell_equipment_evidence", res.get("cell_equipment_evidence"))
        if recheck.get("site_type") in _positive_site_types():
            res["site_type"] = recheck["site_type"]
            res["site_confidence"] = recheck.get(
                "site_confidence", res.get("site_confidence"))
            res["site_evidence"] = recheck.get(
                "site_evidence", res.get("site_evidence"))
        normalize_model_result(res)
        _step_done("equipment recheck", _brief_pass_result(res))
    else:
        _step_done("equipment recheck", "still no cell gear")
    return res


_STEALTH_FORM_CUES = (
    "monopalm",
    "mono-palm",
    "monopine",
    "mono-pine",
    "faux palm",
    "faux-pine",
    "faux pine",
    "palm frond",
    "pine frond",
    "canister",
    "antenna bay",
    "antenna bays",
    "stealth canister",
    "pine-needle",
    "pine needle",
    "faux frond",
    "synthetic frond",
    "artificial palm",
    "artificial pine",
    "disguised as a palm",
    "disguised as a pine",
    "palm-style",
    "pine-style",
    "palm tower",
    "pine tower",
    "fake palm",
    "fake pine",
)

_STEALTH_HARDWARE_CUES = (
    "cabinet",
    "cabinets",
    "compound",
    "fenced pad",
    "equipment pad",
    "shroud",
    "canister",
    "antenna bulge",
    "antenna bay",
    "antenna bays",
)


def _text_has_telecom_cue(text: str) -> bool:
    lowered = (text or "").lower()
    cues = (
        "antenna",
        "sector",
        "rru",
        "microwave",
        "backhaul",
        "panel",
        "parapet mast",
        "radio",
        "telecom",
    )
    return any(cue in lowered for cue in cues)


def _text_has_stealth_form(text: str) -> bool:
    lowered = (text or "").lower()
    if any(cue in lowered for cue in _STEALTH_FORM_CUES):
        return True
    # Pairing "palm"/"pine" with stealth is specific; "palm"+"tower" is not
    # (street palms next to a real monopole).
    if "stealth" in lowered and ("palm" in lowered or "pine" in lowered):
        return True
    return False


def _text_has_stealth_hardware(text: str) -> bool:
    lowered = (text or "").lower()
    return any(cue in lowered for cue in _STEALTH_HARDWARE_CUES)


def _text_denies_telecom(text: str) -> bool:
    """True when evidence explicitly rules out cellular hardware (not just HVAC)."""
    lowered = (text or "").lower()
    deny_phrases = (
        "no sector",
        "no panel",
        "no antenna",
        "no cellular",
        "no telecom",
        "no microwave",
        "no rru",
        "hvac only",
        "only hvac",
        "mechanicals only",
        "no evidence of panel",
        "no identifiable cellular",
        "rather than cellular",
        "not cellular",
        "none of the components match",
        "without any distinct",
    )
    return any(phrase in lowered for phrase in deny_phrases)


def maybe_recheck_rooftop_cell_false_positive(
    provider: str, clients: dict, res: dict, views: list,
    *, prior_nearmap_no_cell: bool = False,
) -> dict:
    """Downgrade rooftop cell=true only when telecom is explicitly ruled out.

    HVAC commonly coexists with cell gear — presence of mechanical plant alone
    must not clear a positive cell_equipment call.

    Always re-check rooftop cell=true: Gemini often self-cites "antenna/sector"
    on HVAC, so keyword skip is unsafe.
    """
    if str(res.get("site_type") or "").strip().lower() != "rooftop":
        return res
    if res.get("cell_equipment") is not True:
        return res
    if not _rooftop_cell_confirmed(res):
        return res
    if not views:
        return res

    prior_evidence = " ".join(
        str(res.get(key) or "")
        for key in ("cell_equipment_evidence", "site_evidence")
    )
    if prior_nearmap_no_cell:
        _step_done("cell FP recheck", "forced (prior Nearmap no-cell)")

    recheck = classify_site(
        provider, clients, views, prompt=EQUIPMENT_FALSE_POSITIVE_RECHECK_PROMPT
    )
    new_cell = recheck.get("cell_equipment")
    recheck_evidence = str(recheck.get("cell_equipment_evidence") or "")

    if new_cell is True:
        res["cell_equipment_confidence"] = recheck.get(
            "cell_equipment_confidence", res.get("cell_equipment_confidence")
        )
        res["cell_equipment_evidence"] = recheck.get(
            "cell_equipment_evidence", res.get("cell_equipment_evidence")
        )
        normalize_model_result(res)
        _step_done("cell FP recheck", "confirmed cellular")
        return res

    if new_cell is False and not _text_denies_telecom(recheck_evidence):
        # Soft false: only keep coexistence true when the first pass already
        # cited antennas/panels AND we have a compact Nearmap oblique box.
        # Self-authored Gemini "sector" prose alone is not enough.
        if not (
            _text_has_telecom_cue(prior_evidence)
            and has_locked_oblique_asset_box(res)
        ):
            res["cell_equipment"] = False
            if "cell_equipment_confidence" in recheck:
                res["cell_equipment_confidence"] = recheck.get(
                    "cell_equipment_confidence"
                )
            res["cell_equipment_evidence"] = recheck_evidence or (
                "FP recheck: no locked oblique box; soft-false treated as HVAC"
            )
            res["asset_box_2d"] = None
            res["asset_view"] = None
            normalize_model_result(res)
            _step_done(
                "cell FP recheck",
                "downgraded to false (no locked oblique box)",
            )
            return res
        if recheck_evidence:
            res["cell_equipment_evidence"] = (
                f"{prior_evidence.strip()} | FP recheck kept true (coexistence): "
                f"{recheck_evidence}"
            ).strip(" |")
        normalize_model_result(res)
        _step_done("cell FP recheck", "kept true (HVAC+cell coexistence)")
        return res

    if new_cell is None and _text_has_telecom_cue(prior_evidence) and (
        has_locked_oblique_asset_box(res)
    ):
        # Ambiguous second pass should not erase a prior antenna citation
        # when we already locked an oblique gear box.
        _step_done("cell FP recheck", "kept true (prior antenna cues + oblique box)")
        return res

    res["cell_equipment"] = new_cell  # false with hard deny, or null
    if "cell_equipment_confidence" in recheck:
        res["cell_equipment_confidence"] = recheck.get("cell_equipment_confidence")
    if recheck_evidence:
        res["cell_equipment_evidence"] = recheck_evidence
    if new_cell is not True:
        # Drop misleading HVAC boxes when cell is no longer confirmed.
        if not has_locked_oblique_asset_box(res):
            res["asset_box_2d"] = None
            res["asset_view"] = None
    normalize_model_result(res)
    if new_cell is False:
        _step_done("cell FP recheck", "downgraded to false")
    else:
        _step_done("cell FP recheck", "downgraded to null")
    return res


_PROMOTE_TO_STEALTH_SUBTYPES = frozenset(
    {"flagpole", "other_tower", "monopole", "unclear"}
)


def gate_weak_stealth_tower_claim(res: dict) -> dict:
    """Demote speculative stealth tower calls (common Gemini false positive).

    Stealth requires a purpose-built disguise (monopalm/monopine/canister).
    Architecture "likely conceals" guesses are demoted. A flagpole/monopole
    described as a palm/pine/canister mast is promoted to stealth so it is
    not uploaded as Flagpole.
    """
    if str(res.get("site_type") or "").strip().lower() != "tower":
        return res

    evidence = " ".join(
        str(res.get(key) or "")
        for key in ("cell_equipment_evidence", "site_evidence")
    ).lower()
    strong_form = _text_has_stealth_form(evidence)
    subtype = str(res.get("tower_subtype") or "").strip().lower()
    if subtype in _PROMOTE_TO_STEALTH_SUBTYPES and strong_form:
        res["tower_subtype"] = "stealth"
        subtype = "stealth"
        prior = str(res.get("site_evidence") or "").strip()
        note = "stealth gate: disguised mast (palm/pine/canister)"
        res["site_evidence"] = f"{prior} | {note}".strip(" |")
        _step_done("stealth gate", note)

    if subtype != "stealth":
        return res

    speculative = any(
        phrase in evidence
        for phrase in (
            "likely conceal",
            "likely hides",
            "typical for",
            "as is typical",
            "probably conceal",
            "appears to conceal",
            "may conceal",
            "could conceal",
        )
    )
    has_cues = _text_has_telecom_cue(evidence) or _text_has_stealth_hardware(
        evidence
    )

    # Keep a described monopalm/monopine/canister. Speculative architecture
    # without those forms still gets demoted.
    if strong_form and not speculative:
        return res
    if strong_form and has_cues:
        return res

    prior = str(res.get("site_evidence") or "").strip()
    res["tower_subtype"] = "other_tower"
    note = "stealth gate: speculative or weak disguise cues"
    if speculative or not strong_form:
        # Do not keep cell=true on architecture-as-stealth guesses.
        if res.get("cell_equipment") is True and (speculative or not has_cues):
            res["cell_equipment"] = None
            res["cell_gear_kind"] = "unclear"
            res["cell_equipment_evidence"] = (
                f"{res.get('cell_equipment_evidence') or ''} | {note}"
            ).strip(" |")
    res["site_evidence"] = f"{prior} | {note}".strip(" |")
    normalize_model_result(res)
    _step_done("stealth gate", note)
    return res


def gate_weak_rooftop_cell_claim(res: dict) -> dict:
    """Downgrade untrustworthy rooftop cell=true claims before crop/dual-model.

    Requires telecom evidence cues and a compact Nearmap localization box.
    Oblique boxes are preferred; Vert-only true needs high conf + cues.
    Clears misleading HVAC boxes when the claim is demoted.
    """
    if str(res.get("site_type") or "").strip().lower() != "rooftop":
        return res
    if res.get("cell_equipment") is not True:
        return res

    evidence = " ".join(
        str(res.get(key) or "")
        for key in ("cell_equipment_evidence", "site_evidence")
    )
    has_cues = _text_has_telecom_cue(evidence)
    oblique_box = has_locked_oblique_asset_box(res)
    valid = get_valid_asset_box(res)
    view = str(res.get("asset_view") or "").strip().lower()
    vert_box = bool(valid) and (
        "top-down" in view or ("vert" in view and "oblique" not in view)
    )
    conf = normalize_confidence(res.get("cell_equipment_confidence")) or 0.0

    if oblique_box and has_cues:
        return res
    if vert_box and has_cues and conf >= 0.85:
        return res

    prior = str(res.get("cell_equipment_evidence") or "").strip()
    reason_bits = []
    if not has_cues:
        reason_bits.append("no telecom cues")
    if not oblique_box and not vert_box:
        reason_bits.append("no usable Nearmap box")
    elif vert_box and conf < 0.85:
        reason_bits.append("Vert box needs conf>=0.85 or oblique")
    note = "first-pass gate: " + (", ".join(reason_bits) or "weak cell claim")
    res["cell_equipment"] = None
    res["cell_equipment_evidence"] = f"{prior} | {note}".strip(" |")
    res["cell_gear_kind"] = "unclear"
    res["asset_box_2d"] = None
    res["asset_view"] = None
    res["dual_model_resolution"] = res.get("dual_model_resolution") or "first_pass_gate"
    normalize_model_result(res)
    _step_done("first-pass cell gate", note)
    return res


def maybe_recheck_rooftop_cell_crop(
    provider: str, clients: dict, res: dict, views: list
) -> dict:
    """Magnify the asset box and re-ask cell vs HVAC on the crop alone."""
    if str(res.get("site_type") or "").strip().lower() != "rooftop":
        return res
    if res.get("cell_equipment") is not True:
        return res
    valid = get_valid_asset_box(res)
    if not valid:
        _step_done("cell crop recheck", "skipped (invalid box)")
        return res
    # Persist coerced geometry so later stages share the same box.
    res["asset_box_2d"] = valid

    picked = pick_view_for_asset_box(res, views)
    if picked is None:
        _step_done("cell crop recheck", "skipped (no source image)")
        return res
    source_label, source_img = picked
    # Do not silently crop NAIP when the model claimed an oblique box.
    if asset_view_is_nearmap_oblique(res.get("asset_view")) and (
        "naip" in str(source_label).lower()
    ):
        _step_done("cell crop recheck", "skipped (oblique box view missing)")
        return res

    crop = _crop_zoom(source_img, valid, pad_frac=CELL_CONFIRM_PAD_FRAC)
    crop_views = [
        (f"zoom crop ({source_label})", crop),
        (source_label, source_img),
    ]
    recheck = classify_site(
        provider, clients, crop_views, prompt=ROOFTOP_CELL_CROP_RECHECK_PROMPT
    )
    new_cell = recheck.get("cell_equipment")
    if new_cell is True:
        if recheck.get("cell_equipment_confidence") is not None:
            res["cell_equipment_confidence"] = recheck.get("cell_equipment_confidence")
        if recheck.get("cell_equipment_evidence"):
            res["cell_equipment_evidence"] = (
                f"{res.get('cell_equipment_evidence') or ''} | crop: "
                f"{recheck.get('cell_equipment_evidence')}"
            ).strip(" |")
        if recheck.get("cell_gear_kind"):
            res["cell_gear_kind"] = recheck.get("cell_gear_kind")
        normalize_model_result(res)
        _step_done("cell crop recheck", f"confirmed cellular on {source_label}")
        return res

    # Do not hard-downgrade on a tight/inconclusive crop — dual-model localize
    # still gets a chance. Keep Gemini cell=true + box.
    note = str(recheck.get("cell_equipment_evidence") or "crop inconclusive")
    res["cell_equipment_evidence"] = (
        f"{res.get('cell_equipment_evidence') or ''} | crop recheck kept "
        f"for dual-model ({note})"
    ).strip(" |")
    _step_done("cell crop recheck", "inconclusive — keep for dual-model")
    return res


def _oblique_views_only(views: list) -> list:
    """Prefer Nearmap oblique chips for box repair / cell localization."""
    obliques = []
    for label, img in views:
        lower = str(label).lower()
        if "naip" in lower:
            continue
        if "oblique" in lower or any(
            d in lower for d in ("north", "east", "south", "west")
        ):
            obliques.append((label, img))
    return obliques


def rooftop_box_is_usable(res: dict, views: list | None = None) -> bool:
    """True when asset_box_2d is geometrically valid and matches a real view."""
    if not get_valid_asset_box(res):
        return False
    view = str(res.get("asset_view") or "").strip()
    if not view:
        return False
    if views is None:
        return True
    picked = pick_view_for_asset_box(res, views)
    if picked is None:
        return False
    label, _img = picked
    if asset_view_is_nearmap_oblique(view) and "naip" in str(label).lower():
        return False
    return True


def maybe_repair_rooftop_asset_box(
    provider: str, clients: dict, res: dict, views: list
) -> dict:
    """Re-ask for a tight oblique box when rooftop localization is incomplete.

    Runs when site_type=rooftop and either cell=true or a strong rooftop call
    lacks a usable box. Without a box, crop + dual-model cannot be trustworthy.
    """
    if str(res.get("site_type") or "").strip().lower() != "rooftop":
        return res
    if rooftop_box_is_usable(res, views):
        # Persist coerced geometry.
        valid = get_valid_asset_box(res)
        if valid:
            res["asset_box_2d"] = valid
        return res

    conf = normalize_confidence(res.get("site_confidence")) or 0.0
    needs_repair = res.get("cell_equipment") is True or conf >= 0.7
    if not needs_repair:
        _step_done("box repair", "skipped (weak rooftop, no cell claim)")
        return res

    repair_views = _oblique_views_only(views)
    if not repair_views:
        _step_done("box repair", "skipped (no Nearmap obliques)")
        return res

    repair = classify_site(
        provider, clients, repair_views, prompt=ROOFTOP_BOX_REPAIR_PROMPT
    )
    normalize_model_result(repair)

    # Always absorb a usable repaired box.
    if coerce_asset_box(repair.get("asset_box_2d")) and repair.get("asset_view"):
        trial = dict(res)
        trial["asset_box_2d"] = repair.get("asset_box_2d")
        trial["asset_view"] = repair.get("asset_view")
        if rooftop_box_is_usable(trial, views) or rooftop_box_is_usable(
            trial, repair_views
        ):
            res["asset_box_2d"] = coerce_asset_box(repair.get("asset_box_2d"))
            res["asset_view"] = repair.get("asset_view")
            if repair.get("cell_equipment") is True:
                res["cell_equipment"] = True
                if repair.get("cell_equipment_confidence") is not None:
                    res["cell_equipment_confidence"] = repair.get(
                        "cell_equipment_confidence"
                    )
                if repair.get("cell_equipment_evidence"):
                    res["cell_equipment_evidence"] = repair.get(
                        "cell_equipment_evidence"
                    )
                if repair.get("cell_gear_kind"):
                    res["cell_gear_kind"] = repair.get("cell_gear_kind")
            elif repair.get("cell_equipment") is False:
                res["cell_equipment"] = False
                if repair.get("cell_equipment_evidence"):
                    res["cell_equipment_evidence"] = repair.get(
                        "cell_equipment_evidence"
                    )
                res["cell_gear_kind"] = repair.get("cell_gear_kind") or "none"
            normalize_model_result(res)
            _step_done(
                "box repair",
                f"repaired on {res.get('asset_view')} cell={res.get('cell_equipment')!r}",
            )
            return res

    if repair.get("cell_equipment") is False:
        res["cell_equipment"] = False
        if repair.get("cell_equipment_evidence"):
            res["cell_equipment_evidence"] = repair.get("cell_equipment_evidence")
        res["cell_gear_kind"] = "none"
        res["asset_box_2d"] = None
        res["asset_view"] = None
        normalize_model_result(res)
        _step_done("box repair", "no cellular gear — cell=false")
        return res

    _step_done("box repair", "failed (still no usable box)")
    return res


def enforce_rooftop_cell_requires_box(res: dict, views: list | None = None) -> dict:
    """Rooftop cell=true is incomplete without a usable localization box."""
    if str(res.get("site_type") or "").strip().lower() != "rooftop":
        return res
    if res.get("cell_equipment") is not True:
        return res
    if rooftop_box_is_usable(res, views):
        valid = get_valid_asset_box(res)
        if valid:
            res["asset_box_2d"] = valid
        return res
    res["cell_equipment"] = None
    res["cell_models_agree"] = False
    res["dual_model_resolution"] = "box_required"
    prior = str(res.get("cell_equipment_evidence") or "").strip()
    note = "Cell claim cleared: no usable asset_box_2d on a matching view."
    res["cell_equipment_evidence"] = f"{prior} | {note}".strip(" |")
    normalize_model_result(res)
    _step_done("box gate", "cell cleared (box required)")
    return res


def box_iou(box_a: list[int] | None, box_b: list[int] | None) -> float:
    """Intersection-over-union for [ymin, xmin, ymax, xmax] boxes in 0-1000 space."""
    a = coerce_asset_box(box_a)
    b = coerce_asset_box(box_b)
    if not a or not b:
        return 0.0
    ay0, ax0, ay1, ax1 = a
    by0, bx0, by1, bx1 = b
    iy0, ix0 = max(ay0, by0), max(ax0, bx0)
    iy1, ix1 = min(ay1, by1), min(ax1, bx1)
    iw, ih = max(0, ix1 - ix0), max(0, iy1 - iy0)
    inter = iw * ih
    if inter <= 0:
        return 0.0
    area_a = max(0, ax1 - ax0) * max(0, ay1 - ay0)
    area_b = max(0, bx1 - bx0) * max(0, by1 - by0)
    union = area_a + area_b - inter
    if union <= 0:
        return 0.0
    return inter / union


def confirm_rooftop_cell_with_claude(
    res: dict,
    clients: dict,
    views: list,
    *,
    already_escalated: bool,
    allow_soft_keep: bool = True,
    from_wide_rescue: bool = False,
    used_crop: bool = False,
    allow_gemini_solo: bool = True,
    all_views: list | None = None,
) -> tuple[dict, str | None, bool]:
    """Tiered Claude check for rooftop/tower cell=true.

    1) Strong Gemini (optional) → skip Claude when allow_gemini_solo.
    2) usable Gemini crop → Claude votes true/false on THAT box only.
       A crop no is not final for rooftops: fall back to full-scene localize
       so a too-tight crop cannot unqualify a real array. Clear HVAC after
       localize still vetoes.
    3) No crop → Claude localizes; cell=true requires Claude box + IoU.

    Enrichment auto-apply should pass allow_soft_keep=False and
    allow_gemini_solo=False so rooftop HVAC FPs still need Claude hard-agree.
    Towers with Gemini site_confidence >= GEMINI_SOLO_CELL_CONF skip Claude
    even when allow_gemini_solo is False (Gemini already locked the tower).
    """
    site = str(res.get("site_type") or "").strip().lower()
    if site not in {"rooftop", "tower"}:
        return res, None, False

    def _apply_soft_keep(reason: str) -> tuple[dict, str | None, bool]:
        res["cell_equipment"] = True
        pre_conf = res.get("gemini_pre_escalation_cell_conf")
        pre_ev = res.get("gemini_pre_escalation_evidence")
        pre_gear = res.get("gemini_pre_escalation_gear")
        if pre_conf is not None:
            res["cell_equipment_confidence"] = pre_conf
        elif res.get("cell_equipment_confidence") is None:
            pass
        if pre_ev:
            res["cell_equipment_evidence"] = (
                f"{pre_ev} | Claude veto soft-kept (Nearmap+Gemini)"
            )
        if pre_gear:
            res["cell_gear_kind"] = pre_gear
        res["gemini_cell_equipment"] = True
        res["cell_models_agree"] = True
        res["dual_model_resolution"] = "soft_keep_gemini"
        normalize_model_result(res)
        _step_done("dual-model cell", reason)
        return res, "claude", True

    def _soft_keep_candidate() -> dict:
        trial = dict(res)
        trial["cell_equipment"] = True
        if res.get("gemini_pre_escalation_cell_conf") is not None:
            trial["cell_equipment_confidence"] = res.get(
                "gemini_pre_escalation_cell_conf"
            )
        if res.get("gemini_pre_escalation_evidence"):
            trial["cell_equipment_evidence"] = res.get(
                "gemini_pre_escalation_evidence"
            )
        return trial

    def _apply_claude_veto(claude_res: dict, *, reason: str) -> tuple[dict, str | None, bool]:
        # Crop veto: never soft-keep the same boxed region Claude rejected.
        can_claimed = (
            not used_crop
            and should_keep_claimed_gemini_tower(
                res, from_wide_rescue=from_wide_rescue
            )
        )
        if can_claimed:
            if gemini_evidence or claude_res.get("cell_equipment_evidence"):
                res["cell_equipment_evidence"] = (
                    f"{gemini_evidence or ''} | claimed-site keep Gemini "
                    f"(Claude missed disguised mast); Claude: "
                    f"{claude_res.get('cell_equipment_evidence') or 'cell=false'}"
                ).strip(" |")
            res["cell_equipment"] = True
            pre_conf = res.get("gemini_pre_escalation_cell_conf")
            pre_gear = res.get("gemini_pre_escalation_gear")
            if pre_conf is not None:
                res["cell_equipment_confidence"] = pre_conf
            if pre_gear:
                res["cell_gear_kind"] = pre_gear
            res["gemini_cell_equipment"] = True
            res["claude_cell_equipment"] = claude_res.get("cell_equipment")
            res["cell_models_agree"] = False
            res["dual_model_resolution"] = "claimed_site_keep_gemini"
            normalize_model_result(res)
            _step_done(
                "dual-model cell",
                "claimed-site keep Gemini (stealth/canister mast)",
            )
            return res, "claude", False
        can_soft = (
            allow_soft_keep
            and not used_crop
            and should_soft_keep_gemini_cell(
                res, from_wide_rescue=from_wide_rescue
            )
        )
        if can_soft:
            if gemini_evidence or claude_res.get("cell_equipment_evidence"):
                res["cell_equipment_evidence"] = (
                    f"{gemini_evidence or ''} | Claude veto soft-kept "
                    f"(Nearmap+Gemini); Claude: "
                    f"{claude_res.get('cell_equipment_evidence') or 'cell=false'}"
                ).strip(" |")
            return _apply_soft_keep(
                "disagree → soft-keep Gemini (Nearmap+Gemini)"
            )
        res["cell_equipment"] = claude_res.get("cell_equipment")
        if claude_res.get("cell_equipment_evidence"):
            res["cell_equipment_evidence"] = claude_res.get(
                "cell_equipment_evidence"
            )
        if claude_res.get("cell_gear_kind"):
            res["cell_gear_kind"] = claude_res.get("cell_gear_kind")
        if "cell_equipment_confidence" in claude_res:
            res["cell_equipment_confidence"] = claude_res.get(
                "cell_equipment_confidence"
            )
        if not has_locked_oblique_asset_box(res):
            res["asset_box_2d"] = None
            res["asset_view"] = None
        res["dual_model_resolution"] = "claude_veto"
        res["cell_models_agree"] = False
        normalize_model_result(res)
        _step_done("dual-model cell", reason)
        return res, "claude", False

    if res.get("cell_equipment") is not True:
        res["claude_cell_equipment"] = res.get("cell_equipment")
        res["gemini_cell_equipment"] = res.get("gemini_pre_escalation_cell")
        if (
            already_escalated
            and allow_soft_keep
            and res.get("gemini_pre_escalation_cell") is True
            and should_soft_keep_gemini_cell(
                _soft_keep_candidate(), from_wide_rescue=from_wide_rescue
            )
        ):
            return _apply_soft_keep(
                "disagree → soft-keep Gemini (Nearmap+Gemini)"
            )
        res["cell_models_agree"] = False
        return res, ("claude" if already_escalated else None), False

    if site == "rooftop" and not _has_nearmap_context(res):
        res["claude_cell_equipment"] = None
        res["cell_models_agree"] = False
        _step_done("dual-model cell", "skipped (NAIP-only rooftop)")
        return res, None, False

    gemini_cell = True
    res["gemini_cell_equipment"] = True
    gemini_conf = res.get("cell_equipment_confidence")
    gemini_evidence = res.get("cell_equipment_evidence")
    gemini_gear = res.get("cell_gear_kind")
    gemini_box = get_valid_asset_box(res)
    gemini_view = str(res.get("asset_view") or "").strip()
    # Preserve Gemini snapshot for soft-keep if not already stamped.
    res.setdefault("gemini_pre_escalation_cell", True)
    res.setdefault("gemini_pre_escalation_cell_conf", gemini_conf)
    res.setdefault("gemini_pre_escalation_evidence", gemini_evidence)
    res.setdefault("gemini_pre_escalation_gear", gemini_gear)

    if already_escalated:
        agree = res.get("cell_equipment") is True
        res["claude_cell_equipment"] = res.get("cell_equipment")
        res["cell_models_agree"] = agree
        if agree:
            res["dual_model_resolution"] = res.get("dual_model_resolution") or "agree"
        _step_done(
            "dual-model cell",
            "agree" if agree else "disagree (Claude negative)",
        )
        return res, "claude", agree

    if should_skip_claude_for_gemini_tower(
        res, from_wide_rescue=from_wide_rescue
    ):
        res["claude_cell_equipment"] = None
        res["cell_models_agree"] = True
        res["dual_model_resolution"] = "gemini_strong_solo"
        normalize_model_result(res)
        _step_done(
            "dual-model cell",
            f"skipped Claude (Gemini tower conf>={GEMINI_SOLO_CELL_CONF:g})",
        )
        return res, "gemini_strong_solo", True

    site_conf = normalize_confidence(res.get("site_confidence"))
    if site_conf is None or site_conf < CLAUDE_ESCALATE_MIN_SITE_CONF:
        res["claude_cell_equipment"] = None
        res["cell_models_agree"] = False
        _step_done(
            "dual-model cell",
            f"skipped Claude (Gemini site conf<{CLAUDE_ESCALATE_MIN_SITE_CONF:g})",
        )
        return res, None, False

    if allow_gemini_solo and should_trust_gemini_cell_solo(
        res, from_wide_rescue=from_wide_rescue
    ):
        res["claude_cell_equipment"] = None
        res["cell_models_agree"] = True
        res["dual_model_resolution"] = "gemini_strong_solo"
        normalize_model_result(res)
        _step_done(
            "dual-model cell",
            f"skipped Claude (Gemini solo conf>={GEMINI_SOLO_CELL_CONF:g})",
        )
        return res, "gemini_strong_solo", True

    if "claude" not in clients or clients.get("claude") is None:
        res["cell_models_agree"] = False
        _step_done("dual-model cell", "skipped (no Claude client)")
        return res, None, False

    if used_crop:
        prompt = cell_confirm_prompt(site, used_crop=True)
        mode = "crop_check"
    else:
        prompt = cell_confirm_prompt(site, used_crop=False)
        mode = "localize_iou"
    if site == "tower" and source_expects_asset(res):
        prompt += CLAIMED_SITE_TOWER_CONFIRM_NOTE

    claude_res = classify_site(
        "claude",
        clients,
        views,
        prompt=prompt,
        claude_model=CLAUDE_CROP_MODEL if used_crop else CLAUDE_ESCALATION_MODEL,
    )
    claude_cell = claude_res.get("cell_equipment") is True
    res["claude_cell_equipment"] = claude_res.get("cell_equipment")
    res["claude_dual_mode"] = mode

    if not claude_cell:
        loc_views = all_views or []
        if used_crop and site == "rooftop" and loc_views:
            loc_res = classify_site(
                "claude",
                clients,
                loc_views,
                prompt=ROOFTOP_CELL_LOCALIZE_CONFIRM_PROMPT,
                claude_model=CLAUDE_ESCALATION_MODEL,
            )
            loc_box = get_valid_asset_box(loc_res)
            if loc_res.get("cell_equipment") is True and loc_box:
                res["claude_cell_equipment"] = True
                res["claude_dual_mode"] = "crop_then_localize"
                if loc_res.get("cell_equipment_confidence") is not None:
                    res["cell_equipment_confidence"] = loc_res.get(
                        "cell_equipment_confidence"
                    )
                if loc_res.get("cell_equipment_evidence"):
                    res["cell_equipment_evidence"] = (
                        f"{gemini_evidence or ''} | crop inconclusive; "
                        f"localize: {loc_res.get('cell_equipment_evidence')}"
                    ).strip(" |")
                if loc_res.get("cell_gear_kind"):
                    res["cell_gear_kind"] = loc_res.get("cell_gear_kind")
                res["asset_box_2d"] = loc_box
                res["asset_view"] = loc_res.get("asset_view") or res.get("asset_view")
                res["cell_equipment"] = True
                res["cell_models_agree"] = True
                res["dual_model_resolution"] = "agree_localize"
                normalize_model_result(res)
                _step_done(
                    "dual-model cell",
                    "crop no → localize agree (tight crop fallback)",
                )
                return res, "claude", True
            claude_res = loc_res
            res["claude_cell_equipment"] = loc_res.get("cell_equipment")
        return _apply_claude_veto(
            claude_res,
            reason=(
                "crop+localize veto"
                if used_crop and site == "rooftop"
                else "crop veto (HVAC/not cell)"
                if used_crop
                else "localize veto"
            ),
        )

    # Claude says true.
    if used_crop:
        # Keep Gemini's box/coords; Claude only confirmed the crop contents.
        if claude_res.get("cell_equipment_confidence") is not None:
            res["cell_equipment_confidence"] = claude_res.get(
                "cell_equipment_confidence"
            )
        if claude_res.get("cell_equipment_evidence"):
            res["cell_equipment_evidence"] = claude_res.get(
                "cell_equipment_evidence"
            )
        if claude_res.get("cell_gear_kind"):
            res["cell_gear_kind"] = claude_res.get("cell_gear_kind")
        res["cell_equipment"] = True
        res["cell_models_agree"] = True
        res["dual_model_resolution"] = "agree_crop"
        normalize_model_result(res)
        _step_done("dual-model cell", "agree (Claude crop check)")
        return res, "claude", True

    # Localize path: Claude must produce a usable box.
    claude_box = get_valid_asset_box(claude_res)
    claude_view = str(claude_res.get("asset_view") or "").strip()
    if not claude_box:
        claude_res = dict(claude_res)
        claude_res["cell_equipment"] = False
        claude_res["cell_equipment_evidence"] = (
            str(claude_res.get("cell_equipment_evidence") or "")
            + " | localize: Claude true without usable box"
        ).strip(" |")
        return _apply_claude_veto(
            claude_res, reason="localize true without box"
        )

    if gemini_box and gemini_view and claude_view and site != "tower":
        same_view = (
            gemini_view.lower() in claude_view.lower()
            or claude_view.lower() in gemini_view.lower()
        )
        if same_view:
            iou = box_iou(gemini_box, claude_box)
            res["claude_gemini_box_iou"] = round(iou, 3)
            if iou < GEMINI_CLAUDE_BOX_IOU:
                claude_res = dict(claude_res)
                claude_res["cell_equipment"] = False
                claude_res["cell_equipment_evidence"] = (
                    f"{claude_res.get('cell_equipment_evidence') or ''} | "
                    f"box IoU {iou:.2f} < {GEMINI_CLAUDE_BOX_IOU:g} vs Gemini"
                ).strip(" |")
                return _apply_claude_veto(
                    claude_res, reason=f"box IoU {iou:.2f} mismatch"
                )

    # Towers: Gemini often boxes the full mast/compound; Claude boxes the
    # antenna array. Prefer Claude's tighter box. Rooftops keep Gemini when
    # IoU already aligned above; otherwise adopt Claude localization.
    if site == "tower" or not gemini_box:
        res["asset_box_2d"] = claude_box
        res["asset_view"] = claude_view or res.get("asset_view")
    if claude_res.get("cell_equipment_confidence") is not None:
        res["cell_equipment_confidence"] = claude_res.get(
            "cell_equipment_confidence"
        )
    if claude_res.get("cell_equipment_evidence"):
        res["cell_equipment_evidence"] = claude_res.get("cell_equipment_evidence")
    if claude_res.get("cell_gear_kind"):
        res["cell_gear_kind"] = claude_res.get("cell_gear_kind")
    res["cell_equipment"] = True
    res["cell_models_agree"] = True
    res["dual_model_resolution"] = "agree_localize"
    normalize_model_result(res)
    _step_done("dual-model cell", "agree (Claude localize)")
    return res, "claude", True


# ----------------------------- geocoding ------------------------------------

def _has_coordinates(row) -> bool:
    lat, lon = row.get("lat"), row.get("lon")
    if lat is None or lon is None:
        return False
    if isinstance(lat, float) and pd.isna(lat):
        return False
    if isinstance(lon, float) and pd.isna(lon):
        return False
    if str(lat).strip() == "" or str(lon).strip() == "":
        return False
    return True


def _clean_address(row) -> str | None:
    address = row.get("address")
    if address is None or (isinstance(address, float) and pd.isna(address)):
        return None
    text = str(address).strip()
    return text or None


def resolve_row_coordinates(row) -> tuple[float, float, dict]:
    """Return (lat, lon, geocode_metadata). metadata is empty when coords given."""
    if _has_coordinates(row):
        return float(row["lat"]), float(row["lon"]), {}

    address = _clean_address(row)
    if not address:
        raise ValueError(
            "each row needs lat+lon or a non-empty address column")

    try:
        from ingest.geocoder import geocode_address
    except ImportError as exc:
        raise ValueError(
            "address geocoding requires ingest.geocoder; pass lat+lon instead"
        ) from exc

    try:
        geo = geocode_address(address)
    except RuntimeError as exc:
        raise ValueError(str(exc)) from exc

    meta = {k: v for k, v in geo.items() if k not in ("lat", "lng", "lon")}
    meta["input_address"] = address
    lon = geo.get("lon", geo["lng"])
    return geo["lat"], lon, meta


def validate_input_csv(df: pd.DataFrame):
    """Ensure each row has id and either coordinates or an address."""
    if "id" not in df.columns:
        raise SystemExit(f"{INPUT_CSV} is missing required column: id")
    has_coords = "lat" in df.columns and "lon" in df.columns
    has_address = "address" in df.columns
    if not has_coords and not has_address:
        raise SystemExit(
            f"{INPUT_CSV} needs lat+lon columns, an address column, or both")

    missing = []
    for _, row in df.iterrows():
        if _has_coordinates(row):
            continue
        if _clean_address(row):
            continue
        missing.append(row["id"])
    if missing:
        raise SystemExit(
            f"{INPUT_CSV}: these rows have no lat/lon and no address: "
            f"{missing[:5]}{'...' if len(missing) > 5 else ''}")

# ----------------------------- imagery stage --------------------------------

_catalog = None

def get_catalog():
    global _catalog
    if _catalog is None:
        _catalog = Client.open(STAC_URL, modifier=planetary_computer.sign_inplace)
    return _catalog


def _naip_image_meta(item) -> dict:
    """Pull acquisition/refresh fields from NAIP STAC item properties."""
    acquired = item.datetime.date() if item.datetime else None
    props = item.properties or {}
    age_years = None
    if acquired is not None:
        age_years = round((date.today() - acquired).days / 365.25, 1)
    return {
        "image_date": acquired.isoformat() if acquired else None,
        "naip_year": props.get("naip:year"),
        "naip_state": props.get("naip:state"),
        "naip_gsd_m": props.get("gsd"),
        "image_age_years": age_years,
        "naip_chip_m": None,
    }


def fetch_chip(lat: float, lon: float, chip_m: float = CHIP_SIZE_M):
    """Return (PIL.Image, meta, geo) for the newest NAIP scene at a point, or
    (None, None, None) if no imagery covers the location.

    `meta` includes acquisition date / NAIP year / GSD from STAC photo metadata.
    `geo` holds the chip's CRS and projected bounds so a detection box drawn on
    the image can be converted back to real-world coordinates.
    """
    search = get_catalog().search(
        collections=[COLLECTION],
        intersects={"type": "Point", "coordinates": [lon, lat]},
    )
    items = sorted(search.items(), key=lambda i: i.datetime, reverse=True)
    if not items:
        return None, None, None

    item = items[0]
    href = item.assets["image"].href

    with rasterio.open(href) as src:
        # NAIP rasters are in UTM; project the WGS84 point into the raster CRS
        transformer = Transformer.from_crs("EPSG:4326", src.crs, always_xy=True)
        x, y = transformer.transform(lon, lat)
        half = chip_m / 2.0
        window = from_bounds(x - half, y - half, x + half, y + half, src.transform)
        # Read RGB bands only; boundless handles points near scene edges
        data = src.read([1, 2, 3], window=window, boundless=True, fill_value=0)
        geo = {"crs": str(src.crs),
               "x_min": x - half, "x_max": x + half,
               "y_min": y - half, "y_max": y + half,
               "chip_m": chip_m}

    img = Image.fromarray(np.transpose(data, (1, 2, 0)).astype(np.uint8))
    meta = _naip_image_meta(item)
    meta["naip_chip_m"] = chip_m
    return img, meta, geo


def _naip_view_label(chip_m: float | None = None) -> str:
    """Label used in the view list and for matching asset_view to geo."""
    if chip_m is not None and chip_m > CHIP_SIZE_M:
        return f"NAIP top-down (wide {int(chip_m)}m)"
    return "NAIP top-down"


def _is_naip_view(asset_view: str | None) -> bool:
    return bool(asset_view) and str(asset_view).startswith("NAIP top-down")


def box_to_latlon(geo: dict, box) -> tuple[float, float, float] | None:
    """Convert a [ymin, xmin, ymax, xmax] box in 0-1000 normalized image
    coordinates on the NAIP chip into (lat, lon, offset_m), where offset_m is
    the distance from the box center to the chip center (the input coordinate).
    Returns None if the box is malformed."""
    try:
        ymin, xmin, ymax, xmax = (float(v) for v in box[:4])
    except (TypeError, ValueError):
        return None
    if not (0 <= ymin <= ymax <= 1000 and 0 <= xmin <= xmax <= 1000):
        return None
    # Normalized box center -> projected coordinates (y axis is flipped:
    # image row 0 is the chip's northern edge / max projected y)
    cx_n = (xmin + xmax) / 2000.0
    cy_n = (ymin + ymax) / 2000.0
    x = geo["x_min"] + cx_n * (geo["x_max"] - geo["x_min"])
    y = geo["y_max"] - cy_n * (geo["y_max"] - geo["y_min"])
    to_wgs84 = Transformer.from_crs(geo["crs"], "EPSG:4326", always_xy=True)
    lon, lat = to_wgs84.transform(x, y)
    center_x = (geo["x_min"] + geo["x_max"]) / 2.0
    center_y = (geo["y_min"] + geo["y_max"]) / 2.0
    offset_m = math.hypot(x - center_x, y - center_y)
    return lat, lon, offset_m


def box_to_latlon_centered(
    lat: float, lon: float, chip_m: float, box
) -> tuple[float, float, float] | None:
    """Geocode a chip box when the AOI is a square centered on lat/lon.

    Used for Nearmap Vert and as an approximation for Nearmap obliques: the
    mosaic is still fetched around the same AOI, so the box center is a nearby
    pin (perspective error is usually tens of meters, not hundreds).
    """
    try:
        ymin, xmin, ymax, xmax = (float(v) for v in box[:4])
        side = float(chip_m)
    except (TypeError, ValueError):
        return None
    if side <= 0:
        return None
    if ymin > ymax:
        ymin, ymax = ymax, ymin
    if xmin > xmax:
        xmin, xmax = xmax, xmin
    if not (0 <= ymin < ymax <= 1000 and 0 <= xmin < xmax <= 1000):
        return None
    cx = (xmin + xmax) / 2000.0
    cy = (ymin + ymax) / 2000.0
    east_m = (cx - 0.5) * side
    south_m = (cy - 0.5) * side
    dlat = -south_m / 111_320.0
    cos_lat = math.cos(math.radians(lat))
    if abs(cos_lat) < 1e-6:
        return None
    dlon = east_m / (111_320.0 * cos_lat)
    offset_m = math.hypot(east_m, south_m)
    return lat + dlat, lon + dlon, offset_m


def locate_asset_box_latlon(
    *,
    lat: float,
    lon: float,
    box,
    box_view: str | None,
    naip_geo: dict | None = None,
    nearmap_aoi_m: float | None = None,
) -> tuple[float, float, float, str] | None:
    """Map asset_box_2d → (lat, lon, offset_m, source) when possible.

    Preference: true NAIP geo → Nearmap Vert → Nearmap oblique AOI approximation.
    """
    valid = coerce_asset_box(box)
    if valid is None and isinstance(box, (list, tuple)) and len(box) >= 4:
        # Allow slightly-oversized boxes through for geocode only.
        try:
            ymin, xmin, ymax, xmax = (int(round(float(v))) for v in box[:4])
            if ymin > ymax:
                ymin, ymax = ymax, ymin
            if xmin > xmax:
                xmin, xmax = xmax, xmin
            if 0 <= ymin < ymax <= 1000 and 0 <= xmin < xmax <= 1000:
                valid = [ymin, xmin, ymax, xmax]
        except (TypeError, ValueError):
            valid = None
    if not valid:
        return None

    if _is_naip_view(box_view) and naip_geo:
        located = box_to_latlon(naip_geo, valid)
        if located:
            return located[0], located[1], located[2], "naip_asset_box"

    chip_m = float(nearmap_aoi_m or NEARMAP_CHIP_M)
    view = str(box_view or "").lower()
    if "top-down" in view or ("vert" in view and "oblique" not in view):
        located = box_to_latlon_centered(lat, lon, chip_m, valid)
        if located:
            return located[0], located[1], located[2], "nearmap_vert_box"

    if asset_view_is_nearmap_oblique(box_view):
        located = box_to_latlon_centered(lat, lon, chip_m, valid)
        if located:
            return located[0], located[1], located[2], "nearmap_oblique_box"

    return None


def nearmap_full_blocks_rescue(
    res: dict, *, nearmap_tier: str, has_obliques: bool
) -> bool:
    """True when full Nearmap+obliques locked onto a specific HVAC-only rooftop.

    Only blocks wide/zoom when we have a compact Nearmap oblique asset box —
    that means we are on the right building and should not invent cell gear
    from a wider NAIP scout.

    Without a locked box (common when the SF pin sits in a parking lot or
    empty pavement next to a mall), wide/zoom + pin re-center must still run
    so a nearby rooftop/tower can be recovered.
    """
    if str(nearmap_tier or "").strip().lower() != "full" or not has_obliques:
        return False
    site = str(res.get("site_type") or "").strip().lower()
    if site != "rooftop" or res.get("cell_equipment") is not False:
        return False
    return has_locked_oblique_asset_box(res)


def has_locked_oblique_asset_box(res: dict) -> bool:
    """True when a compact Nearmap oblique box names a specific roof/host."""
    if not asset_view_is_nearmap_oblique(res.get("asset_view")):
        return False
    return get_valid_asset_box(res) is not None


def needs_pin_offset_scout(res: dict) -> bool:
    """True when a rooftop pin may have missed facade/parapet gear.

    Rooftops buy a Nearmap wide AOI on the host building. Other/unclear
    without Nearmap coverage use cheap NAIP wide/zoom instead.
    """
    if str(res.get("site_type") or "").strip().lower() != "rooftop":
        return False
    return res.get("cell_equipment") is not True


def scout_result_wins(prior: dict, scout: dict) -> bool:
    """True when a wide/zoom/address scout should replace the pin result.

    A rooftop/tower label is not enough: confidence must clear TIER_CONF_MEDIUM
    and must not drop versus an already-positive prior.
    """
    scout_site = str(scout.get("site_type") or "").strip().lower()
    prior_site = str(prior.get("site_type") or "").strip().lower()
    scout_conf = normalize_confidence(scout.get("site_confidence")) or 0.0
    prior_conf = normalize_confidence(prior.get("site_confidence")) or 0.0
    if scout_site in _positive_site_types():
        if scout_conf < TIER_CONF_MEDIUM:
            return False
        if prior_site in _positive_site_types() and scout_conf < prior_conf:
            return False
        return True
    return scout_conf > prior_conf


def needs_naip_rescue(res: dict) -> bool:
    """True when a weak NAIP other/unclear still warrants NAIP wide/zoom.

    Very low-confidence ``unclear`` (below TIER_CONF_MEDIUM) is not a signal —
    do not spend wide/zoom hunting an empty chip.
    """
    site = str(res.get("site_type") or "").strip().lower()
    conf = normalize_confidence(res.get("site_confidence"))
    if site == "unclear":
        return conf is not None and conf >= TIER_CONF_MEDIUM
    if site == "other":
        return not confident_no_asset(res)
    return False


def confident_no_asset(res: dict) -> bool:
    """True when Gemini already locked a high-confidence empty/other chip.

    Outreach-verified (high source trust) only locks empty at the Gemini
    solo bar (0.90). Weaker other calls still get zoom/Claude.
    """
    if str(res.get("site_type") or "").strip().lower() != "other":
        return False
    conf = normalize_confidence(res.get("site_confidence"))
    if conf is None:
        return False
    lock = (
        GEMINI_SOLO_CELL_CONF
        if source_expects_asset(res)
        else TIER_CONF_HIGH
    )
    return conf >= lock


def stamp_naip_screen(res: dict) -> dict:
    """Keep the NAIP-screen label after Nearmap overwrites site_type."""
    if res.get("naip_screen_site_type") not in (None, ""):
        return res
    res["naip_screen_site_type"] = res.get("site_type")
    res["naip_screen_site_confidence"] = res.get("site_confidence")
    res["naip_screen_cell_equipment"] = res.get("cell_equipment")
    return res


def needs_flash_confirm(res: dict) -> bool:
    """Promote the NAIP lite screen to Flash for anything that is not a confident other.

    Low-confidence ``unclear`` stays on the screen result — Flash will not invent
    a tower from a 0.1 chip.
    """
    site = str(res.get("site_type") or "").strip().lower()
    if site in _positive_site_types():
        return True
    if site == "unclear":
        conf = normalize_confidence(res.get("site_confidence"))
        return conf is not None and conf >= TIER_CONF_MEDIUM
    if site == "other":
        return not confident_no_asset(res)
    return False


def asset_view_is_nearmap_oblique(asset_view: str | None) -> bool:
    view = str(asset_view or "").strip().lower()
    if not view or "naip" in view:
        return False
    if "oblique" in view:
        return True
    return any(d in view for d in ("north", "east", "south", "west"))


def should_keep_claimed_gemini_tower(res: dict, *, from_wide_rescue: bool = False) -> bool:
    """True when outreach-verified Gemini found a disguised mast Claude missed.

    Rooftop HVAC vetoes still stand. Only keep a ground tower whose evidence
    already names a monopalm/monopine/canister (or equivalent stealth form).
    """
    if from_wide_rescue:
        return False
    if not source_expects_asset(res):
        return False
    if str(res.get("site_type") or "").strip().lower() != "tower":
        return False
    if (
        res.get("cell_equipment") is not True
        and res.get("gemini_pre_escalation_cell") is not True
    ):
        return False
    evidence = " ".join(
        str(res.get(key) or "")
        for key in (
            "gemini_pre_escalation_evidence",
            "cell_equipment_evidence",
            "site_evidence",
        )
    )
    return _text_has_stealth_form(evidence) or _text_has_stealth_hardware(evidence)


def should_soft_keep_gemini_cell(res: dict, *, from_wide_rescue: bool) -> bool:
    """Allow Gemini cell=true to win a Claude veto when Nearmap+Gemini are strong.

    Requires full Nearmap obliques, high cell conf, telecom cues, and a compact
    asset box drawn on a Nearmap oblique (not NAIP / whole-roof boxes).
    Never soft-keep after an unvalidated wide-AOI resurrection.
    """
    if from_wide_rescue:
        return False
    if str(res.get("nearmap_tier") or "").strip().lower() != "full":
        return False
    views = str(res.get("nearmap_views") or "").lower()
    if not any(d in views for d in ("north", "east", "south", "west")):
        return False
    if not asset_view_is_nearmap_oblique(res.get("asset_view")):
        return False
    valid = get_valid_asset_box(res)
    if not valid:
        return False
    ymin, xmin, ymax, xmax = valid
    # Soft-keep requires a compact gear box (not a half-roof blob).
    if (ymax - ymin) > 400 or (xmax - xmin) > 400:
        return False
    conf = normalize_confidence(res.get("cell_equipment_confidence"))
    if conf is None or conf < GEMINI_SOFT_KEEP_CELL_CONF:
        return False
    evidence = " ".join(
        str(res.get(key) or "")
        for key in ("cell_equipment_evidence", "site_evidence")
    )
    return _text_has_telecom_cue(evidence)


def should_skip_claude_for_gemini_tower(
    res: dict, *, from_wide_rescue: bool = False
) -> bool:
    """Skip Claude when Gemini already locked a tower at >= GEMINI_SOLO_CELL_CONF.

    Wide-AOI / zoom rescues still go through Claude (wrong-neighbor risk).
    Rooftops are not skipped here — HVAC FPs need the dual-model crop vote.
    """
    if from_wide_rescue:
        return False
    if str(res.get("site_type") or "").strip().lower() != "tower":
        return False
    if res.get("cell_equipment") is not True:
        return False
    conf = normalize_confidence(res.get("site_confidence"))
    return conf is not None and conf >= GEMINI_SOLO_CELL_CONF


def gemini_confidence_locks_claude(res: dict) -> bool:
    """True when Gemini site_confidence is high enough to skip full-scene Claude.

    ``unclear`` never locks. Rooftop cell-unconfirmed is checked first in
    ``escalation_reason`` and still escalates.
    """
    if str(res.get("site_type") or "").strip().lower() == "unclear":
        return False
    conf = normalize_confidence(res.get("site_confidence"))
    return conf is not None and conf >= GEMINI_SOLO_CELL_CONF


def should_trust_gemini_cell_solo(res: dict, *, from_wide_rescue: bool = False) -> bool:
    """Skip Claude dual-confirm when Gemini already locked a strong rooftop cell.

    Requires soft-keep-strength Nearmap localization plus conf >= GEMINI_SOLO_CELL_CONF.
    Saves Claude cost and avoids Claude false vetoes on clear antenna sites.
    """
    if res.get("cell_equipment") is not True:
        return False
    if not should_soft_keep_gemini_cell(res, from_wide_rescue=from_wide_rescue):
        return False
    conf = normalize_confidence(res.get("cell_equipment_confidence"))
    if conf is None or conf < GEMINI_SOLO_CELL_CONF:
        return False
    return has_locked_oblique_asset_box(res)


def align_site_evidence_with_cell(res: dict) -> dict:
    """Keep site_evidence from claiming cellular gear when cell call is not true."""
    if res.get("cell_equipment") is True:
        return res
    site_ev = str(res.get("site_evidence") or "")
    if not site_ev:
        return res
    if _text_has_telecom_cue(site_ev):
        res["site_evidence"] = (
            "Host structure visible on imagery; cellular gear was not confirmed "
            f"(final cell={res.get('cell_equipment')!r})."
        )
        _step_done("site evidence align", "cleared cellular claims (cell not true)")
    return res


def pick_view_for_asset_box(
    res: dict, views: list
) -> tuple[str, Image.Image] | None:
    """Return (label, image) matching asset_view, preferring Nearmap obliques."""
    if not views:
        return None
    target = str(res.get("asset_view") or "").strip().lower()
    if target:
        for label, img in views:
            if target in str(label).strip().lower():
                return label, img
    # Prefer any Nearmap oblique over NAIP/vert fallbacks.
    for label, img in views:
        lower = str(label).lower()
        if any(d in lower for d in ("north", "east", "south", "west")) or "oblique" in lower:
            return label, img
    for label, img in views:
        if "vert" in str(label).lower() or "top-down" in str(label).lower():
            if "naip" not in str(label).lower():
                return label, img
    return None


def build_cell_confirm_views(res: dict, views: list) -> tuple[list, bool]:
    """Crop the asset box for dual-model confirm when possible.

    Towers skip crop: Gemini's mast/compound box often misses the antenna
    array (or lands on foliage / a tank rim), and the rooftop HVAC crop
    prompt then false-vetoes a tower that is obvious in the full Nearmap
    stack. Rooftops still crop so Claude votes on the boxed HVAC-vs-panel
    region.

    Returns (views_for_confirm, used_crop).
    """
    if str(res.get("site_type") or "").strip().lower() == "tower":
        return views, False
    valid = get_valid_asset_box(res)
    picked = pick_view_for_asset_box(res, views)
    if not valid or picked is None:
        return views, False
    label, img = picked
    # Rooftop cell confirm should not run on NAIP when an oblique box was claimed.
    if asset_view_is_nearmap_oblique(res.get("asset_view")) and (
        "naip" in str(label).lower()
    ):
        return views, False
    res["asset_box_2d"] = valid
    crop = _crop_zoom(img, valid, pad_frac=CELL_CONFIRM_PAD_FRAC)
    return [(f"cell crop ({label})", crop), (label, img)], True


_nearmap_session = requests.Session()
_nearmap_coverage_cache: dict[tuple[float, float], tuple[bool, str | None]] = {}


def _nearmap_get(url: str) -> requests.Response:
    """GET with header auth (keeps the API key out of logged URLs) and a
    short retry on rate-limit/transient errors."""
    for attempt in range(3):
        resp = _nearmap_session.get(
            url, headers={"Authorization": f"Apikey {NEARMAP_API_KEY}"},
            timeout=60)
        if resp.status_code in (429, 502, 503) and attempt < 2:
            time.sleep(2 * (attempt + 1))
            continue
        return resp
    return resp


def nearmap_point_coverage(
    lat: float, lon: float
) -> tuple[bool, str | None]:
    """Return (has_survey, capture_date). Fail-open on transport errors.

    Cached per ~1 m so Vert then oblique fetches at the same pin do not
    re-hit the coverage API.
    """
    if not NEARMAP_API_KEY:
        return False, None
    key = (round(float(lat), 5), round(float(lon), 5))
    cached = _nearmap_coverage_cache.get(key)
    if cached is not None:
        return cached
    try:
        resp = _nearmap_get(
            NEARMAP_COVERAGE_POINT_URL.format(lon=lon, lat=lat) + "?limit=1"
        )
        if resp.status_code in (401, 403):
            result = (True, None)
        elif not resp.ok:
            result = (False, None)
        else:
            surveys = resp.json().get("surveys") or []
            capture = None
            if surveys:
                capture = surveys[0].get("captureDate")
            result = (bool(surveys), capture)
    except Exception:
        result = (True, None)
    _nearmap_coverage_cache[key] = result
    return result


def _tile_range(lat: float, lon: float, half_m: float, zoom: int):
    """Slippy-tile x/y index range covering a half_m-radius box at a zoom."""
    dlat = half_m / 111_320.0
    dlon = half_m / (111_320.0 * math.cos(math.radians(lat)))
    n = 2 ** zoom

    def tile_xy(la, lo):
        x = (lo + 180.0) / 360.0 * n
        y = (1.0 - math.asinh(math.tan(math.radians(la))) / math.pi) / 2.0 * n
        return x, y

    x_west, y_north = tile_xy(lat + dlat, lon - dlon)
    x_east, y_south = tile_xy(lat - dlat, lon + dlon)
    return int(x_west), int(x_east), int(y_north), int(y_south)


def fetch_nearmap_views(lat: float, lon: float, chip_m: float = NEARMAP_CHIP_M,
                        views: list[str] | None = None):
    """Fetch Nearmap content for a point via the Tile API: high-res vertical
    plus 45-degree oblique panoramas (N/E/S/W), stitched from XYZ tiles.

    Returns ({view_name: PIL.Image}, capture_date). Empty dict when the key is
    not set or the location has no Nearmap coverage.

    Optional `views` limits which orientations to fetch (e.g. ["Vert"] or
    OBLIQUE_VIEWS). Defaults to all NEARMAP_VIEWS when omitted.
    """
    if not NEARMAP_API_KEY:
        return {}, None

    has_survey, capture_date = nearmap_point_coverage(lat, lon)
    if not has_survey:
        return {}, None

    fetch_views = views if views is not None else NEARMAP_VIEWS
    result = {}
    for view in fetch_views:
        zoom = NEARMAP_VERT_ZOOM if view == "Vert" else NEARMAP_OBLIQUE_ZOOM
        x0, x1, y0, y1 = _tile_range(lat, lon, chip_m / 2.0, zoom)
        cols, rows = x1 - x0 + 1, y1 - y0 + 1

        # Canvas dimensions follow the tile orientation: East/West mosaics
        # have the slippy x axis running vertically
        if view in ("East", "West"):
            canvas = Image.new("RGB", (rows * 256, cols * 256))
        else:
            canvas = Image.new("RGB", (cols * 256, rows * 256))

        got_any = False
        for ty in range(y0, y1 + 1):
            for tx in range(x0, x1 + 1):
                resp = _nearmap_get(NEARMAP_TILE_URL.format(
                    content=view, z=zoom, x=tx, y=ty))
                if resp.status_code == 404:   # no coverage for this tile/view
                    continue
                resp.raise_for_status()
                tile = Image.open(io.BytesIO(resp.content)).convert("RGB")
                got_any = True
                if view in ("Vert", "North"):     # north-up
                    pos = ((tx - x0) * 256, (ty - y0) * 256)
                elif view == "South":             # south-up: both axes flip
                    pos = ((x1 - tx) * 256, (y1 - ty) * 256)
                elif view == "East":              # east-up: up = +x, right = +y
                    pos = ((ty - y0) * 256, (x1 - tx) * 256)
                else:                             # west-up: up = -x, right = -y
                    pos = ((y1 - ty) * 256, (tx - x0) * 256)
                canvas.paste(tile, pos)

        if not got_any:
            continue
        if view != "Vert":
            # Compensate the 45-degree foreshortening (256 -> 192 height)
            canvas = canvas.resize(
                (canvas.width, max(1, int(canvas.height * 0.75))))
        canvas.thumbnail((NEARMAP_MAX_PX, NEARMAP_MAX_PX))
        result[view] = canvas

    if not result:
        return {}, None
    return result, capture_date

# --------------------------- classification stage ---------------------------

def _image_block(img: Image.Image) -> dict:
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=90)
    data = base64.standard_b64encode(buf.getvalue()).decode("ascii")
    return {
        "type": "image",
        "source": {
            "type": "base64",
            "media_type": "image/jpeg",
            "data": data,
        },
    }


def _parse_json_fallback(text: str, default: dict) -> dict:
    text = (text or "").strip()
    start, end = text.find("{"), text.rfind("}")
    if start != -1 and end > start:
        text = text[start:end + 1]
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return {**default, "site_evidence": f"unparseable model reply: {text[:200]}"}


def _extract_tool_result(resp, tool_name: str, default: dict) -> dict:
    for block in resp.content:
        if block.type == "tool_use" and block.name == tool_name:
            if isinstance(block.input, dict):
                return dict(block.input)
    text_parts = [block.text for block in resp.content if block.type == "text"]
    return _parse_json_fallback("\n".join(text_parts), default)


def _call_claude_json(client: Anthropic, content: list, schema: dict,
                      tool_name: str, retries: int = 3,
                      model: str | None = None) -> dict:
    """Shared Claude vision call with tool-based JSON and retry logic."""
    global _model_idx
    attempt = 0
    default = {"site_type": "unclear", "site_confidence": 0.0}
    while True:
        use_model = model or MODELS[_model_idx]
        try:
            resp = client.messages.create(
                model=use_model,
                max_tokens=1000,
                tools=[{
                    "name": tool_name,
                    "description": "Return structured analysis as JSON.",
                    "input_schema": schema,
                }],
                tool_choice={"type": "tool", "name": tool_name},
                messages=[{"role": "user", "content": content}],
            )
            res = _extract_tool_result(resp, tool_name, default)
            normalize_model_result(res)
            res["model"] = use_model
            return res
        except anthropic.RateLimitError:
            if model:
                raise
            if attempt < retries:
                attempt += 1
                wait = 15 * attempt
                _out(f"  rate limit on {use_model}, retrying in {wait}s "
                      f"({attempt}/{retries})...")
                time.sleep(wait)
                continue
            if _model_idx + 1 < len(MODELS):
                _out(f"  {use_model} rate limited -> hopping to {MODELS[_model_idx + 1]}")
                _model_idx += 1
                attempt = 0
                continue
            raise
        except anthropic.APIStatusError as e:
            if model:
                raise
            if e.status_code == 404 and _model_idx + 1 < len(MODELS):
                _out(f"  {use_model} not found (404) -> hopping to {MODELS[_model_idx + 1]}")
                _model_idx += 1
                attempt = 0
                continue
            if e.status_code == 404:
                raise SystemExit(
                    f"\nClaude model '{use_model}' returned 404 (not found). "
                    f"Tried: {', '.join(MODELS)}\n"
                    "Set CLAUDE_MODELS to valid IDs, e.g. "
                    "claude-sonnet-4-6,claude-haiku-4-5-20251001\n"
                    "See https://docs.anthropic.com/en/docs/about-claude/models/overview"
                ) from e
            if e.status_code in (429, 529, 503, 500) and attempt < retries:
                attempt += 1
                wait = 15 * attempt
                _out(f"  transient {e.status_code}, retrying in {wait}s "
                      f"({attempt}/{retries})...")
                time.sleep(wait)
                continue
            if e.status_code in (429, 529) and _model_idx + 1 < len(MODELS):
                _out(f"  {use_model} overloaded -> hopping to {MODELS[_model_idx + 1]}")
                _model_idx += 1
                attempt = 0
                continue
            raise


def _gemini_image_part(img: Image.Image) -> genai_types.Part:
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=90)
    return genai_types.Part.from_bytes(data=buf.getvalue(), mime_type="image/jpeg")


def _gemini_http_status(exc: Exception) -> int | None:
    """Extract an HTTP status from a Gemini SDK or transport error."""
    for attr in ("code", "status_code"):
        value = getattr(exc, attr, None)
        if isinstance(value, int):
            return value
    response = getattr(exc, "response", None)
    if response is not None:
        value = getattr(response, "status_code", None)
        if isinstance(value, int):
            return value
    message = str(exc)
    if " 429 " in message or "429" in message[:8]:
        return 429
    if " 503 " in message:
        return 503
    return None


def _gemini_retry_after_s(exc: Exception) -> float | None:
    response = getattr(exc, "response", None)
    if response is None:
        return None
    headers = getattr(response, "headers", None)
    if not headers:
        return None
    raw = headers.get("retry-after") or headers.get("Retry-After")
    if raw is None:
        return None
    try:
        return max(1.0, float(raw))
    except (TypeError, ValueError):
        return None


def _gemini_retry_wait_s(attempt: int, exc: Exception) -> float:
    """Backoff for transient Gemini rate limits (429) and outages (503)."""
    retry_after = _gemini_retry_after_s(exc)
    if retry_after is not None:
        return retry_after + random.uniform(0.0, 2.0)
    wait = GEMINI_RETRY_BASE_S * (2 ** max(0, attempt - 1))
    return min(wait, 120.0) + random.uniform(0.0, 3.0)


def _is_gemini_3(model: str | None) -> bool:
    return "gemini-3" in str(model or "").lower()


def _is_screen_model(model: str | None) -> bool:
    use = str(model or "").strip()
    screen = str(GEMINI_SCREEN_MODEL or "").strip()
    return "lite" in use.lower() or (bool(screen) and use == screen)


def _gemini_thinking_level(model: str | None = None) -> str:
    """Gemini 3.x thinking_level for this model (NAIP screen vs Nearmap/confirm)."""
    if _is_screen_model(model or GEMINI_MODEL):
        return _parse_thinking_level(
            os.environ.get("GEMINI_SCREEN_THINKING_LEVEL"),
            GEMINI_SCREEN_THINKING_LEVEL,
        )
    return _parse_thinking_level(
        os.environ.get("GEMINI_THINKING_LEVEL"), GEMINI_THINKING_LEVEL
    )


def _gemini_thinking_budget(model: str | None = None) -> int:
    """Resolved thinking token budget for a Gemini vision model.

    Default is 0 (thinking is billed). Set GEMINI_THINKING_BUDGET=1024 to restore
    Flash thinking. Lite models always stay at 0.
    """
    raw = GEMINI_THINKING_BUDGET_ENV or os.environ.get("GEMINI_THINKING_BUDGET", "").strip()
    use = str(model or GEMINI_MODEL or "").strip().lower()
    if "lite" in use:
        return 0
    if raw and raw.lower() not in {"auto", "default"}:
        try:
            return max(0, int(float(raw)))
        except (TypeError, ValueError):
            return 0
    return 0


def _gemini_generate_config(
    schema: dict, model: str | None = None
) -> genai_types.GenerateContentConfig:
    """Structured JSON config. Gemini 3.x uses thinking_level, not thinking_budget."""
    use_model = model or GEMINI_MODEL
    budget = _gemini_thinking_budget(use_model)
    name = str(use_model or "").lower()
    kwargs: dict = {
        "response_mime_type": "application/json",
        "response_schema": schema,
        "max_output_tokens": 8192 if budget > 0 else 1000,
    }
    if name.startswith("gemini-2.0"):
        pass
    elif _is_gemini_3(use_model):
        # thinking_budget on 3.x lite/preview returns 400 INVALID_ARGUMENT.
        level = _gemini_thinking_level(use_model)
        kwargs["thinking_config"] = genai_types.ThinkingConfig(thinking_level=level)
        kwargs["max_output_tokens"] = 1000 if level == "MINIMAL" else 8192
    else:
        kwargs["thinking_config"] = genai_types.ThinkingConfig(
            thinking_budget=budget
        )
    return genai_types.GenerateContentConfig(**kwargs)


def _call_gemini_json(client: genai.Client, contents: list, schema: dict,
                      retries: int | None = None,
                      model: str | None = None) -> dict:
    """Gemini vision call with structured JSON output (single model)."""
    max_retries = GEMINI_RETRIES if retries is None else retries
    use_model = model or GEMINI_MODEL
    attempt = 0
    while True:
        try:
            resp = client.models.generate_content(
                model=use_model,
                contents=contents,
                config=_gemini_generate_config(schema, model=use_model),
            )
            break
        except Exception as exc:
            status = _gemini_http_status(exc)
            if status in (429, 503) and attempt < max_retries:
                attempt += 1
                wait = _gemini_retry_wait_s(attempt, exc)
                label = "rate limit" if status == 429 else "service unavailable"
                _out(f"  transient Gemini {status} ({label}), retrying in "
                      f"{wait:.0f}s ({attempt}/{max_retries})...")
                time.sleep(wait)
                continue
            raise
    text = (resp.text or "").strip()
    start, end = text.find("{"), text.rfind("}")
    if start != -1 and end > start:
        text = text[start:end + 1]
    try:
        res = json.loads(text)
    except json.JSONDecodeError:
        res = {"site_type": "unclear", "site_confidence": 0.0,
               "site_evidence": f"unparseable model reply: {text[:200]}"}
    normalize_model_result(res)
    res["model"] = use_model
    return res


def _views_to_claude_content(views: list[tuple[str, Image.Image]], prompt: str) -> list:
    content = []
    for label, img in views:
        content.append({"type": "text", "text": f"View: {label}"})
        content.append(_image_block(img))
    content.append({"type": "text", "text": prompt})
    return content


def _views_to_gemini_contents(views: list[tuple[str, Image.Image]], prompt: str) -> list:
    contents = []
    for label, img in views:
        contents.append(f"View: {label}")
        contents.append(_gemini_image_part(img))
    contents.append(prompt)
    return contents


def downscale_image(img: Image.Image, max_px: int) -> Image.Image:
    """Return a copy capped at max_px on the long side (original unchanged)."""
    if img is None or max_px <= 0:
        return img
    if max(img.size) <= max_px:
        return img
    copy = img.copy()
    copy.thumbnail((max_px, max_px), Image.Resampling.LANCZOS)
    return copy


def trim_views_for_model(
    views: list,
    *,
    max_obliques: int | None = None,
    max_px: int | None = None,
) -> list:
    """Keep one top-down + a few obliques (or crops) and downscale for the API."""
    if not views:
        return views
    max_obliques = MODEL_MAX_OBLIQUES if max_obliques is None else max_obliques
    max_px = MODEL_IMAGE_MAX_PX if max_px is None else max_px
    top: list = []
    obliques: list = []
    other: list = []
    for label, img in views:
        lower = str(label).lower()
        if "zoom crop" in lower or "cell crop" in lower:
            other.append((label, img))
            continue
        is_naip = "naip" in lower
        is_oblique = (not is_naip) and (
            "oblique" in lower
            or any(d in lower for d in ("north", "east", "south", "west"))
        )
        if is_oblique:
            obliques.append((label, img))
        elif is_naip or "vert" in lower or "top-down" in lower:
            top.append((label, img))
        else:
            other.append((label, img))
    nearmap_vert = [
        item
        for item in top
        if "naip" not in str(item[0]).lower()
    ]
    chosen_top = (nearmap_vert[:1] if nearmap_vert else top[:1])
    selected = chosen_top + obliques[: max(0, max_obliques)] + other
    return [(label, downscale_image(img, max_px)) for label, img in selected]


def classify_site(provider: str, clients: dict,
                  views: list[tuple[str, Image.Image]],
                  prompt: str = CLASSIFICATION_PROMPT, retries: int = 3,
                  scan: bool = False, claude_model: str | None = None,
                  gemini_model: str | None = None,
                  trim_views: bool = True,
                  image_max_px: int | None = None) -> dict:
    """Classify one asset via Gemini or Claude using the same prompt."""
    send_views = (
        trim_views_for_model(views, max_px=image_max_px)
        if trim_views
        else list(views)
    )
    if scan:
        claude_schema, gemini_schema = SCAN_SCHEMA, GEMINI_SCAN_SCHEMA
        tool_name = "scan_candidates"
        classify_prompt = _active_scan_prompt()
    else:
        claude_schema, gemini_schema = _classification_response_schemas()
        tool_name = "classify_site"
        classify_prompt = prompt
    if provider == "gemini":
        contents = _views_to_gemini_contents(send_views, classify_prompt)
        return _call_gemini_json(
            clients["gemini"], contents, gemini_schema, retries,
            model=gemini_model or GEMINI_MODEL,
        )
    content = _views_to_claude_content(send_views, classify_prompt)
    return _call_claude_json(
        clients["claude"], content, claude_schema, tool_name, retries,
        model=claude_model)


def site_confidence_band(res: dict) -> str:
    """Map numeric site_confidence to high | medium | low for tier gating."""
    conf = normalize_confidence(res.get("site_confidence"))
    if conf is None:
        return "low"
    if conf >= TIER_CONF_HIGH:
        return "high"
    if conf >= TIER_CONF_MEDIUM:
        return "medium"
    return "low"


def _is_rooftop(res: dict) -> bool:
    return str(res.get("site_type") or "").strip().lower() == "rooftop"


def _rooftop_cell_confirmed(res: dict) -> bool:
    """True when rooftop cell gear is true at/above ROOFTOP_CELL_CONF_MIN.

    Missing cell_equipment_confidence still counts as confirmed when
    cell_equipment is True (model sometimes omits the numeric field).
    """
    if res.get("cell_equipment") is not True:
        return False
    cell_conf = normalize_confidence(res.get("cell_equipment_confidence"))
    if cell_conf is None:
        return True
    return cell_conf >= ROOFTOP_CELL_CONF_MIN


def tier_confident_stop(res: dict) -> bool:
    """True when tiered fetch can stop without pulling the next Nearmap tier.

    Towers: medium+ site conf and cell_equipment decided (true or false).
    Rooftops: medium+ site conf and confirmed cellular gear (true + conf bar).
    cell=false/null on a rooftop must continue — facade/parapet gear is often
    invisible on NAIP/Vert alone.
    """
    if res.get("site_type") not in _positive_site_types():
        return False
    if site_confidence_band(res) == "low":
        return False
    if _is_rooftop(res):
        return _rooftop_cell_confirmed(res)
    if res.get("cell_equipment") is None:
        return False
    return True


def db_backed_naip_tower_skip_nearmap_reason(
    res: dict, *, db_backed: bool = False
) -> str | None:
    """Skip Nearmap when FCC/TowerSource plus NAIP already decided a tower.

    Medium+ site confidence and cell_equipment true/false is enough — the
    database anchors the pin. Imagery-only towers still buy obliques.
    """
    if not db_backed:
        return None
    if str(res.get("site_type") or "").strip().lower() != "tower":
        return None
    if not tier_confident_stop(res):
        return None
    return "DB-hit NAIP tower medium+ cell decided"


def rooftop_naip_cell_skip_nearmap_reason(
    res: dict, naip_age_years: float | None = None
) -> str | None:
    """Skip Nearmap when NAIP already locked rooftop cell at the Gemini solo bar."""
    if not _is_rooftop(res):
        return None
    if res.get("cell_equipment") is not True:
        return None
    cell_conf = normalize_confidence(res.get("cell_equipment_confidence"))
    if cell_conf is None or cell_conf < GEMINI_SOLO_CELL_CONF:
        return None
    if site_confidence_band(res) == "low":
        return None
    if naip_age_blocks_early_stop(res, naip_age_years):
        return None
    return f"NAIP rooftop cell conf>={GEMINI_SOLO_CELL_CONF:g}"


def naip_empty_osm_skip_nearmap_reason(
    res: dict, *, db_backed: bool = False, osm_info: dict | None = None
) -> str | None:
    """Skip Nearmap when NAIP is empty and OSM shows no building or tower.

    Fail-open: missing/failed OSM does not skip. DB-backed claimed towers
    still buy Nearmap if NAIP saw nothing.
    """
    if db_backed:
        return None
    site = str(res.get("site_type") or "").strip().lower()
    if site not in {"other", "unclear"}:
        return None
    try:
        from enrichment.osm_prefilter import osm_suggests_empty_chip
    except Exception:
        return None
    if not osm_suggests_empty_chip(osm_info):
        return None
    return "NAIP empty + OSM no building/tower"


def skip_nearmap_after_naip_reason(
    res: dict,
    *,
    db_backed: bool = False,
    osm_tower: bool = False,
    osm_info: dict | None = None,
    naip_age_years: float | None = None,
) -> str | None:
    """First skip-Nearmap reason after the NAIP pass, else None."""
    locked = locked_gemini_tower_skip_nearmap_reason(
        res, db_backed=db_backed, osm_tower=osm_tower
    )
    if locked:
        return locked
    medium = db_backed_naip_tower_skip_nearmap_reason(res, db_backed=db_backed)
    if medium:
        return medium
    roof = rooftop_naip_cell_skip_nearmap_reason(res, naip_age_years)
    if roof:
        return roof
    return naip_empty_osm_skip_nearmap_reason(
        res, db_backed=db_backed, osm_info=osm_info
    )


def rooftop_requires_nearmap_tiers(res: dict) -> bool:
    """True when the remaining rooftop path should fetch Vert + obliques.

    Locked NAIP rooftop cell (>= GEMINI_SOLO_CELL_CONF) skips earlier via
    rooftop_naip_cell_skip_nearmap_reason. Callers that reach this point
    still need paid imagery for facade confirmation.
    """
    return _is_rooftop(res)


def tower_cell_requires_nearmap_obliques(
    res: dict, *, db_backed: bool = False
) -> bool:
    """Whether a tower+cell call must continue to Nearmap obliques.

    Imagery-only (no FCC/TowerSource) always continues: pin-centered Vert
    often frames a neighbor rooftop or a different pole than the SF pin.

    DB-hit Gemini towers that already decided cell at medium+ site conf
    may stop on NAIP. Rooftops are handled separately by
    rooftop_requires_nearmap_tiers.
    """
    if str(res.get("site_type") or "").strip().lower() != "tower":
        return False
    if res.get("cell_equipment") is not True:
        return False
    if db_backed and db_backed_naip_tower_skip_nearmap_reason(
        res, db_backed=True
    ):
        return False
    return True


def naip_age_blocks_early_stop(
    res: dict,
    naip_age_years: float | None,
    *,
    max_age_years: float | None = None,
    high_conf_override: float | None = None,
) -> bool:
    """True when stale NAIP must not early-stop Nearmap (rooftop equipment risk).

    Towers can still stop on old NAIP. Rooftops on imagery older than
    NAIP_MAX_AGE_YEARS continue to Nearmap, unless confidence is at/above
    NAIP_AGE_HIGH_CONF_OVERRIDE (definitive asset ID on NAIP alone).
    Early-stop still also requires tier_confident_stop (equipment decided).
    Note: classify_with_tiers skips Nearmap for rooftops only when NAIP
    cell is already locked; stale NAIP still continues via this gate.
    """
    if naip_age_years is None:
        return False
    try:
        age = float(naip_age_years)
    except (TypeError, ValueError):
        return False
    limit = NAIP_MAX_AGE_YEARS if max_age_years is None else float(max_age_years)
    if age <= limit:
        return False
    if not _is_rooftop(res):
        return False
    conf = normalize_confidence(res.get("site_confidence"))
    override = (
        NAIP_AGE_HIGH_CONF_OVERRIDE
        if high_conf_override is None
        else float(high_conf_override)
    )
    if conf is not None and conf >= override:
        return False
    return True


def _has_nearmap_context(res: dict) -> bool:
    """True when this result already used Nearmap (not NAIP-only rescue)."""
    tier = str(res.get("nearmap_tier") or "").strip().lower()
    if tier in {"full", "wide_aoi", "vert_only"}:
        return True
    views = str(res.get("nearmap_views") or "").strip()
    if views:
        return True
    return has_locked_oblique_asset_box(res)


def _nearmap_empty_without_cell(res: dict) -> bool:
    """Full Nearmap other/unclear with no confirmed cell — skip Claude."""
    from enrichment.cost_policy import nearmap_empty_without_cell

    return nearmap_empty_without_cell(
        res.get("site_type"),
        res.get("nearmap_tier"),
        res.get("cell_equipment"),
    )


def escalation_reason(res: dict) -> str | None:
    """Why a Gemini result should escalate to Claude; None if no escalation."""
    site = str(res.get("site_type") or "").strip().lower()
    conf = normalize_confidence(res.get("site_confidence"))
    # Carrier-claimed / high-trust: Gemini empty on Nearmap still needs Claude
    # unless the pack locked other at >= 0.90. Medium-trust empties skip.
    if site in {"other", "unclear"} and _has_nearmap_context(res):
        if gemini_confidence_locks_claude(res):
            return None
        if _nearmap_empty_without_cell(res):
            if (
                source_expects_asset(res)
                and conf is not None
                and conf < GEMINI_SOLO_CELL_CONF
            ):
                return "nearmap_claimed_site_empty"
            return None
        if conf is not None and conf >= CLAUDE_ESCALATE_MIN_SITE_CONF:
            return "nearmap_claimed_site_empty"
        return None
    if conf is None or conf < CLAUDE_ESCALATE_MIN_SITE_CONF:
        return None
    # Rooftops: Claude is for Nearmap HVAC vs antenna, not a NAIP substitute.
    if _is_rooftop(res):
        if not _has_nearmap_context(res):
            return None
        if res.get("cell_equipment") is not True:
            return "rooftop_cell_unconfirmed"
        if not _rooftop_cell_confirmed(res):
            return "rooftop_low_cell_confidence"
    elif res.get("cell_equipment") is False:
        # Towers (and non-rooftop): definitive no-gear — skip Claude.
        return None
    # Gemini already locked the call at >= 0.9 — don't overwrite with Claude.
    if gemini_confidence_locks_claude(res):
        return None
    if site in {"other", "unclear"}:
        return None
    if site_confidence_band(res) == "low":
        return "low_confidence"
    return None


def _brief_pass_result(res: dict) -> str:
    """Compact site_type/confidence for step-done lines."""
    site = str(res.get("site_type") or "—").strip() or "—"
    subtype = res.get("tower_subtype")
    if site == "tower" and subtype not in (None, "", "nan"):
        site = f"{site}/{subtype}"
    conf = res.get("site_confidence")
    try:
        conf_s = f"{float(conf):.2f}"
    except (TypeError, ValueError):
        conf_s = "—"
    return f"{site} conf={conf_s}"


def _step_done(step: str, detail: str | None = None) -> None:
    """Operator progress: mark a per-site pipeline step complete."""
    if detail:
        _out(f"         {step} — done ({detail})", important=True)
    else:
        _out(f"         {step} — done", important=True)


def _classify_pass(
    provider: str,
    clients: dict,
    views: list,
    prompt: str,
    input_confidence: str,
    *,
    screen: bool = False,
) -> dict:
    """One Gemini/Claude classify pass; optional lite screen then Flash confirm."""
    gemini_model = None
    if provider == "gemini":
        gemini_model = GEMINI_SCREEN_MODEL if screen else GEMINI_MODEL
    res = classify_site(
        provider,
        clients,
        views,
        prompt=prompt,
        gemini_model=gemini_model,
        image_max_px=SCREEN_IMAGE_MAX_PX if screen else None,
    )
    if (
        screen
        and provider == "gemini"
        and GEMINI_SCREEN_MODEL != GEMINI_MODEL
        and needs_flash_confirm(res)
    ):
        res = classify_site(
            provider, clients, views, prompt=prompt, gemini_model=GEMINI_MODEL
        )
        _step_done("Flash confirm", _brief_pass_result(res))
    res["input_confidence"] = normalize_input_confidence(input_confidence)
    res = maybe_recheck_equipment(provider, clients, res, views, input_confidence)
    res["input_confidence"] = normalize_input_confidence(input_confidence)
    return res


def _osm_features(lat: float, lon: float) -> dict:
    """Nearby OSM building/tower flags. Fail-open to ok=False."""
    empty = {
        "ok": False,
        "has_building": False,
        "has_tower_or_mast": False,
        "communication_tower": False,
    }
    try:
        from enrichment.osm_prefilter import lookup_osm_features
    except Exception:
        return empty
    try:
        return lookup_osm_features(lat, lon)
    except Exception:
        return empty


def locked_gemini_tower_skip_nearmap_reason(
    res: dict, *, db_backed: bool = False, osm_tower: bool = False
) -> str | None:
    """Skip-Nearmap label for a locked Gemini tower, or None.

    FCC/TowerSource (``db_backed``) is a DB-hit. OSM is not — it uses a
    separate string so terminals and metrics do not count it as a database match.
    """
    if not should_skip_claude_for_gemini_tower(res):
        return None
    if rooftop_requires_nearmap_tiers(res):
        return None
    if db_backed:
        return f"DB-hit Gemini tower conf>={GEMINI_SOLO_CELL_CONF:g}"
    if osm_tower:
        return "OSM tower + Gemini lock — no obliques"
    return None


def classify_with_tiers(lat: float, lon: float, img: Image.Image | None,
                        provider: str, clients: dict, prompt: str,
                        input_confidence: str,
                        build_views,
                        naip_age_years: float | None = None,
                        db_backed: bool = False,
                        osm_tower: bool = False) -> tuple[dict, dict, str | None, str, list]:
    """NAIP screen, then Nearmap Vert+obliques unless a cheaper lock applies.

    Skip paid imagery when: a DB-backed NAIP tower already decided cell at
    medium+ conf; NAIP rooftop cell is locked at the Gemini solo bar; NAIP
    is empty and OSM has no building/tower; an OSM communication tower plus
    Gemini lock; or there is no API key. Remaining claimed sites and weaker
    rooftops still buy Vert+obliques.
    """
    nearmap_views: dict = {}
    nearmap_date = None

    views = build_views({})
    _step_done("NAIP")
    res = _classify_pass(
        provider, clients, views, prompt, input_confidence, screen=True
    )
    _step_done("classify (NAIP)", _brief_pass_result(res))
    stamp_naip_screen(res)

    skip = skip_nearmap_after_naip_reason(
        res, db_backed=db_backed, osm_tower=False, naip_age_years=naip_age_years
    )
    if skip:
        _step_done("Nearmap skipped", skip)
        return res, nearmap_views, nearmap_date, "naip_only", views

    osm_info = None
    if not db_backed:
        osm_info = _osm_features(lat, lon)
        if osm_info.get("communication_tower"):
            osm_tower = True
            _step_done("OSM", "communication tower nearby")
        elif osm_info.get("ok") and not (
            osm_info.get("has_building") or osm_info.get("has_tower_or_mast")
        ):
            _step_done("OSM", "no building/tower")
        skip = skip_nearmap_after_naip_reason(
            res,
            db_backed=False,
            osm_tower=osm_tower,
            osm_info=osm_info,
            naip_age_years=naip_age_years,
        )
        if skip:
            _step_done("Nearmap skipped", skip)
            return res, nearmap_views, nearmap_date, "naip_only", views

    if rooftop_requires_nearmap_tiers(res):
        _step_done(
            "Nearmap required",
            "rooftop — Vert+obliques for cellular confirmation",
        )
    else:
        _step_done(
            "Nearmap required",
            "claimed site — Vert+obliques (NAIP cannot rule out gear)",
        )

    if not NEARMAP_API_KEY:
        return res, nearmap_views, nearmap_date, "naip_only", views

    vert_views, vert_date = fetch_nearmap_views(lat, lon, views=["Vert"])
    nearmap_views.update(vert_views)
    nearmap_date = vert_date or nearmap_date
    if not nearmap_views:
        _step_done("Nearmap vert", "no coverage")
        return res, nearmap_views, nearmap_date, "no_coverage", views
    _step_done("Nearmap vert")

    missing = [v for v in OBLIQUE_VIEWS if v not in nearmap_views]
    if missing:
        oblique_views, ob_date = fetch_nearmap_views(lat, lon, views=missing)
        nearmap_views.update(oblique_views)
        nearmap_date = ob_date or nearmap_date
    has_obliques = any(v in nearmap_views for v in OBLIQUE_VIEWS)
    if has_obliques:
        _step_done("Nearmap obliques")
    else:
        _step_done("Nearmap obliques", "no coverage")

    views = build_views(nearmap_views)
    res = _classify_pass(
        provider, clients, views, prompt, input_confidence, screen=False
    )
    res = gate_weak_rooftop_cell_claim(res)
    res = gate_weak_stealth_tower_claim(res)
    _step_done("classify (Nearmap combined)", _brief_pass_result(res))
    return (
        res,
        nearmap_views,
        nearmap_date,
        "full" if has_obliques else "vert_only",
        views,
    )


def classify_chip(client: Anthropic, views: list[tuple[str, Image.Image]],
                  prompt: str = CLASSIFICATION_PROMPT, retries: int = 3) -> dict:
    """Legacy wrapper — prefer classify_site()."""
    return classify_site(
        "claude", {"claude": client}, views, prompt=prompt, retries=retries)


def _valid_box(box) -> list[int] | None:
    """Validate a zoom-scout candidate box (larger minimum side)."""
    try:
        ymin, xmin, ymax, xmax = (int(v) for v in box[:4])
    except (TypeError, ValueError):
        return None
    if not (0 <= ymin < ymax <= 1000 and 0 <= xmin < xmax <= 1000):
        return None
    if (ymax - ymin) < ZOOM_MIN_FRAC * 1000 or (xmax - xmin) < ZOOM_MIN_FRAC * 1000:
        return None
    return [ymin, xmin, ymax, xmax]


def coerce_asset_box(box) -> list[int] | None:
    """Normalize and validate a rooftop/tower asset_box_2d.

    Fixes inverted corners and rejects empty / whole-scene boxes. Uses a
    smaller minimum than zoom scout so tight antenna mounts stay valid.
    """
    if box is None or box == "":
        return None
    if isinstance(box, str):
        try:
            box = json.loads(box)
        except json.JSONDecodeError:
            return None
    if not isinstance(box, (list, tuple)) or len(box) < 4:
        return None
    try:
        ymin, xmin, ymax, xmax = (int(round(float(v))) for v in box[:4])
    except (TypeError, ValueError):
        return None
    if ymin > ymax:
        ymin, ymax = ymax, ymin
    if xmin > xmax:
        xmin, xmax = xmax, xmin
    ymin = max(0, min(1000, ymin))
    xmin = max(0, min(1000, xmin))
    ymax = max(0, min(1000, ymax))
    xmax = max(0, min(1000, xmax))
    if ymin >= ymax or xmin >= xmax:
        return None
    min_side = ASSET_BOX_MIN_FRAC * 1000
    if (ymax - ymin) < min_side or (xmax - xmin) < min_side:
        return None
    if (ymax - ymin) > ASSET_BOX_MAX_SIDE or (xmax - xmin) > ASSET_BOX_MAX_SIDE:
        return None
    return [ymin, xmin, ymax, xmax]


def get_valid_asset_box(res: dict) -> list[int] | None:
    return coerce_asset_box(res.get("asset_box_2d"))


def _grid_boxes(grid: int = ZOOM_GRID) -> list[list[int]]:
    """Return normalized boxes for an NxN grid covering the full image."""
    step = 1000 // grid
    boxes = []
    for row in range(grid):
        for col in range(grid):
            ymin = row * step
            xmin = col * step
            ymax = 1000 if row == grid - 1 else (row + 1) * step
            xmax = 1000 if col == grid - 1 else (col + 1) * step
            boxes.append([ymin, xmin, ymax, xmax])
    return boxes


def _crop_zoom(
    img: Image.Image, box: list[int], *, pad_frac: float | None = None
) -> Image.Image:
    """Magnify a normalized box from a source image to ZOOM_OUTPUT_PX."""
    w, h = img.size
    ymin, xmin, ymax, xmax = box
    pad = ZOOM_PAD_FRAC if pad_frac is None else float(pad_frac)
    pad_y = int((ymax - ymin) * pad)
    pad_x = int((xmax - xmin) * pad)
    ymin = max(0, ymin - pad_y)
    xmin = max(0, xmin - pad_x)
    ymax = min(1000, ymax + pad_y)
    xmax = min(1000, xmax + pad_x)
    left = int(xmin / 1000.0 * w)
    upper = int(ymin / 1000.0 * h)
    right = max(left + 1, int(xmax / 1000.0 * w))
    lower = max(upper + 1, int(ymax / 1000.0 * h))
    crop = img.crop((left, upper, right, lower))
    crop = crop.resize((ZOOM_OUTPUT_PX, ZOOM_OUTPUT_PX), Image.Resampling.LANCZOS)
    return crop


def scout_candidates(provider: str, clients: dict, label: str,
                     img: Image.Image) -> list[dict]:
    """Ask the vision model to propose candidate regions on one top-down image."""
    views = [(label, img)]
    gemini_model = GEMINI_SCREEN_MODEL if provider == "gemini" else None
    res = classify_site(
        provider,
        clients,
        views,
        scan=True,
        gemini_model=gemini_model,
        image_max_px=SCREEN_IMAGE_MAX_PX,
    )
    return res.get("candidates") or []


def _anchor_candidates() -> list[dict]:
    """Default crops around the recorded coordinate (chip center) and the
    band just below center, where assets often sit when coords are approximate."""
    return [
        {"box_2d": [350, 350, 650, 650],
         "reason": "coordinate anchor (center)"},
        {"box_2d": [480, 380, 680, 620],
         "reason": "coordinate anchor (just below center)"},
    ]


def select_zoom_candidates(
    scouted: list[dict],
    *,
    max_crops: int = ZOOM_MAX_CANDIDATES,
) -> list[dict]:
    """Prefer scout boxes; fill remaining slots with pin-center anchors."""
    merged: list[dict] = []
    seen: set[tuple] = set()

    def _add(cands: list[dict]) -> None:
        for cand in cands:
            box = _valid_box(cand.get("box_2d"))
            if box is None:
                continue
            key = tuple(box)
            if key in seen:
                continue
            seen.add(key)
            merged.append(cand)
            if len(merged) >= max_crops:
                return

    _add(list(scouted or []))
    if len(merged) < max_crops:
        _add(_anchor_candidates())
    if not merged:
        _add([{"box_2d": b, "reason": "grid sweep"} for b in _grid_boxes()])
    return merged[:max_crops]


def build_zoom_views(asset_id: str, source_label: str, source_img: Image.Image,
                     candidates: list[dict]) -> list[tuple[str, Image.Image]]:
    """Turn scout candidates into magnified zoom crops; save each to chips/."""
    zoom_views = []
    seen = set()
    for i, cand in enumerate(candidates[:ZOOM_MAX_CANDIDATES], start=1):
        box = _valid_box(cand.get("box_2d"))
        if box is None:
            continue
        key = tuple(box)
        if key in seen:
            continue
        seen.add(key)
        crop = _crop_zoom(source_img, box)
        reason = (cand.get("reason") or "candidate").replace("\n", " ")[:80]
        path = CHIP_DIR / f"{asset_id}_zoom_{i}.jpg"
        crop.save(path, quality=92)
        zoom_views.append((f"Zoom crop {i} ({reason})", crop))
    return zoom_views


def run_zoom_stage(provider: str, clients: dict, asset_id: str,
                   context_views: list[tuple[str, Image.Image]],
                   source_label: str, source_img: Image.Image,
                   max_crops: int = ZOOM_MAX_CANDIDATES) -> tuple[dict, int]:
    """Scout + magnify + re-classify. Returns (result dict, zoom crop count)."""
    scouted = scout_candidates(provider, clients, source_label, source_img)
    if not scouted:
        _out(f"  [{asset_id}] scout found no extra candidates")
    candidates = select_zoom_candidates(scouted, max_crops=max_crops)

    zoom_views = build_zoom_views(asset_id, source_label, source_img, candidates)
    if not zoom_views:
        return {"site_type": "unclear", "site_confidence": 0.0,
                "site_evidence": "Zoom stage could not build valid crops."}, 0

    context = context_views[:1] if context_views else []
    res = classify_site(
        provider, clients, context + zoom_views, prompt=_active_zoom_prompt())
    res["classification_stage"] = "zoom"
    return res, len(zoom_views)


def _row_error(record: dict) -> str | None:
    """Return a non-empty error string, or None if the row succeeded."""
    err = record.get("error")
    if err is None or (isinstance(err, float) and pd.isna(err)):
        return None
    err = str(err).strip()
    return err or None


def _format_asset_label(record: dict) -> str:
    label = str(record.get("label", "")).strip()
    aid = record.get("id", "")
    return f"{aid} ({label})" if label else str(aid)


def _format_located(record: dict) -> str:
    off = record.get("asset_offset_m")
    if off is not None and not (isinstance(off, float) and pd.isna(off)):
        return f"{off:.0f} m off"
    if record.get("asset_view"):
        return f"on {record['asset_view']}"
    return "—"


def _format_cell_equip(record: dict) -> str:
    ce = record.get("cell_equipment")
    ev = str(record.get("cell_equipment_evidence") or "").strip()
    if ce is True:
        return f"true — {ev}" if ev else "true"
    if ce is False:
        return f"false — {ev}" if ev else "false"
    if ce is None:
        return "unknown"
    return str(ce)


def _format_confidence(record: dict) -> str | float:
    conf = normalize_confidence(record.get("site_confidence"))
    if conf is None:
        return "—"
    return round(conf, 2)


def pick_review_image_path(asset_id: str, record: dict) -> Path | None:
    """Pick the best saved chip for stakeholder review (oblique > NAIP > zoom)."""
    chip = CHIP_DIR
    nm = (record.get("nearmap_views") or "").lower()
    asset_view = (record.get("asset_view") or "").lower()

    oblique_dirs = {
        "north": "north", "east": "east", "south": "south", "west": "west",
    }
    if record.get("cell_equipment") is True:
        for name, suffix in oblique_dirs.items():
            if name in nm or name in asset_view:
                path = chip / f"{asset_id}_nearmap_{suffix}.jpg"
                if path.exists():
                    return path

    for name, suffix in oblique_dirs.items():
        if name in asset_view:
            path = chip / f"{asset_id}_nearmap_{suffix}.jpg"
            if path.exists():
                return path

    if "nearmap top-down" in asset_view or "vert" in nm:
        path = chip / f"{asset_id}_nearmap_vert.jpg"
        if path.exists():
            return path

    chip_path = record.get("chip_path")
    if chip_path:
        path = Path(chip_path)
        if path.exists():
            return path

    for name in (
        f"{asset_id}_nearmap_vert.jpg",
        f"{asset_id}_NAIP_wide.jpg",
        f"{asset_id}_NAIP.jpg",
    ):
        path = chip / name
        if path.exists():
            return path

    zooms = sorted(chip.glob(f"{asset_id}_zoom_*.jpg"))
    return zooms[0] if zooms else None


def build_stakeholder_row(record: dict) -> dict:
    err = _row_error(record)
    if err:
        return {
            "Asset": _format_asset_label(record),
            "Site type": "error",
            "Conf": "—",
            "Located": "—",
            "Cell equip": err[:120],
            "Views": record.get("view_count", 0),
            "Review image": record.get("review_image"),
        }
    if record.get("site_type") == "no_imagery":
        return {
            "Asset": _format_asset_label(record),
            "Site type": "no imagery",
            "Conf": "—",
            "Located": "—",
            "Cell equip": "—",
            "Views": 0,
            "Review image": None,
        }
    return {
        "Asset": _format_asset_label(record),
        "Site type": record.get("site_type"),
        "Conf": _format_confidence(record),
        "Located": _format_located(record),
        "Cell equip": _format_cell_equip(record),
        "Views": record.get("view_count", 0),
        "Review image": record.get("review_image"),
    }


def write_stakeholder_report(results: list[dict], report_csv: str, report_xlsx: str):
    """Write a clean CSV + Excel workbook with embedded review images."""
    rows = [build_stakeholder_row(r) for r in results]
    report_df = pd.DataFrame(rows)
    report_df.to_csv(report_csv, index=False)

    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"Stakeholder CSV written to {report_csv} "
              "(install openpyxl for Excel export)")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["Asset", "Site type", "Conf", "Located", "Cell equip", "Views", "Photo"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[row_idx].height = 95
        ws.cell(row=row_idx, column=1, value=row["Asset"])
        ws.cell(row=row_idx, column=2, value=row["Site type"])
        ws.cell(row=row_idx, column=3, value=row["Conf"])
        ws.cell(row=row_idx, column=4, value=row["Located"])
        ws.cell(row=row_idx, column=5, value=row["Cell equip"])
        ws.cell(row=row_idx, column=6, value=row["Views"])

        img_path = row.get("Review image")
        if img_path and Path(img_path).exists():
            thumb = CHIP_DIR / f"_thumb_{Path(img_path).name}"
            with Image.open(img_path) as im:
                im = im.convert("RGB")
                im.thumbnail((160, 120))
                thumb_w, thumb_h = im.size
                im.save(thumb, quality=85)
            xl_img = XLImage(str(thumb))
            xl_img.width, xl_img.height = thumb_w, thumb_h
            col = get_column_letter(7)
            ws.add_image(xl_img, f"{col}{row_idx}")

    widths = {"A": 22, "B": 12, "C": 8, "D": 14, "E": 42, "F": 8, "G": 24}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    try:
        wb.save(report_xlsx)
    except PermissionError:
        alt = str(Path(report_xlsx).with_stem(Path(report_xlsx).stem + "_updated"))
        wb.save(alt)
        print(f"Stakeholder report written to {report_csv} and {alt}")
        print(f"  (Could not overwrite {report_xlsx} — file is open. Close it and "
              f"re-run --regenerate-report to refresh.)")
        return

    print(f"Stakeholder report written to {report_csv} and {report_xlsx}")


def write_executive_summary(results: list[dict], assets_df: pd.DataFrame):
    """Write a stakeholder-friendly markdown summary of the latest run."""
    total = len(assets_df)
    errors = [r for r in results if _row_error(r)]
    classified = [r for r in results if r.get("site_type") not in
                  (None, "no_imagery") and not _row_error(r)]
    towers = sum(1 for r in classified if r.get("site_type") == "tower")
    rooftops = sum(1 for r in classified if r.get("site_type") == "rooftop")
    cell_hits = sum(1 for r in classified if r.get("cell_equipment") is True)
    located = [r for r in classified if r.get("asset_lat") is not None
               or r.get("asset_view")]

    lines = [
        "# Asset Classifier — Executive Summary",
        "",
        f"*Generated: {time.strftime('%B %d, %Y at %I:%M %p')}*",
        "",
        "## At a glance",
        "",
        "We built an automated pipeline that takes a list of coordinates and "
        "determines whether each location is a **tower site** or **rooftop "
        "cellular site**, whether **cellular equipment is visible**, and where "
        "the asset sits relative to the recorded point.",
        "",
        f"- **Assets evaluated:** {total}",
        f"- **Successfully classified:** {len(classified)}",
        f"- **Tower sites identified:** {towers}",
        f"- **Rooftop sites identified:** {rooftops}",
        f"- **Cellular equipment detected:** {cell_hits}",
        f"- **Assets with a located position:** {len(located)}",
        f"- **Errors:** {len(errors)}",
        "",
        "## Pilot run highlights",
        "",
        "This proof-of-concept run combined **free public NAIP imagery**, "
        "**Nearmap high-resolution + 45° oblique views**, and **Anthropic Claude "
        "vision AI** to classify six sample assets (2 urban NJ, 4 rural W/C).",
        "",
        "**What worked well:**",
        "",
        "- **Rooftop detection (urban):** Nearmap obliques revealed rooftop "
        "antenna sectors on a classical building (asset_001, 95% confidence)",
        "- **Tower detection (rural):** Monopoles identified from shadow "
        "signatures even when 60 m from the recorded coordinate (asset_003)",
        "- **Disguised towers:** A monopine (tower disguised as a pine tree) "
        "was correctly identified only from oblique imagery (asset_004)",
        "- **Off-center assets:** The pipeline searches the full image, not "
        "just the center — critical when coordinates are imprecise",
        "",
        "**Improvement in progress:**",
        "",
        "- asset_005 (rural Oregon) was missed in wide imagery but a human "
        "reviewer confirmed a lattice tower in the top-right of the NAIP chip. "
        "A **two-stage zoom** pass (now implemented) magnifies suspicious "
        "regions before re-classifying — designed specifically for this case.",
        "",
        "## How it works",
        "",
        "```",
        "Coordinates (CSV)",
        "       |",
        "       v",
        "  +----+----+",
        "  |  NAIP   |  Wide public aerial (~250 m, ~1 m resolution)",
        "  +----+----+",
        "       |",
        "       v",
        "  +----+----+",
        "  | Nearmap |  High-res top-down + 45-degree obliques (urban/suburban)",
        "  +----+----+",
        "       |",
        "       v",
        "  +----+----+",
        "  | Claude  |  AI vision: classify site, locate asset, detect equipment",
        "  +----+----+",
        "       |",
        "       v (if rural / still unidentified)",
        "  +----+----+",
        "  |  Zoom   |  Magnify suspicious regions and re-classify",
        "  +----+----+",
        "       |",
        "       v",
        "  results.csv + review chips + this summary",
        "```",
        "",
        "### Imagery sources",
        "",
        "| Source | What it provides | Why it matters |",
        "|---|---|---|",
        "| **NAIP** (free, public) | Wide top-down context around each point | "
        "Catches off-center towers; cheap baseline for the full US |",
        "| **Nearmap** (subscription) | ~7 cm top-down + 45° oblique views | "
        "Makes rooftop antennas and disguised towers (e.g. monopines) visible |",
        "| **Claude** (Anthropic) | Structured classification from multi-image input | "
        "Turns imagery into site type, equipment call, and location |",
        "",
        "### Confidence safeguards",
        "",
        "1. **Whole-image search** — never assumes the asset is at the exact center",
        "2. **Multi-view fusion** — NAIP context + Nearmap detail + oblique angles",
        "3. **Rural fallback** — widens Nearmap area when only vertical imagery exists",
        "4. **Two-stage zoom** — magnifies subtle structures the wide view missed",
        "5. **Human review chips** — every image sent to the model is saved for audit",
        "",
        "## Results by asset",
        "",
        "| Asset | Region | Site type | Confidence | Cell equip. | Located | Method | Key finding |",
        "|---|---|---|---:|---|---|---|---|",
    ]

    label_map = {k: str(v).strip()
                 for k, v in zip(assets_df["id"], assets_df.get("label", pd.Series(dtype=str)))}
    for r in results:
        aid = r.get("id", "")
        region = label_map.get(aid, str(r.get("label", "")).strip())
        err = _row_error(r)
        if err:
            short = err if len(err) <= 60 else err[:60] + "…"
            row = f"| {aid} | {region} | **ERROR** | — | — | — | — | {short} |"
        elif r.get("site_type") == "no_imagery":
            row = f"| {aid} | {region} | No imagery | — | — | — | — | Outside coverage |"
        else:
            conf = r.get("site_confidence")
            conf_s = f"{conf:.0%}" if isinstance(conf, (int, float)) else "—"
            cell = r.get("cell_equipment")
            cell_s = {True: "Yes", False: "No", None: "Unknown"}.get(cell, "—")
            if r.get("asset_lat") is not None:
                off = r.get("asset_offset_m")
                if off is not None and not (isinstance(off, float) and pd.isna(off)):
                    loc = f"{off:.0f} m off"
                else:
                    loc = "yes"
            elif r.get("asset_view"):
                loc = f"on {r['asset_view']}"
            else:
                loc = "—"
            stage = r.get("classification_stage") or "primary"
            if r.get("nearmap_aoi_m") == NEARMAP_FALLBACK_CHIP_M:
                stage = "wide AOI"
            evidence = (r.get("site_evidence") or "")[:90]
            if len(r.get("site_evidence") or "") > 90:
                evidence += "…"
            row = (f"| {aid} | {region} | {r.get('site_type', '—')} | {conf_s} | "
                   f"{cell_s} | {loc} | {stage} | {evidence} |")
        lines.append(row)

    lines.extend([
        "",
        "## Operational notes",
        "",
        f"- **Nearmap data usage (pilot run):** ~15 MB for {total} assets "
        "(under 1% of the 2.49 GB/month subscription allowance)",
        "- **Review folder:** saved images in `chips/` — NAIP chips named "
        "`*_NAIP.jpg`, Nearmap views `*_nearmap_*.jpg`, zoom crops `*_zoom_*.jpg`",
        "- **Machine-readable output:** `results.csv` for downstream systems",
        "",
        "## Known limitations",
        "",
        "- Rural sites may have Nearmap vertical imagery only (no 45° obliques)",
        "- Very small lattice towers can still be missed until the zoom stage runs",
        "- Recorded coordinates can be tens of meters off the true asset",
        "- AI calls should be spot-checked on low-confidence results (< 60%)",
        "",
        "## Recommended next steps",
        "",
        "1. Spot-check chips for any low-confidence or unexpected classifications",
        "2. Scale to the full asset list once stakeholders approve the approach",
        "3. Feed confirmed results back into the coordinate/enrichment workflow",
        "",
    ])

    Path(EXECUTIVE_SUMMARY_MD).write_text("\n".join(lines), encoding="utf-8")


def setup_run_directory(prefix: str, run_dir: str | None) -> Path:
    """Create or reopen a timestamped run folder for all outputs."""
    if run_dir:
        run_root = Path(run_dir)
        if not run_root.is_dir():
            raise SystemExit(f"Run directory not found: {run_root}")
        _out(f"Resuming run folder: {run_root}")
    else:
        stamp = time.strftime("%Y-%m-%d_%H%M%S")
        run_root = RUNS_DIR / f"{stamp}_{prefix}"
        run_root.mkdir(parents=True, exist_ok=True)
        _out(f"Created run folder: {run_root}")
    (run_root / "chips").mkdir(exist_ok=True)
    return run_root


def _print_run_banner(total: int, pending: int, skipped: int, input_csv: str,
                      run_dir: Path, output_csv: str, report_csv: str | None):
    """Startup summary so the operator can monitor the run in the terminal."""
    if QUIET:
        return
    est_min = max(1, round(pending * 0.35))  # ~12s pacing + API work per site
    print("\n" + "=" * 60, flush=True)
    print("  SITE CLASSIFIER RUN", flush=True)
    print("=" * 60, flush=True)
    print(f"  Run folder: {run_dir}", flush=True)
    print(f"  Input:      {input_csv}", flush=True)
    print(f"  Detail CSV: {output_csv}", flush=True)
    if report_csv:
        print(f"  Report:     {report_csv} (+ Excel with photos)", flush=True)
    print(f"  Chips:      {run_dir / 'chips'}", flush=True)
    print(f"  Total:      {total} assets | {skipped} already done | {pending} to run",
          flush=True)
    print(f"  Est. time:  ~{est_min} min for {pending} remaining", flush=True)
    print("=" * 60 + "\n", flush=True)


def _print_asset_start(idx: int, total: int, asset_id: str, row):
    """Mark the start of each asset in the terminal log."""
    addr = _clean_address(row)
    coord = (f"{row['lat']}, {row['lon']}" if _has_coordinates(row)
             else (addr[:55] + "…" if addr and len(addr) > 55 else addr))
    if QUIET:
        label = addr[:64] if addr else (coord or asset_id)
        _out(f"  [{idx}/{total}] {label}", important=True)
        return
    print(f"\n>>> [{idx}/{total}] {asset_id} — {coord}", flush=True)


def _print_asset_result(record: dict):
    """One-line success/failure summary after each asset."""
    if QUIET:
        return
    err = _row_error(record)
    if err:
        print(f"    RESULT: ERROR — {err[:100]}", flush=True)
        return
    site = record.get("site_type", "—")
    conf = record.get("site_confidence")
    conf_s = f"{conf:.2f}" if isinstance(conf, (int, float)) else "—"
    cell = record.get("cell_equipment")
    cell_s = {True: "yes", False: "no", None: "?"}.get(cell, str(cell))
    located = _format_located(record)
    views = record.get("view_count", 0)
    photo = "photo saved" if record.get("review_image") else "no photo"
    print(f"    RESULT: {site} | conf {conf_s} | cell {cell_s} | "
          f"{located} | {views} views | {photo}", flush=True)


def _print_run_complete(results: list[dict], run_dir: Path, output_csv: str,
                        report_csv: str | None, report_xlsx: str | None,
                        summary_md: str):
    """Final terminal summary when the batch finishes."""
    if QUIET:
        return
    errors = sum(1 for r in results if _row_error(r))
    ok = len(results) - errors
    towers = sum(1 for r in results if r.get("site_type") == "tower")
    rooftops = sum(1 for r in results if r.get("site_type") == "rooftop")
    cell_hits = sum(1 for r in results if r.get("cell_equipment") is True)
    print("\n" + "=" * 60, flush=True)
    print("  RUN COMPLETE", flush=True)
    print("=" * 60, flush=True)
    print(f"  Folder:   {run_dir}", flush=True)
    print(f"  Classified: {ok}/{len(results)}  |  Errors: {errors}", flush=True)
    print(f"  Towers: {towers}  |  Rooftops: {rooftops}  |  Cell equip: {cell_hits}",
          flush=True)
    print(f"  Detail:   {output_csv}", flush=True)
    if report_csv:
        print(f"  Report:   {report_csv}", flush=True)
        print(f"  Excel:    {report_xlsx}  (includes embedded photos)", flush=True)
    print(f"  Summary:  {summary_md}", flush=True)
    print("=" * 60 + "\n", flush=True)


def regenerate_reports_from_detail(run_root: Path, output_csv: str,
                                   report_csv: str | None, report_xlsx: str | None,
                                   summary_md: str, input_csv: str):
    """Fix confidence values and rebuild report files without API calls."""
    if not os.path.exists(output_csv):
        raise SystemExit(f"No detail file found: {output_csv}")

    results = pd.read_csv(output_csv).to_dict("records")
    fixed = 0
    for record in results:
        before = (record.get("site_confidence"), record.get("cell_equipment_confidence"))
        normalize_model_result(record)
        after = (record.get("site_confidence"), record.get("cell_equipment_confidence"))
        if before != after:
            fixed += 1

    pd.DataFrame(results).to_csv(output_csv, index=False)

    input_path = run_root / Path(input_csv).name
    assets_df = (pd.read_csv(input_path) if input_path.exists()
                 else pd.read_csv(input_csv))

    write_executive_summary(results, assets_df)
    if report_csv and report_xlsx:
        write_stakeholder_report(results, report_csv, report_xlsx)

    print(f"\nRegenerated reports in {run_root}", flush=True)
    print(f"  Fixed confidence on {fixed} row(s)", flush=True)
    print(f"  Detail:  {output_csv}", flush=True)
    if report_csv:
        print(f"  Report:  {report_csv}", flush=True)
        print(f"  Excel:   {report_xlsx}", flush=True)
    print(f"  Summary: {summary_md}", flush=True)

def classify_with_routing(provider: str, clients: dict, views: list,
                          prompt: str, input_confidence: str,
                          *, escalate: bool = True, screen: bool = False
                          ) -> tuple[dict, str, str | None, str | None]:
    """Run primary classification; optionally escalate Gemini -> Claude.

    Pass escalate=False for intermediate imagery stages so Gemini is fully
    exhausted (NAIP/Nearmap/zoom) before a single final Claude call.
    ``screen=True`` uses GEMINI_SCREEN_MODEL then Flash-confirm when needed.
    """
    primary_model = provider
    res = _classify_pass(
        provider, clients, views, prompt, input_confidence, screen=screen
    )

    escalation_model = None
    escalation_reason_str = None
    if (escalate and BIFURCATED_AI and not GEMINI_ONLY
            and provider == "gemini"):
        res, escalation_model, escalation_reason_str = maybe_escalate_to_claude(
            res, clients, views, prompt, input_confidence, allow=True
        )
    return res, primary_model, escalation_model, escalation_reason_str


def cheap_second_opinion_disagrees(
    res: dict, clients: dict, views: list, prompt: str
) -> bool:
    """True when Flash (or a second Gemini pass) calls a positive vs other/unclear.

    If the last call was already GEMINI_MODEL, type alone is not disagreement.
    """
    site = str(res.get("site_type") or "").strip().lower()
    if site not in {"other", "unclear"}:
        return False
    if confident_no_asset(res):
        return False
    if "gemini" not in clients or clients.get("gemini") is None:
        return False
    last_model = str(res.get("model") or "").strip()
    if last_model == GEMINI_MODEL:
        return False
    second = classify_site(
        "gemini", clients, views, prompt=prompt, gemini_model=GEMINI_MODEL
    )
    return str(second.get("site_type") or "").strip().lower() in _positive_site_types()


def should_attempt_claude_escalation(res: dict) -> bool:
    """True when Claude (or a Gemini second opinion) is worth spending.

    Low-confidence NAIP ``unclear`` and NAIP-only rooftops are holdouts, not
    Claude jobs. Full Nearmap other/unclear with no cell also skips the
    cheap second-opinion path — Claude does not recover those.
    """
    if escalation_reason(res):
        return True
    if _nearmap_empty_without_cell(res):
        return False
    site = str(res.get("site_type") or "").strip().lower()
    if site not in {"other", "unclear"} or confident_no_asset(res):
        return False
    if site == "unclear":
        conf = normalize_confidence(res.get("site_confidence"))
        return conf is not None and conf >= TIER_CONF_MEDIUM
    return True


def maybe_escalate_to_claude(res: dict, clients: dict, views: list, prompt: str,
                             input_confidence: str,
                             *, allow: bool
                             ) -> tuple[dict, str | None, str | None]:
    """Single late Claude escalation after all Gemini imagery stages."""
    if not allow:
        return res, None, None
    if not should_attempt_claude_escalation(res):
        return res, None, None
    reason = escalation_reason(res)
    if not reason:
        if cheap_second_opinion_disagrees(res, clients, views, prompt):
            reason = "second_opinion_disagree"
        else:
            return res, None, None
    escalated = classify_site(
        "claude", clients, views, prompt=prompt,
        claude_model=CLAUDE_ESCALATION_MODEL)
    escalated = maybe_recheck_equipment(
        "claude", clients, escalated, views, input_confidence)
    _step_done("escalate (Claude)", _brief_pass_result(escalated))
    return escalated, "claude", reason


def _effective_provider(primary_model: str, escalation_model: str | None) -> str:
    return escalation_model or primary_model


def _canonical_to_classify_row(canonical: dict) -> dict:
    """Build one classifier input row from an orchestrator canonical record."""
    site_id = canonical.get("id") or (
        f"site_{abs(hash(canonical.get('address', ''))) % 10**8:08d}"
    )
    lon = canonical.get("lng") if canonical.get("lng") is not None else canonical.get("lon")
    meta = canonical.get("permit_metadata") or {}
    return {
        "id": site_id,
        "lat": canonical.get("lat"),
        "lon": lon,
        "address": canonical.get("address", ""),
        "label": canonical.get("label") or meta.get("label") or "",
        "input_confidence": (
            canonical.get("input_confidence")
            or meta.get("input_confidence")
            or ""
        ),
    }


def _detail_row_to_result(row: dict, canonical: dict | None = None) -> dict:
    """Normalize a results_detail row for orchestrator (no pandas NaN surprises)."""
    result = {}
    for key, value in row.items():
        if isinstance(value, float) and value != value:  # NaN
            result[key] = None
        elif isinstance(value, str) and value.strip().lower() == "nan":
            result[key] = None
        else:
            result[key] = value
    canonical = canonical or {}
    result["permit_metadata"] = canonical.get("permit_metadata", {})
    if canonical.get("source_url"):
        result["source_url"] = canonical["source_url"]
    return result


def classify_records(
    canonicals: list[dict],
    run_dir: Path | str | None = None,
    *,
    quiet: bool = True,
) -> list[dict]:
    """Classify many orchestrator records in one classifier run (one shared input CSV)."""
    if not canonicals:
        return []

    run_path = Path(run_dir) if run_dir else (RUNS_DIR / "orchestrator")
    run_path.mkdir(parents=True, exist_ok=True)
    (run_path / "chips").mkdir(exist_ok=True)

    input_rows = [_canonical_to_classify_row(c) for c in canonicals]
    input_csv = run_path / "classify_input.csv"
    pd.DataFrame(input_rows).to_csv(input_csv, index=False)

    argv = [
        "asset_classifier",
        "--input", str(input_csv),
        "--run-dir", str(run_path),
    ]
    if quiet:
        argv.append("--quiet")
    prior_argv = sys.argv
    try:
        sys.argv = argv
        main()
    finally:
        sys.argv = prior_argv

    detail_path = run_path / "results_detail.csv"
    if not detail_path.exists():
        raise RuntimeError(f"Classifier did not produce detail output in {run_path}")

    detail = pd.read_csv(detail_path)
    results: list[dict] = []
    for input_row, canonical in zip(input_rows, canonicals):
        site_id = str(input_row["id"])
        if "id" in detail.columns:
            matched = detail[detail["id"].astype(str) == site_id]
            raw = (matched.iloc[-1] if not matched.empty else detail.iloc[-1]).to_dict()
        else:
            raw = detail.iloc[-1].to_dict()
        results.append(_detail_row_to_result(raw, canonical))
    return results


def classify_record(
    canonical: dict,
    run_dir: Path | str | None = None,
    *,
    quiet: bool = True,
) -> dict:
    """Classify one orchestrator canonical record (thin wrapper over classify_records)."""
    return classify_records([canonical], run_dir=run_dir, quiet=quiet)[0]


def main():
    global INPUT_CSV, OUTPUT_CSV, EXECUTIVE_SUMMARY_MD, CHIP_DIR, RUN_DIR, QUIET

    # Gemini SDK / httpx INFO lines (AFC banner, "HTTP Request: POST …") drown
    # useful classify progress; keep warnings+.
    for _noisy in ("google_genai", "google_genai.models", "httpx", "httpcore"):
        logging.getLogger(_noisy).setLevel(logging.WARNING)

    parser = argparse.ArgumentParser(description="Classify cell sites from aerial imagery")
    parser.add_argument("--input", "-i", default=INPUT_CSV,
                        help=f"Input CSV (default: {INPUT_CSV})")
    parser.add_argument("--run-dir", default=None,
                        help="Resume an existing timestamped run folder under runs/")
    parser.add_argument("--output", "-o", default=None,
                        help="Detail output CSV filename inside the run folder")
    parser.add_argument("--report-csv", default=None,
                        help="Stakeholder summary CSV filename inside the run folder")
    parser.add_argument("--report-xlsx", default=None,
                        help="Stakeholder Excel with photos inside the run folder")
    parser.add_argument("--regenerate-report", action="store_true",
                        help="Fix confidence values and rebuild reports from detail CSV")
    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Suppress banners/tqdm/API chatter; keep per-step '— done' lines",
    )
    args = parser.parse_args()
    QUIET = bool(args.quiet)

    if args.regenerate_report and not args.run_dir:
        raise SystemExit("--regenerate-report requires --run-dir pointing at an "
                         "existing run folder (e.g. runs\\2026-06-12_184824_WI)")

    INPUT_CSV = args.input
    stem = Path(INPUT_CSV).stem
    prefix = stem.replace("_assets", "") if stem.endswith("_assets") else stem

    run_root = setup_run_directory(prefix, args.run_dir)
    RUN_DIR = run_root
    CHIP_DIR = run_root / "chips"

    detail_name = (Path(args.output).name if args.output
                   else f"{prefix}_results_detail.csv"
                   if stem.endswith("_assets") else "results_detail.csv")
    OUTPUT_CSV = str(run_root / detail_name)

    report_csv = None
    report_xlsx = None
    if stem.endswith("_assets") or args.report_csv or args.report_xlsx:
        report_csv = str(run_root / (Path(args.report_csv).name if args.report_csv
                                       else f"{prefix}_results.csv"))
        report_xlsx = str(run_root / (Path(args.report_xlsx).name if args.report_xlsx
                                        else f"{prefix}_results.xlsx"))

    EXECUTIVE_SUMMARY_MD = str(run_root / (
        f"{prefix}_EXECUTIVE_SUMMARY.md"
        if stem.endswith("_assets") else "EXECUTIVE_SUMMARY.md"))

    if not args.run_dir and not args.regenerate_report:
        shutil.copy2(INPUT_CSV, run_root / Path(INPUT_CSV).name)

    if args.regenerate_report:
        regenerate_reports_from_detail(
            run_root, OUTPUT_CSV, report_csv, report_xlsx,
            EXECUTIVE_SUMMARY_MD, INPUT_CSV)
        return

    primary_provider, allow_claude_escalation = resolve_ai_mode()

    if primary_provider == "gemini" and not os.environ.get("GEMINI_API_KEY"):
        raise SystemExit(
            "GEMINI_API_KEY is not set. Get a key at "
            "https://aistudio.google.com/apikey"
        )
    if (primary_provider == "claude" or allow_claude_escalation) and not os.environ.get(
        "ANTHROPIC_API_KEY"
    ):
        raise SystemExit(
            "ANTHROPIC_API_KEY is not set. Get a key at https://console.anthropic.com/\n"
            '  PowerShell: $env:ANTHROPIC_API_KEY="sk-ant-..."\n'
            "  bash/zsh:   export ANTHROPIC_API_KEY=sk-ant-..."
        )

    clients: dict[str, object] = {}
    if primary_provider == "gemini" or allow_claude_escalation:
        clients["gemini"] = genai.Client(api_key=os.environ.get("GEMINI_API_KEY"))
    if primary_provider == "claude" or allow_claude_escalation:
        clients["claude"] = Anthropic()

    if NAIP_ONLY:
        _out("[BREAKPOINT] NAIP_ONLY=1 — Nearmap fetch disabled. "
              "Running on NAIP imagery only.")
        _out(f"  primary NAIP chip: {int(CHIP_SIZE_M)}m")
    if TOWER_ONLY:
        _out("TOWER_ONLY=1 — tower detection mode (rooftop hosts -> other).")
    if not ZOOM_STAGE:
        _out("ZOOM_STAGE=0 — two-stage zoom disabled (single-pass classification).")
    if WIDE_AOI_STAGE:
        _out(f"WIDE_AOI_STAGE=1 — widen NAIP to {int(NAIP_WIDE_CHIP_M)}m "
              "and re-classify on other/unclear.")
    else:
        _out("WIDE_AOI_STAGE=0 — no NAIP zoom-out retry.")
    if GEMINI_ONLY:
        _out(f"GEMINI_ONLY=1 — Gemini only ({GEMINI_MODEL}), no Claude escalation.")
    elif allow_claude_escalation:
        _out(f"BIFURCATED_AI=1 — Gemini first ({GEMINI_MODEL}), "
              f"Claude escalation ({CLAUDE_ESCALATION_MODEL})")
        _out(
            f"  Gemini thinking_budget={_gemini_thinking_budget()} · "
            f"solo dual-skip conf>={GEMINI_SOLO_CELL_CONF:g}"
        )
    elif primary_provider == "claude":
        _out(f"Claude only ({MODELS[0]})")
    if NEARMAP_TIERED and not NAIP_ONLY:
        _out("NEARMAP_TIERED=1 — tiered Nearmap fetch enabled "
              "(NAIP -> Vert -> obliques)")
        _out(f"  rooftop on NAIP age > {NAIP_MAX_AGE_YEARS:g}y requires Nearmap "
              f"unless conf >= {NAIP_AGE_HIGH_CONF_OVERRIDE:g} "
              f"(NAIP_MAX_AGE_YEARS / NAIP_AGE_HIGH_CONF_OVERRIDE)")

    _out(f"Input:  {INPUT_CSV}")
    _out(f"Output: {OUTPUT_CSV} (detail/resume)")
    if report_csv:
        _out(f"Report: {report_csv} + {report_xlsx}")

    df = pd.read_csv(INPUT_CSV)
    validate_input_csv(df)

    # Resume support: keep successfully classified rows from a previous run and
    # skip them, so quota is only spent on assets that still need work
    results = []
    done_ids = set()
    if os.path.exists(OUTPUT_CSV):
        prev = pd.read_csv(OUTPUT_CSV)
        if "site_type" in prev.columns:
            mask = prev["site_type"].notna()
            if "error" in prev.columns:
                mask &= prev["error"].isna()
            kept = prev[mask]
            results = kept.to_dict("records")
            done_ids = set(kept["id"])
            if done_ids:
                _out(f"Resuming: {len(done_ids)} assets already done, "
                      f"{len(df) - len(done_ids)} remaining")

    pending_rows = [row for _, row in df.iterrows() if row["id"] not in done_ids]
    _print_run_banner(len(df), len(pending_rows), len(done_ids),
                      INPUT_CSV, run_root, OUTPUT_CSV, report_csv)

    progress = tqdm(
        pending_rows,
        desc="Progress",
        unit="site",
        dynamic_ncols=True,
        disable=QUIET,
    )
    for row in progress:
        progress.set_postfix_str(str(row["id"]), refresh=False)
        _print_asset_start(len(done_ids) + progress.n + 1, len(df), row["id"], row)
        # Carry all input columns (id, lat, lon, address, label, ...) into results
        record = row.to_dict()
        try:
            lat, lon, geocode_meta = resolve_row_coordinates(row)
            record["lat"] = lat
            record["lon"] = lon
            record.update(geocode_meta)
            if geocode_meta:
                _out(f"    geocoded ({geocode_meta.get('geocode_source')}): "
                      f"{lat:.6f}, {lon:.6f}")

            img, naip_meta, naip_geo = fetch_chip(lat, lon)
            img_date = (naip_meta or {}).get("image_date")
            naip_chip_m = (naip_meta or {}).get("naip_chip_m") or CHIP_SIZE_M

            nearmap_views, nearmap_date = {}, None
            if not NAIP_ONLY:
                try:
                    if not NEARMAP_TIERED:
                        nearmap_views, nearmap_date = fetch_nearmap_views(lat, lon)
                except Exception as e:
                    _out(f"  [{row['id']}] nearmap fetch failed: {e}", important=True)

            if img is None and not nearmap_views:
                record["site_type"] = "no_imagery"
                if naip_meta:
                    record.update({k: v for k, v in naip_meta.items() if k != "naip_chip_m"})
                results.append(record)
                _print_asset_result(record)
                pd.DataFrame(results).to_csv(OUTPUT_CSV, index=False)
                time.sleep(GEMINI_DELAY_S if primary_provider == "gemini" else API_DELAY_S)
                continue

            def build_views(nm_views, naip_img=None, chip_m=None):
                """Save Nearmap chips (overwriting any narrow-AOI versions)
                and assemble the labeled view list for the model."""
                v = []
                source_img = img if naip_img is None else naip_img
                side_m = naip_chip_m if chip_m is None else chip_m
                if source_img is not None:
                    v.append((_naip_view_label(side_m), source_img))
                for name, vimg in nm_views.items():
                    vpath = CHIP_DIR / f"{row['id']}_nearmap_{name.lower()}.jpg"
                    vimg.save(vpath, quality=90)
                    label = ("Nearmap top-down" if name == "Vert"
                             else f"Nearmap oblique ({name})")
                    v.append((label, vimg))
                return v

            chip_path = None
            if img is not None:
                chip_path = CHIP_DIR / f"{row['id']}_NAIP.jpg"
                img.save(chip_path, quality=90)
                if img_date:
                    _out(f"    NAIP acquired {img_date}"
                          f" (age {naip_meta.get('image_age_years')}y,"
                          f" gsd={naip_meta.get('naip_gsd_m')}m,"
                          f" chip={int(naip_chip_m)}m)")

            label_hint = str(row.get("label", "")).strip().lower()
            input_confidence = normalize_input_confidence(row.get("input_confidence"))
            prompt = build_classification_prompt(row)

            primary_model = primary_provider
            escalation_model = None
            escalation_reason_str = None

            if NAIP_ONLY:
                views = build_views({})
                _step_done("NAIP")
                nearmap_tier = "naip_only"
                res, primary_model, escalation_model, escalation_reason_str = (
                    classify_with_routing(
                        primary_provider, clients, views, prompt,
                        input_confidence, escalate=False, screen=True))
                _step_done("classify (NAIP)", _brief_pass_result(res))
            elif NEARMAP_TIERED:
                # Gemini exhausts NAIP -> Vert -> obliques before any Claude.
                res, nearmap_views, nearmap_date, nearmap_tier, views = (
                    classify_with_tiers(
                        lat, lon, img, primary_provider, clients, prompt,
                        input_confidence, build_views,
                        naip_age_years=(naip_meta or {}).get("image_age_years")))
                primary_model = primary_provider
            else:
                nearmap_tier = "full" if nearmap_views else "naip_only"
                views = build_views(nearmap_views)
                _step_done("imagery ready")
                res, primary_model, escalation_model, escalation_reason_str = (
                    classify_with_routing(
                        primary_provider, clients, views, prompt,
                        input_confidence, escalate=False))
                _step_done("classify", _brief_pass_result(res))

            nearmap_aoi_m = NEARMAP_CHIP_M if nearmap_views else None
            classification_stage = "primary"
            zoom_count = 0
            # Keep Gemini through wide/zoom retries; Claude only at the end.
            stage_provider = primary_provider

            # Wide-AOI fallback: pin may sit in a parking lot while the mall
            # rooftop / tower is 100–250m away. Scout even when pin-centered
            # Nearmap already had obliques, unless we locked an HVAC-only roof.
            has_obliques = any(n != "Vert" for n in nearmap_views)
            nearmap_blocks_rescue = nearmap_full_blocks_rescue(
                res, nearmap_tier=nearmap_tier, has_obliques=has_obliques
            )
            pin_offset_scout = (
                not nearmap_blocks_rescue and needs_pin_offset_scout(res)
            )
            naip_rescue = (
                not nearmap_blocks_rescue
                and not pin_offset_scout
                and needs_naip_rescue(res)
            )
            if confident_no_asset(res):
                _step_done("rescue skipped", "confident NAIP other")
            if nearmap_blocks_rescue:
                _step_done(
                    "Nearmap settled",
                    "locked HVAC rooftop — skip wide/zoom rescue",
                )
            if pin_offset_scout and not NAIP_ONLY and NEARMAP_API_KEY:
                try:
                    wide_views, wide_date = fetch_nearmap_views(
                        lat, lon, NEARMAP_FALLBACK_CHIP_M)
                except Exception as e:
                    wide_views, wide_date = {}, None
                    _out(f"  [{row['id']}] wide nearmap fetch failed: {e}", important=True)
                if wide_views:
                    _step_done(f"Nearmap wide AOI ({NEARMAP_FALLBACK_CHIP_M}m)")
                    nearmap_views = wide_views
                    nearmap_date = wide_date or nearmap_date
                    nearmap_aoi_m = NEARMAP_FALLBACK_CHIP_M
                    views = build_views(wide_views)
                    res, _, _, _ = classify_with_routing(
                        stage_provider, clients, views, prompt,
                        input_confidence, escalate=False)
                    _step_done("classify (wide AOI)", _brief_pass_result(res))
                    classification_stage = "wide_aoi"
                    nearmap_tier = "wide_aoi"

            # NAIP zoom-out: only when Nearmap never ran (coverage miss).
            # Rooftop pin-offset already bought a wide Nearmap AOI.
            wide_was_other = False
            if (naip_rescue and WIDE_AOI_STAGE and img is not None
                    and not confident_no_asset(res)
                    and res.get("site_type") in ("other", "unclear")
                    and res.get("cell_equipment") is not True
                    and NAIP_WIDE_CHIP_M > CHIP_SIZE_M):
                try:
                    wide_img, wide_meta, wide_geo = fetch_chip(
                        lat, lon, chip_m=NAIP_WIDE_CHIP_M)
                except Exception as e:
                    wide_img, wide_meta, wide_geo = None, None, None
                    _out(f"  [{row['id']}] wide NAIP fetch failed: {e}", important=True)
                if wide_img is not None:
                    _step_done(f"NAIP wide ({int(NAIP_WIDE_CHIP_M)}m)")
                    wide_path = CHIP_DIR / f"{row['id']}_NAIP_wide.jpg"
                    wide_img.save(wide_path, quality=90)
                    views = build_views(
                        nearmap_views, naip_img=wide_img, chip_m=NAIP_WIDE_CHIP_M)
                    wide_res, _, _, _ = classify_with_routing(
                        stage_provider, clients, views, prompt,
                        input_confidence, escalate=False)
                    _step_done("classify (NAIP wide)", _brief_pass_result(wide_res))
                    wide_wins = scout_result_wins(res, wide_res)
                    if wide_wins:
                        res = wide_res
                        img = wide_img
                        naip_geo = wide_geo
                        naip_chip_m = NAIP_WIDE_CHIP_M
                        if wide_meta:
                            naip_meta = {**(naip_meta or {}), **wide_meta}
                            img_date = wide_meta.get("image_date") or img_date
                        chip_path = wide_path
                        classification_stage = "wide_aoi"
                        nearmap_tier = "naip_wide"
                    if str(wide_res.get("site_type") or "").strip().lower() == "other":
                        wide_was_other = True

            # Two-stage zoom: scout suspicious regions, magnify, re-classify.
            force_zoom = label_hint == "stealth"
            zoom_types = (
                ("other", "unclear", "rooftop")
                if pin_offset_scout
                else ("other", "unclear")
            )
            if ((pin_offset_scout or naip_rescue or force_zoom) and ZOOM_STAGE
                    and not confident_no_asset(res)
                    and not wide_was_other
                    and (
                        force_zoom
                        or (
                            res.get("site_type") in zoom_types
                            and res.get("cell_equipment") is not True
                        )
                    )):
                source_label, source_img = None, None
                if nearmap_views.get("Vert"):
                    source_label, source_img = "Nearmap top-down", nearmap_views["Vert"]
                elif img is not None:
                    source_label, source_img = "NAIP top-down", img
                if source_img is not None:
                    zoom_res, zoom_count = run_zoom_stage(
                        stage_provider, clients, row["id"], views,
                        source_label, source_img,
                        max_crops=3 if force_zoom else ZOOM_MAX_CANDIDATES)
                    _step_done("zoom crops", _brief_pass_result(zoom_res))
                    zoom_wins = (
                        scout_result_wins(res, zoom_res)
                        or (force_zoom and zoom_res.get("cell_equipment") is True)
                        or force_zoom
                    )
                    if zoom_wins:
                        res = zoom_res
                        classification_stage = "zoom"
                        nearmap_tier = "zoom"

            # Claude only after Gemini has exhausted imagery stages.
            if allow_claude_escalation and should_attempt_claude_escalation(res):
                res, escalation_model, escalation_reason_str = (
                    maybe_escalate_to_claude(
                        res, clients, views, prompt, input_confidence,
                        allow=True))

            fp_provider = (
                "claude" if escalation_model else primary_provider
            )
            res = gate_weak_rooftop_cell_claim(res)
            res = gate_weak_stealth_tower_claim(res)
            res = maybe_repair_rooftop_asset_box(fp_provider, clients, res, views)
            res = gate_weak_rooftop_cell_claim(res)
            res = gate_weak_stealth_tower_claim(res)
            res = enforce_rooftop_cell_requires_box(res, views)
            confirm_views, used_crop = build_cell_confirm_views(res, views)
            res, dual_model, cell_agree = confirm_rooftop_cell_with_claude(
                res,
                clients,
                confirm_views,
                already_escalated=bool(escalation_model),
                allow_soft_keep=True,
                from_wide_rescue=classification_stage
                in {"wide_aoi", "zoom", "pin_recenter_rejected"}
                or nearmap_tier in {"naip_wide", "wide_aoi"},
                used_crop=used_crop,
                all_views=views,
            )
            if dual_model and not escalation_model:
                escalation_model = dual_model
                escalation_reason_str = escalation_reason_str or (
                    "gemini_high_conf_tower"
                    if dual_model == "gemini_strong_solo"
                    else "rooftop_dual_model_cell"
                )
            res["cell_models_agree"] = cell_agree
            if res.get("cell_equipment") is True:
                res = enforce_rooftop_cell_requires_box(res, views)
            res = align_site_evidence_with_cell(res)

            # Convert the detection box to real-world coordinates when possible
            # (NAIP geo, Nearmap Vert, or Nearmap oblique AOI approximation).
            asset_lat = asset_lon = asset_offset_m = None
            asset_coord_source = None
            box, box_view = res.get("asset_box_2d"), res.get("asset_view")
            if box:
                located = locate_asset_box_latlon(
                    lat=lat,
                    lon=lon,
                    box=box,
                    box_view=box_view,
                    naip_geo=naip_geo,
                    nearmap_aoi_m=nearmap_aoi_m,
                )
                if located:
                    asset_lat, asset_lon, asset_offset_m, asset_coord_source = located

            record.update({
                "image_date": img_date,
                "naip_year": (naip_meta or {}).get("naip_year"),
                "naip_state": (naip_meta or {}).get("naip_state"),
                "naip_gsd_m": (naip_meta or {}).get("naip_gsd_m"),
                "image_age_years": (naip_meta or {}).get("image_age_years"),
                "naip_chip_m": naip_chip_m,
                "nearmap_date": nearmap_date,
                "nearmap_views": ",".join(nearmap_views) or None,
                "nearmap_aoi_m": nearmap_aoi_m,
                "chip_path": str(chip_path) if chip_path else None,
                "view_count": len(views),
                "site_type": res.get("site_type"),
                "tower_subtype": res.get("tower_subtype"),
                "site_confidence": res.get("site_confidence"),
                "site_evidence": res.get("site_evidence"),
                "asset_lat": asset_lat,
                "asset_lon": asset_lon,
                "asset_offset_m": (round(asset_offset_m, 1)
                                   if asset_offset_m is not None else None),
                "asset_coord_source": asset_coord_source,
                "asset_box_2d": json.dumps(box) if box else None,
                "asset_view": box_view,
                "cell_equipment": res.get("cell_equipment"),
                "cell_equipment_confidence": res.get("cell_equipment_confidence"),
                "cell_equipment_evidence": res.get("cell_equipment_evidence"),
                "classification_stage": classification_stage,
                "zoom_crops": zoom_count or None,
                "input_confidence": input_confidence,
                "source_trust_mismatch": (
                    input_confidence == "high"
                    and res.get("cell_equipment") is False),
                "model": res.get("model"),
                "nearmap_tier": nearmap_tier,
                "primary_model": primary_model,
                "escalation_model": escalation_model,
                "escalation_reason": escalation_reason_str,
            })
            review_path = pick_review_image_path(row["id"], record)
            if review_path:
                record["review_image"] = str(review_path)
            loc = (f"({asset_lat:.6f},{asset_lon:.6f}, {asset_offset_m:.0f}m off)"
                   if asset_lat is not None else f"(box on: {box_view})")
            if not QUIET:
                print(f"    {record['site_type']} ({record['site_confidence']}) {loc} "
                      f"| cell: {record['cell_equipment']} | stage: {classification_stage}",
                      flush=True)
            _print_asset_result(record)

        except Exception as e:
            record["error"] = str(e)
            if not QUIET:
                print(f"    ERROR: {e}", flush=True)
            else:
                _out(f"         ERROR — {e}", important=True)
            _print_asset_result(record)

        results.append(record)
        # Rewrite after every row so a mid-run crash never loses completed work
        pd.DataFrame(results).to_csv(OUTPUT_CSV, index=False)
        inter_asset_delay = GEMINI_DELAY_S if primary_provider == "gemini" else API_DELAY_S
        time.sleep(inter_asset_delay)

    write_executive_summary(results, df)
    if report_csv and report_xlsx:
        write_stakeholder_report(results, report_csv, report_xlsx)
    _print_run_complete(results, run_root, OUTPUT_CSV, report_csv, report_xlsx,
                      EXECUTIVE_SUMMARY_MD)


if __name__ == "__main__":
    main()
